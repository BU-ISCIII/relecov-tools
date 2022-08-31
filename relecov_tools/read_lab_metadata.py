#!/usr/bin/env python
from itertools import islice

import json
import logging

# from queue import Empty

# from turtle import heading
import rich.console

import openpyxl
import os
import sys
import re
import relecov_tools.utils
from relecov_tools.config_json import ConfigJson
import relecov_tools.json_schema

log = logging.getLogger(__name__)
stderr = rich.console.Console(
    stderr=True,
    style="dim",
    highlight=False,
    force_terminal=relecov_tools.utils.rich_force_colors(),
)


class RelecovMetadata:
    def __init__(self, metadata_file=None, sample_list_file=None, output_folder=None):
        if metadata_file is None:
            self.metadata_file = relecov_tools.utils.prompt_path(
                msg="Select the excel file which contains metadata"
            )
        else:
            self.metadata_file = metadata_file
        if not os.path.exists(self.metadata_file):
            log.error("Metadata file %s does not exist ", self.metadata_file)
            stderr.print("[red] Metadata file " + self.meta_file + " does not exist")
            sys.exit(1)
        if sample_list_file is None:
            self.sample_list_file = relecov_tools.utils.prompt_path(
                msg="Select the file which contains the sample information"
            )
        else:
            self.sample_list_file = sample_list_file
        if not os.path.exists(self.sample_list_file):
            log.error("Sample information file %s does not exist ", self.metadata_file)
            stderr.print(
                "[red] Sample information " + self.sample_list_file + " does not exist"
            )
            sys.exit(1)
        if output_folder is None:
            self.output_folder = relecov_tools.utils.prompt_path(
                msg="Select the output folder"
            )
        else:
            self.output_folder = output_folder
        config_json = ConfigJson()
        relecov_schema = config_json.get_topic_data("json_schemas", "relecov_schema")
        relecov_sch_path = os.path.join(
            os.path.dirname(os.path.realpath(__file__)), "schema", relecov_schema
        )
        self.configuration = config_json

        with open(relecov_sch_path, "r") as fh:
            self.relecov_sch_json = json.load(fh)
        self.label_prop_dict = {}
        for prop, values in self.relecov_sch_json["properties"].items():
            try:
                self.label_prop_dict[values["label"]] = prop
            except KeyError:
                continue
        self.schema_name = self.relecov_sch_json["schema"]
        self.schema_version = self.relecov_sch_json["version"]

    def fetch_metadata_file(folder, file_name):
        """Fetch the metadata file folder  Directory to fetch metadata file
        file_name   metadata file name
        """
        wb_file = openpyxl.load_workbook(file_name, data_only=True)
        ws_metadata_lab = wb_file["METADATA_LAB"]
        heading = []
        for cell in ws_metadata_lab[1]:
            heading.append(cell.value)

    def read_json_file(self, j_file):
        """Read json file."""
        with open(j_file, "r") as fh:
            data = json.load(fh)
        return data

    def get_laboratory_data(self, lab_json, geo_loc_json, lab_name):
        """Fetch the laboratory location  and return a dictionary"""
        data = {}
        if lab_name == "":
            data["geo_loc_city"] = ""
            data["geo_loc_latitude"] = ""
            data["geo_loc_longitude"] = ""
            data["geo_loc_country"] = ""
            stderr.print("[red] Empty Originating Laboratory.")
            log.error("Found empty Originating Laboratory")
            return data
        for lab in lab_json:
            if lab_name == lab["collecting_institution"]:
                for key, value in lab.items():
                    data[key] = value
                break

        for city in geo_loc_json:
            try:
                if city["geo_loc_city"] == lab["geo_loc_city"]:
                    data["geo_loc_latitude"] = city["geo_loc_latitude"]
                    data["geo_loc_longitude"] = city["geo_loc_longitude"]
                    data["geo_loc_country"] = data["geo_loc_country"]
                    break
            except KeyError as e:
                print(e)
        return data

    def include_fixed_data(self):
        """Include fixed data that are always the same for each samples"""
        fixed_data = {
            "host_disease": "COVID-19",
            "type": "betacoronavirus",
            "tax_id": "2697049",
            "organism": "Severe acute respiratory syndrome coronavirus 2",
            "common_name": "Severe acute respiratory syndrome",
            "sample_description": "Sample for surveillance",
        }
        fixed_data.update(self.configuration.get_configuration("ENA_configuration"))
        fixed_data["schema_name"] = self.schema_name
        fixed_data["schema_version"] = self.schema_version
        return fixed_data

    def include_fields_already_set(self, row_sample):
        processed_data = {}

        if row_sample["author_submitter"] == "":
            processed_data["collector_name"] = "unknown"
        else:
            processed_data["collector_name"] = row_sample["author_submitter"]
        processed_data["host_subject_id"] = row_sample["microbiology_lab_sample_id"]

        return processed_data

    def include_processed_data(self, metadata):
        """Include the data that requires to be processed to set the value"""
        new_data = {}
        p_data = {
            "host_common_name": {"Human": ["host_scientific_name", "Homo Sapiens"]},
            "collecting_lab_sample_id": [
                "isolate_sample_id",
                metadata["sequencing_sample_id"],
            ],
        }
        seq_inst_plat = {
            "Illumina": [
                "Illumina iSeq 100",
                "Illumina MiSeq",
                "Illumina NextSeq 550",
                "Illumina NextSeq",
                "Illumina NovaSeq 6000",
            ],
            "Oxford Nanopore": ["MinION"],
            "Ion Torrent": ["Ion Torrent S5", "Ion Torrent PGM"],
        }
        for key, values in p_data.items():
            v_data = metadata[key]
            if isinstance(values, dict):
                if v_data in values:
                    new_data[values[v_data][0]] = values[v_data][1]
            else:
                new_data[values[0]] = values[1]
        """New fields that required processing from other field """
        for key, values in seq_inst_plat.items():
            if metadata["sequencing_instrument_model"] in values:

                new_data["sequencing_instrument_platform"] = key
                break

        return new_data

    def add_additional_data(self, metadata, lab_json_file, geo_loc_file):
        """Add the additional information that must be included in final metadata
        metadata Origin metadata
        extra_data  additional data to be included
        result_metadata    final metadata after adding the additional data
        """
        lab_data = {}
        additional_metadata = []

        lab_json = self.read_json_file(lab_json_file)
        geo_loc_json = self.read_json_file(geo_loc_file)
        samples_json = self.read_json_file(self.sample_list_file)
        for row_sample in metadata:
            """Include sample data from sample json"""
            try:
                for key, value in samples_json[
                    row_sample["microbiology_lab_sample_id"]
                    # row_sample["sequencing_sample_id"]
                ].items():

                    row_sample[key] = value
            except KeyError as e:
                stderr.print(
                    "[red] ERROR  fastq information not found in sample json. ", e
                )

            """ Fetch the information related to the laboratory.
                Info is stored in lab_data, to prevent to call get_laboratory_data
                each time for each sample that belongs to the same lab
            """
            if row_sample["collecting_institution"] not in lab_data:
                # from collecting_institution find city, and geo location latitude and longitude
                l_data = self.get_laboratory_data(
                    lab_json, geo_loc_json, row_sample["collecting_institution"]
                )
                row_sample.update(l_data)
                lab_data[row_sample["collecting_institution"]] = l_data
            else:
                row_sample.update(lab_data[row_sample["collecting_institution"]])

            """ Fetch email and address for submitting_institution
            """
            row_sample["submitting_institution"] = row_sample[
                "submitting_institution"
            ].strip()
            if row_sample["submitting_institution"] not in lab_json:
                l_data = self.get_laboratory_data(
                    lab_json, geo_loc_json, row_sample["submitting_institution"]
                )
                # row_sample.update(l_data)
                lab_data[row_sample["submitting_institution"]] = l_data
            sub_data = {}

            sub_data["submitting_institution_email"] = lab_data[
                row_sample["submitting_institution"]
            ]["collecting_institution_email"]
            sub_data["submitting_institution_address"] = lab_data[
                row_sample["submitting_institution"]
            ]["collecting_institution_address"]
            # else:
            #    sub_data = {"collecting_institution_email" : "", "collecting_institution_address": ""}
            row_sample.update(sub_data)

            """ Add Fixed information
            """
            row_sample.update(self.include_fixed_data())

            """ Add fields that are already in other fields
            """
            row_sample.update(self.include_fields_already_set(row_sample))
            """Add information which requires processing
            """
            row_sample.update(self.include_processed_data(row_sample))
            """
            row["isolate"] = row["collecting_lab_sample_id"]
            row["host_scientific_name"] = extra_data["host_scientific_name"][
                row["host_common_name"]
            ]
            row["sequencing_instrument_platform"] = "To change"
            """
            # Add experiment_alias and run_alias
            row_sample["experiment_alias"] = str(
                row_sample["sequence_file_R1_fastq"]
                + "_"
                + row_sample["sequence_file_R2_fastq"]
            )
            row_sample["run_alias"] = str(
                row_sample["sequence_file_R1_fastq"]
                + "_"
                + row_sample["sequence_file_R2_fastq"]
            )
            additional_metadata.append(row_sample)

        return additional_metadata

    def request_information(external_url, request):
        """Get information from external database server using Rest API

        external_url
        request
        """
        pass

    def store_information(external_url, request, data):
        """Update information"""
        pass

    def read_metadata_file(self):
        """Read the input metadata file, changing the metadata heading with
        their property name values defined in schema.
        Convert the date colunms value to the dd/mm/yyyy format.
        Return list of dict with data, and errors
        """
        # exc_format_num = ["Sample ID given for sequencing"]
        wb_file = openpyxl.load_workbook(self.metadata_file, data_only=True)
        ws_metadata_lab = wb_file["METADATA_LAB"]

        # removing the None columns in excel heading row
        heading = [i.value.strip() for i in ws_metadata_lab[4] if i.value]

        # heading = self.update_heading_to_json(heading_without_none, meta_map_json)
        metadata_values = []
        errors = {}
        for row in islice(ws_metadata_lab.values, 4, ws_metadata_lab.max_row):
            sample_data_row = {}
            # Ignore the empty rows
            if row[2] is None:
                continue
            for idx in range(1, len(heading)):
                if "date" in heading[idx].lower():
                    try:
                        sample_data_row[self.label_prop_dict[heading[idx]]] = row[
                            idx
                        ].strftime("%Y-%m-%d")
                    except AttributeError:
                        # check if date is in string format
                        str_date = re.search(r"(\d{4}-\d{2}-\d{2}).*", row[idx])
                        if str_date:
                            sample_data_row[
                                self.label_prop_dict[heading[idx]]
                            ] = str_date.group(1)
                        else:
                            if row[2] not in errors:
                                errors[row[2]] = {}
                            errors[row[2]][heading[idx]] = "Invalid date format"
                            log.error("Invalid date format in sample %s", row[2])
                            stderr.print(
                                "[red] Invalid date format in sample",
                                row[2] + " column " + heading[idx],
                            )
                else:

                    if isinstance(row[idx], float) or isinstance(row[idx], int):
                        val = str(int(row[idx]))
                        try:
                            sample_data_row[self.label_prop_dict[heading[idx]]] = val
                        except TypeError as e:
                            stderr.print("[red] Error when reading " + row[2] + e)
                    else:
                        try:
                            sample_data_row[self.label_prop_dict[heading[idx]]] = (
                                row[idx] if row[idx] else ""
                            )
                        except KeyError as e:
                            stderr.print(
                                "[red] Error when reading " + str(row[2]) + str(e)
                            )
            metadata_values.append(sample_data_row)

        return metadata_values, errors

    def write_json_fo_file(self, data, file_name):
        """Write metadata to json file"""
        os.makedirs(self.output_folder, exist_ok=True)
        json_file = os.path.join(self.output_folder, file_name)
        with open(json_file, "w", encoding="utf-8") as fh:
            fh.write(json.dumps(data, indent=4, sort_keys=True, ensure_ascii=False))
        return True

    def create_metadata_json(self):
        stderr.print("[blue] Reading configuration settings")
        config_json = ConfigJson()
        geo_loc_json = config_json.get_configuration("geo_location_data")
        geo_loc_file = os.path.join(
            os.path.dirname(os.path.realpath(__file__)), "conf", geo_loc_json
        )
        lab_json = config_json.get_configuration("laboratory_data")
        lab_json_file = os.path.join(
            os.path.dirname(os.path.realpath(__file__)), "conf", lab_json
        )
        stderr.print("[blue] Reading Lab Metadata Excel File")
        valid_metadata_rows, errors = self.read_metadata_file()
        if len(errors) > 0:
            stderr.print("[red] Stopped executing because the errors found")
            sys.exit(1)
        # Continue by adding extra information
        stderr.print("[blue] Including additional information")
        completed_metadata = self.add_additional_data(
            valid_metadata_rows,
            lab_json_file,
            geo_loc_file,
        )

        file_name = (
            "processed_"
            + os.path.splitext(os.path.basename(self.metadata_file))[0]
            + ".json"
        )
        stderr.print("[blue] Creating Json file")
        self.write_json_fo_file(completed_metadata, file_name)
        return True
