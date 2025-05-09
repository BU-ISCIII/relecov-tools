#!/usr/bin/env python
import rich.console
from jsonschema import Draft202012Validator, FormatChecker
import sys
import os
import openpyxl
from datetime import datetime

import relecov_tools.utils
import relecov_tools.assets.schema_utils.jsonschema_draft
import relecov_tools.assets.schema_utils.custom_validators
from relecov_tools.config_json import ConfigJson
from relecov_tools.base_module import BaseModule


stderr = rich.console.Console(
    stderr=True,
    style="dim",
    highlight=False,
    force_terminal=relecov_tools.utils.rich_force_colors(),
)


class SchemaValidation(BaseModule):
    def __init__(
        self,
        json_data_file=None,
        json_schema_file=None,
        metadata=None,
        out_folder=None,
        excel_sheet=None,
        registry=None,
    ):
        """Validate json file against the schema"""
        super().__init__(output_directory=out_folder, called_module=__name__)
        config_json = ConfigJson()
        self.log.info("Initiating validation process")
        if json_schema_file is None:
            schema_name = config_json.get_topic_data("json_schemas", "relecov_schema")
            json_schema_file = os.path.join(
                os.path.dirname(os.path.realpath(__file__)), "schema", schema_name
            )

        self.json_schema = relecov_tools.utils.read_json_file(json_schema_file)

        if json_data_file is None:
            json_data_file = relecov_tools.utils.prompt_path(
                msg="Select the json file to be validated"
            )

        if out_folder is None:
            self.out_folder = relecov_tools.utils.prompt_path(
                msg="Select the folder where excel file with invalid data will be saved"
            )
        else:
            self.out_folder = out_folder

        # Read and check json to validate file
        if not os.path.isfile(json_data_file):
            stderr.print("[red] Json file does not exist")
            self.log.error("Json file does not exist")
            sys.exit(1)
        self.json_data_file = json_data_file
        out_path = os.path.dirname(os.path.realpath(self.json_data_file))
        self.lab_code = out_path.split("/")[-2]
        self.logsum = self.parent_log_summary(
            output_location=self.out_folder, unique_key=self.lab_code, path=out_path
        )

        stderr.print("[blue] Reading the json file")
        self.log.info("Reading the json file")
        self.json_data = relecov_tools.utils.read_json_file(json_data_file)
        if not isinstance(self.json_data, list):
            stderr.print(f"[red]Invalid json file content in {json_data_file}.")
            stderr.print("Should be a list of dicts. Create it with read-lab-metadata")
            self.log.error(f"[red]Invalid json file content in {json_data_file}.")
            self.log.error(
                "Should be a list of dicts. Create it with read-lab-metadata"
            )
            raise TypeError(f"Invalid json file content in {json_data_file}")
        try:
            batch_id = self.get_batch_id_from_data(self.json_data)
        except ValueError:
            raise ValueError(f"Provided json file {json_data_file} is empty")
        except AttributeError as e:
            raise ValueError(f"Invalid json file content in {json_data_file}: {e}")
        self.set_batch_id(batch_id)

        self.metadata = metadata
        try:
            self.sample_id_field = self.get_sample_id_field()
        except ValueError as e:
            self.sample_id_field = None
            self.SAMPLE_FIELD_ERROR = str(e)
        conf_subdata = config_json.get_topic_data("sftp_handle", "metadata_processing")
        if excel_sheet is None:
            try:
                self.excel_sheet = conf_subdata["excel_sheet"]
            except KeyError:
                self.log.error("Default metadata sheet name should be in config file")
                raise
        else:
            self.excel_sheet = excel_sheet

        # Parse file containing sample IDs registry
        if registry is not None and relecov_tools.utils.file_exists(registry):
            self.registry_path = registry
        else:
            config_json = ConfigJson(extra_config=True)
            try:
                default_path = config_json.get_topic_data(
                    "validate_config", "default_sample_id_registry"
                )
            except KeyError:
                default_path = None

            if default_path and relecov_tools.utils.file_exists(default_path):
                self.registry_path = default_path
            else:
                stderr.print(
                    "[yellow]No valid ID registry found. Please select the file manually."
                )
                prompted_path = relecov_tools.utils.prompt_path(
                    "Select the JSON file with registered unique sample IDs"
                )
                if relecov_tools.utils.file_exists(prompted_path):
                    self.registry_path = prompted_path
                else:
                    stderr.print(
                        "[red]No valid ID registry file could be found or selected."
                    )
                    sys.exit(1)

        # Read id registry file
        try:
            self.id_registry = relecov_tools.utils.read_json_file(self.registry_path)
        except Exception as e:
            stderr.print(f"[red]Failed to read ID registry JSON: {e}")
            self.log.error(f"Failed to read ID registry JSON: {e}")
            raise

    def validate_schema(self):
        """Validate json schema against draft"""
        relecov_tools.assets.schema_utils.jsonschema_draft.check_schema_draft(
            self.json_schema, "2020-12"
        )

    def get_sample_id_field(self):
        """Find the name of the field used to track the samples in the given schema"""
        # TODO: Include this field in configuration.json
        sample_id_ontology = "GENEPIO:0000079"
        ontology_match = [
            x
            for x, y in self.json_schema["properties"].items()
            if y.get("ontology") == sample_id_ontology
        ]
        if ontology_match:
            sample_id_field = ontology_match[0]
        else:
            error_text = f"No valid sample ID field ({sample_id_ontology}) in schema"
            raise ValueError(error_text)
        return sample_id_field

    def validate_instances(self):
        """Validate data instances against a validated JSON schema"""

        # Create validator
        validator = Draft202012Validator(
            self.json_schema, format_checker=FormatChecker()
        )
        schema_props = self.json_schema["properties"]

        validated_json_data = []
        invalid_json = []
        errors = {}
        error_keys = {}

        if self.sample_id_field is None:
            log_text = f"Logs keys set to None. Reason: {self.SAMPLE_FIELD_ERROR}"
            self.logsum.add_warning(sample=self.sample_id_field, entry=log_text)

        stderr.print("[blue] Start processing the JSON file")
        self.log.info("Start processing the JSON file")

        # Prepare ID registry (validated samples will be asigned with an unique ID)
        new_ids_to_save = {}

        # Start validation
        for item_row in self.json_data:
            sample_id_value = item_row.get(self.sample_id_field)

            # Collect all errors (don't raise immediately)
            validation_errors = list(validator.iter_errors(item_row))

            # Run the custom validator to check if errors should be ignored
            validation_errors = relecov_tools.assets.schema_utils.custom_validators.validate_with_exceptions(
                self.json_schema, item_row, validation_errors
            )
            if not validation_errors:
                # If sample has validated, then assign an unique id
                if "unique_sample_id" not in item_row:
                    new_id = self.generate_incremental_unique_id(
                        {**self.id_registry, **new_ids_to_save}
                    )
                    item_row["unique_sample_id"] = new_id
                    new_ids_to_save[new_id] = {
                        "sequencing_sample_id": sample_id_value,
                        "lab_code": self.lab_code,
                        "generated_at": datetime.now().isoformat(timespec="seconds"),
                    }
                validated_json_data.append(item_row)
                self.logsum.feed_key(sample=sample_id_value)
            else:
                # Process remaining errors
                for error in validation_errors:
                    try:
                        if error.validator == "required":
                            error_field = list(error.message.split("'"))[1]
                        elif error.validator == "anyOf":
                            missing_fields = []
                            for cond in error.validator_value:
                                if isinstance(cond, dict) and "required" in cond:
                                    missing_fields.extend(cond["required"])
                            error_field = " or ".join(missing_fields)
                        elif error.absolute_path:
                            error_field = str(error.absolute_path[0])
                        else:
                            error_field = error.validator or error.message
                    except Exception as ex:
                        self.log.warning(
                            f"Error extracting error_field from: {error}, {ex}"
                        )
                        error_field = str(error)

                    # Try to get the human-readable label from the schema
                    try:
                        err_field_label = schema_props[error_field]["label"]
                    except KeyError:
                        self.log.error(f"Could not extract label for {error_field}")
                        err_field_label = error_field
                    # Format the error message
                    error.message = error.message.replace(error_field, err_field_label)
                    if (
                        error.validator == "format" and error.validator_value == "date"
                    ):  # Modification of default text warning for invalid date format.
                        error_text = f"Error in column {err_field_label}: '{error.instance}' is not a valid date format. Valid format 'YYYY-MM-DD'"
                    else:
                        error_text = (
                            f"Error in column {err_field_label}: {error.message}"
                        )

                    # Log errors for summary
                    error_keys[error.message] = error_field
                    errors[error.message] = errors.get(error.message, 0) + 1
                    self.logsum.add_error(sample=sample_id_value, entry=error_text)

                # Add the invalid row to the list
                invalid_json.append(item_row)

        # For samples that are valid, save its new unique ID in the registry
        if new_ids_to_save:
            self.save_new_ids(new_ids_to_save)

        # Summarize errors
        stderr.print("[blue] --------------------")
        stderr.print("[blue] VALIDATION SUMMARY")
        stderr.print("[blue] --------------------")
        self.log.info("Validation summary:")
        max_length = 250
        for error_type, count in errors.items():
            field_with_error = error_keys[error_type]
            if (
                "is not a 'date'" in error_type
            ):  # Modification of default text warning for invalid date format.
                error_text = f"{count} samples failed validation for {field_with_error}:\n{error_type.split()[0]} is not a valid date format. Valid format 'YYYY-MM-DD"
            else:
                error_text = f"{count} samples failed validation for {field_with_error}:\n{error_type}"
            truncated_msg = (
                error_text[:max_length] + "..."
                if len(error_text) > max_length
                else error_text
            )
            self.logsum.add_warning(entry=truncated_msg)
            stderr.print(f"[red]{truncated_msg}")
            stderr.print("[red] --------------------")

        return validated_json_data, invalid_json

    def create_invalid_metadata(self, invalid_json, metadata, out_folder):
        """Create a new sub excel file having only the samples that were invalid.
        Samples name are checking the Sequencing sample id which are in
        column B (index 1).
        The rows that match the value collected from json file on tag
        collecting_lab_sample_id are removed from excel
        """
        if self.sample_id_field is None:
            log_text = f"Invalid excel file won't be created: {self.SAMPLE_FIELD_ERROR}"
            self.logsum.add_error(entry=log_text)
            return
        self.log.error("Some of the samples in json metadata were not validated")
        stderr.print("[red] Some of the Samples are not validate")
        if metadata is None:
            metadata = relecov_tools.utils.prompt_path(
                msg="Select the metadata file to select those not-validated samples."
            )
        if not os.path.isfile(metadata):
            self.log.error("Metadata file %s does not exist", metadata)
            stderr.print(
                "[red] Unable to create excel file for invalid samples. Metadata file ",
                metadata,
                " does not exist",
            )
            sys.exit(1)
        sample_list = []
        stderr.print("Start preparation of invalid samples")
        self.log.info("Start preparation of invalid samples")
        for row in invalid_json:
            sample_list.append(str(row[self.sample_id_field]))
        wb = openpyxl.load_workbook(metadata)
        try:
            ws_sheet = wb[self.excel_sheet]
        except KeyError:
            logtxt = f"No sheet named {self.excel_sheet} could be found in {metadata}"
            self.log.error(logtxt)
            raise
        tag = "Sample ID given for sequencing"
        # Check if mandatory colum ($tag) is defined in metadata.
        try:
            id_col = next(
                idx
                for idx, cell in enumerate(ws_sheet[1])
                if cell.value is not None and tag in str(cell.value)
            )
        except StopIteration:
            self.log.error(
                f"Column with tag '{tag}' not found in the second row of the Excel sheet."
            )
            stderr.print(f"[red] Column with tag '{tag}' not found. Cannot continue.")
            raise
        row_to_del = []
        row_iterator = ws_sheet.iter_rows(min_row=2, max_row=ws_sheet.max_row)
        consec_empty_rows = 0
        for row in row_iterator:
            # if no data in 10 consecutive rows, break loop
            if not any(row[x].value for x in range(10)):
                row_to_del.append(row[0].row)
                consec_empty_rows += 1
            if consec_empty_rows > 10:
                break
            consec_empty_rows = 0
            if str(row[id_col].value) not in sample_list:
                row_to_del.append(row[0].row)
        stderr.print("Collected rows to create the excel file")
        if len(row_to_del) > 0:
            row_to_del.sort(reverse=True)
            for idx in row_to_del:
                try:
                    ws_sheet.delete_rows(idx)
                except TypeError as e:
                    self.log.error(
                        "Unable to delete row %s from metadata file because of",
                        idx,
                        e,
                    )
                    stderr.print(f"[red] Unable to delete row {idx} becuase of {e}")
                    sys.exit(1)
        os.makedirs(out_folder, exist_ok=True)
        new_name = "invalid_" + os.path.basename(metadata)
        m_file = os.path.join(out_folder, new_name)
        self.log.info("Saving excel file with the invalid samples")
        stderr.print("Saving excel file with the invalid samples")
        wb.save(m_file)
        return

    def create_validated_json(self, valid_json_data, out_folder):
        """Create a copy of the input json file, keeping only the validated samples

        Args:
            valid_json_data (list(dict)): List of samples metadata as dictionaries
            out_folder (str): path to folder where file will be created
        """
        file_name = "_".join(["validated", os.path.basename(self.json_data_file)])
        file_path = os.path.join(out_folder, file_name)
        self.log.info("Saving Json file with the validated samples in %s", file_path)
        relecov_tools.utils.write_json_to_file(valid_json_data, file_path)
        return

    def generate_incremental_unique_id(self, current_registry, prefix="RLCV"):
        """Generates an incremental unique ID using as baseline the latest record
        in a registry.

        Args:
            current_registry (dict): dict containing sample's unique IDs
            prefix (str, optional): String that preceeds the unique ID. Defaults to "RLCV".

        Returns:
            string: An unique ID
        """
        existing_numbers = []
        for uid in current_registry:
            if uid.startswith(prefix):
                try:
                    number = int(uid.replace(f"{prefix}-", ""))
                    existing_numbers.append(number)
                except ValueError:
                    continue
        next_number = max(existing_numbers, default=0) + 1
        return f"{prefix}-{next_number:09d}"

    def save_new_ids(self, new_ids_dict):
        """Updates sample id registry by adding sample unique ids of new validated samples.

        Args:
            new_ids_dict (dict): Dict of new unique sample IDs.
        """
        updated_registry = {**self.id_registry, **new_ids_dict}
        relecov_tools.utils.write_json_to_file(updated_registry, self.registry_path)
        self.log.info(f"Added {len(new_ids_dict)} new sample IDs to registry")

    def validate(self):
        """Validate samples from metadata, create an excel with invalid samples,
        and a json file with the validated ones
        """
        self.validate_schema()
        valid_json_data, invalid_json = self.validate_instances()
        if not invalid_json:
            stderr.print("[green]Sucessful validation, no invalid file created!!")
            self.log.info("Sucessful validation, no invalid file created.")
        else:
            log_text = "Summary: %s valid and %s invalid samples"
            self.logsum.add_warning(
                entry=log_text % (len(valid_json_data), len(invalid_json))
            )
            self.create_invalid_metadata(invalid_json, self.metadata, self.out_folder)
        if valid_json_data:
            self.create_validated_json(valid_json_data, self.out_folder)
        else:
            log_text = "All the samples were invalid. No valid file created"
            self.logsum.add_error(entry=log_text)
            stderr.print(f"[red]{log_text}")
        self.parent_create_error_summary(called_module="validate")
        return valid_json_data, invalid_json
