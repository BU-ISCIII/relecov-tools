#!/usr/bin/env python
import rich.console
import pandas as pd
import os
import re
import openpyxl
import sys
import json
import difflib
import inspect

import relecov_tools.utils
import relecov_tools.assets.schema_utils.amr_metadata
import relecov_tools.assets.schema_utils.jsonschema_draft
import relecov_tools.assets.schema_utils.metadatalab_template
from relecov_tools.config_json import ConfigJson
from relecov_tools.base_module import BaseModule
from datetime import datetime
from openpyxl.worksheet.datavalidation import DataValidation

pd.set_option("future.no_silent_downcasting", True)

stderr = rich.console.Console(
    stderr=True,
    style="dim",
    highlight=False,
    force_terminal=relecov_tools.utils.rich_force_colors(),
)


def _slugify_project(value):
    """Return a filesystem-friendly project identifier."""
    project = str(value).strip().lower()
    project = re.sub(r"\s+", "_", project)
    project = re.sub(r"[^a-z0-9_-]+", "_", project)
    return project.strip("_")


def _display_project(value):
    """Return a readable project label for generated metadata."""
    return str(value).strip().replace("_", " ").replace("-", " ").title()


class BuildSchema(BaseModule):
    def __init__(
        self,
        input_file=None,
        schema_base=None,
        excel_template=None,
        draft_version=None,
        diff=False,
        output_dir=None,
        version=None,
        project=None,
        non_interactive=False,
        initial_version=False,
    ):
        """
        Initialize the SchemaBuilder class. This class generates a JSON Schema file based on the provided draft version.
        It reads the database definition from an Excel file and allows customization of the schema generation process.
        """
        super().__init__(output_dir=output_dir, called_module=__name__)
        self.excel_file_path = input_file
        self.excel_template = excel_template
        self.non_interactive = non_interactive
        self.initial_version = initial_version
        # Validate params
        if self.initial_version and self.excel_template:
            raise ValueError(
                "--initial-version and --excel_template are incompatible options."
            )
        if not self.excel_file_path or not os.path.isfile(self.excel_file_path):
            self.log.error("A valid Excel file path must be provided.")
            raise ValueError("A valid Excel file path must be provided.")
        if not self.excel_file_path.endswith(".xlsx"):
            self.log.error("The Excel file must have a .xlsx extension.")
            raise ValueError("The Excel file must have a .xlsx extension.")

        # No metadata is being processed so batch_id will be execution date
        self.set_batch_id(self.basemod_date)
        # Cache for generic enum source sheets referenced from the input Excel.
        self._enum_sheet_cache = {}

        # Validate output folder creation
        if not output_dir:
            self.output_dir = relecov_tools.utils.prompt_create_outdir()
        else:
            self.output_dir = os.path.abspath(output_dir)
            if not os.path.exists(self.output_dir):
                self.output_dir = relecov_tools.utils.prompt_create_outdir()

        # Get version option
        if not version:
            # If not defined, then ask via prompt
            self.version = relecov_tools.utils.prompt_text(
                "Write the desired version using semantic versioning:"
            )
        else:
            self.version = version
        if not relecov_tools.utils.validate_semantic_version(self.version):
            raise ValueError("[red]Error: Invalid version format")

        # Validate show diff option
        if diff is False:
            self.show_diff = None
        else:
            self.show_diff = True

        # Validate json schema draft version
        if not draft_version and self.non_interactive:
            self.draft_version = "2020-12"
        else:
            self.draft_version = (
                relecov_tools.assets.schema_utils.jsonschema_draft.check_valid_version(
                    draft_version
                )
            )

        # Get version option
        # Parse build-schema configuration
        self.build_schema_json_file = os.path.join(
            os.path.dirname(__file__), "conf", "build_schema_config.json"
        )

        if project is None:
            project = relecov_tools.utils.prompt_text("Write the desired project:")

        self.project = _slugify_project(project)
        if not self.project:
            raise ValueError("A valid project must be provided.")
        self.project_label = _display_project(self.project)
        self.schema_output_filename = f"{self.project}_schema.json"
        self.template_output_prefix = f"{self.project}_metadata_template"

        available_projects = self.get_available_projects(self.build_schema_json_file)

        # Get collecting institutions and dropdown list
        self._lab_dropdowns, self._lab_uniques = self._load_laboratory_addresses()

        # Config params
        config_build_schema = ConfigJson(self.build_schema_json_file)
        config_data = config_build_schema.get_configuration("projects") or {}
        self.configurables = (
            config_build_schema.get_configuration("configurables") or {}
        )
        if self.project in available_projects:
            self.project_config = config_data.get(self.project, {})
        else:
            self.log.error(
                f"No configuration available for '{self.project}'. Available projects: {', '.join(available_projects)}"
            )
            stderr.print(
                f"[red]No configuration available for '{self.project}'. Available projects: {', '.join(available_projects)}"
            )
            raise ValueError(
                f"No configuration available for '{self.project}'. Available projects: {', '.join(available_projects)}"
            )

        # Validate base schema
        if schema_base is not None:
            if relecov_tools.utils.file_exists(schema_base):
                self.base_schema_path = schema_base
            else:
                self.log.error(
                    f"[Error]Defined base schema file not found: {schema_base}."
                )
                stderr.print(
                    f"[Error]Defined base schema file not found: {schema_base}. Exiting..."
                )
                raise FileNotFoundError(
                    f"Defined base schema file not found: {schema_base}."
                )
        else:
            project_schema = self.project_config.get(
                "schema_file", self.schema_output_filename
            )
            self.base_schema_path = os.path.join(
                os.path.dirname(os.path.realpath(__file__)),
                "schema",
                project_schema,
            )

            if not relecov_tools.utils.file_exists(self.base_schema_path):
                self.log.error(
                    f"[Error]Fatal error. {self.project_label} schema was not found in current relecov-tools installation: {self.base_schema_path}."
                )
                stderr.print(
                    f"[Error]Fatal error. {self.project_label} schema was not found in current relecov-tools installation: {self.base_schema_path}. Exiting..."
                )
                raise FileNotFoundError(
                    f"Fatal error. {self.project_label} schema was not found in current relecov-tools installation: {self.base_schema_path}."
                )

            self.log.info(
                "%s schema successfully found in the installation.",
                self.project_label,
            )
            stderr.print(
                f"[green]{self.project_label} schema successfully found in the installation."
            )

        self._resolve_version_history_template()

    def _resolve_version_history_template(self):
        """Resolve the previous Excel template used to read VERSION history.

        Initial versions skip previous template lookup; regular versions use either
        the explicit template path or the installed project template in assets.
        """
        if self.initial_version:
            self.excel_template = None
            return

        if self.excel_template:
            if not os.path.isfile(self.excel_template):
                raise FileNotFoundError(
                    f"Defined excel template file not found: {self.excel_template}."
                )
            return

        excel_template_path = os.path.join(
            os.path.dirname(os.path.realpath(__file__)), "assets"
        )
        template_prefix = f"{self.project}_metadata_template".lower()
        excel_templates = [
            f
            for f in os.listdir(excel_template_path)
            if f.lower().startswith(template_prefix) and f.lower().endswith(".xlsx")
        ]
        if len(excel_templates) > 1:
            self.log.error(
                "[Error]Fatal error. More than one excel template was found in current relecov-tools installation (assets)"
            )
            stderr.print(
                "[Error]Fatal error. More than one excel template was found in current relecov-tools installation (assets). Exiting..."
            )
            raise FileExistsError(
                "Fatal error. More than one excel template was found in current relecov-tools installation (assets)"
            )
        if not excel_templates:
            msg = (
                f"{self.project_label} excel template was not found in assets. "
                "Use --initial-version to start a new VERSION history explicitly."
            )
            self.log.error(msg)
            stderr.print(f"[red]{msg}")
            raise FileNotFoundError(msg)

        self.excel_template = os.path.join(excel_template_path, excel_templates[0])

    def _load_laboratory_addresses(self):
        """
        Returns two dictionaries with key in the three special fields:
        - dropdowns[field] ........ list ‘<name> [<city>] [<ccn>]’
        - uniques[field] .......... unique names for schema enum

        NOTE:
        For RELECOV, laboratory_address.json stores institution names under
        `collecting_institution`. We intentionally reuse that same source for
        collecting/submitting/sequencing to keep the three schema enums aligned.
        """
        json_path = os.path.join(
            os.path.dirname(__file__),
            "conf",
            "laboratory_address.json",
        )
        with open(json_path, encoding="utf-8") as fh:
            lab_data = json.load(fh)

        fields = [
            "collecting_institution",
            "submitting_institution",
            "sequencing_institution",
        ]
        dropdowns = {f: [] for f in fields}
        uniques = {f: set() for f in fields}

        for ccn, info in lab_data.items():
            city = info.get("geo_loc_city", "").strip()
            name = info.get("collecting_institution", "").strip()
            if not name:
                continue
            dropdown_entry = f"{name} [{city}] [{ccn}]"
            for f in fields:
                dropdowns[f].append(dropdown_entry)
                uniques[f].add(name)

        dropdowns = {
            k: sorted(self._unique_enum_values(v)) for k, v in dropdowns.items()
        }
        uniques = {k: sorted(v) for k, v in uniques.items()}
        return dropdowns, uniques

    def _is_template_only_property(self, property_definition: dict) -> bool:
        """
        Check whether a database definition row should only be included in the template.

        Args:
            property_definition (dict): Definition of one property from the input Excel.

        Returns:
            bool: True if submitting_lab_form is set to ONLY, False otherwise.
        """
        submitting_lab_form = property_definition.get("submitting_lab_form")
        if not isinstance(submitting_lab_form, str):
            return False
        return submitting_lab_form.strip().upper() == "ONLY"

    def _save_amr_genes_json(self):
        """
        Generate the AMR genes metadata JSON file when amr_gene_list is available.

        The AMR-specific parsing and writing logic lives in schema_utils. This
        method only integrates that output into the build-schema workflow.
        """
        output_path = (
            relecov_tools.assets.schema_utils.amr_metadata.save_amr_genes_json(
                self.excel_file_path,
                self.output_dir,
                self.version,
            )
        )
        if not output_path:
            return

        self.log.info("AMR genes JSON saved to %s", output_path)
        stderr.print(f"[green]AMR genes JSON saved to: {output_path}")

    def validate_database_definition(self, json_data):
        """Validate the mandatory features and ensure:
        - No duplicate enum values in the JSON schema.
        - All fields have an example.
        - All examples should have the same type as JSON schema.
        - Date formats follow 'YYYY-MM-DD'.

        Args:
            json_data (dict): The JSON data representing the database definition.

        Returns:
            dict: A dictionary containing errors found, categorized by:
                - Missing features
                - Duplicate enums
                - Missing examples
                - Invalid enum examples
                - Invalid example types
                - Incorrect date formats
        """
        log_errors = {
            "missing_features": {},
            "missing_examples": {},
            "duplicate_enums": {},
            "invalid_enum_examples": {},
            "invalid_example_types": {},
            "invalid_date_formats": {},
        }

        mandatory_features = [
            "enum",
            "examples",
            "ontology_id",
            "type",
            "description",
            "classification",
            "label_name",
            "fill_mode",
            "required (Y/N)",
            "complex_field (Y/N)",
        ]

        # Iterate over properties in json_data
        for prop_name, prop_features in json_data.items():
            missing_features = [
                feature
                for feature in mandatory_features
                if feature not in prop_features
            ]
            if missing_features:
                log_errors["missing_features"][prop_name] = missing_features

            # Check for duplicate enum values
            if prop_features.get("enum"):
                if not pd.isna(prop_features["enum"]):
                    enum_values = prop_features["enum"].split("; ")
                    # Verify that enum has no duplicates
                    if len(enum_values) != len(set(enum_values)):
                        duplicates = [
                            value
                            for value in set(enum_values)
                            if enum_values.count(value) > 1
                        ]
                        log_errors["duplicate_enums"][prop_name] = duplicates

            # Check for missing examples
            example = prop_features.get("examples")
            if example is None:
                log_errors["missing_examples"][prop_name] = ["Missing example."]

            feature_type = prop_features["type"]

            # Check examples are included in the enum values, when an enum exists.
            invalid_enum_examples = self._validate_examples_in_enum(
                prop_name,
                example,
                prop_features.get("enum"),
                feature_type,
            )
            if invalid_enum_examples:
                log_errors["invalid_enum_examples"][prop_name] = invalid_enum_examples

            match feature_type:
                # Check date format for properties with type=string and format=date
                case "string":
                    if "format:date" in str(prop_features.get("options", "")).replace(
                        " ", ""
                    ):
                        if isinstance(example, datetime):
                            example = example.strftime("%Y-%m-%d")
                        if isinstance(example, str):
                            try:
                                datetime.strptime(example, "%Y-%m-%d")
                            except ValueError:
                                if prop_name not in log_errors["invalid_date_formats"]:
                                    log_errors["invalid_date_formats"][prop_name] = []
                                log_errors["invalid_date_formats"][prop_name].append(
                                    f"Invalid date format '{example}', expected 'YYYY-MM-DD'"
                                )

                case "integer" | "number":
                    function_to_convert = float if feature_type == "number" else int
                    try:
                        example = function_to_convert(example)
                    except ValueError:
                        log_errors["invalid_example_types"][prop_name] = [
                            f"Value {example} is not a valid {feature_type}"
                        ]

        # return log errors if any
        if any(log_errors.values()):
            stderr.print("[red]\t- Database Validation Failed")
            # Convert log_errors dictionary to DataFrame
            df_errors = pd.DataFrame(
                [
                    {
                        "Error Category": category,
                        "Field": field,
                        "Details": (
                            ", ".join(details) if isinstance(details, list) else details
                        ),
                    }
                    for category, errors in log_errors.items()
                    for field, details in errors.items()
                ]
            )

            # Save errors to file
            error_file_path = f"{self.output_dir}/schema_validation_errors.csv"
            df_errors.to_csv(error_file_path, index=False, encoding="utf-8")

            # Provide errors to user in rich table format:
            relecov_tools.utils.display_dataframe_to_user(
                name="Schema Validation Errors", dataframe=df_errors
            )
            stderr.print(f"\t- Log errors saved to:\n\t{error_file_path}")

            # Ask user whether to continue or stop execution
            if self.non_interactive or relecov_tools.utils.prompt_yn_question(
                "Errors found in database values. Do you want to continue? (Y/N)"
            ):
                pass
            else:
                return log_errors
        else:
            stderr.print("[green]\t- Database validation passed")

        # If no errors found
        return None

    def _validate_examples_in_enum(
        self,
        property_id: str,
        example_value: any,
        enum_value: any,
        expected_type: str | None,
    ) -> list[str]:
        """Return validation errors for examples that are not present in enum."""
        if self._is_empty_validation_value(enum_value):
            return []
        if self._is_empty_validation_value(example_value):
            return []

        enum_values = self._parse_enum_values(enum_value)
        if not isinstance(enum_values, list) or not enum_values:
            return []

        examples = self._parse_examples_for_validation(example_value)
        examples = self._cast_examples_to_declared_type(
            property_id, expected_type, examples
        )

        enum_lookup = {
            self._normalize_enum_example_value(value) for value in enum_values
        }
        return [
            f"Example '{example}' is not defined in enum."
            for example in examples
            if self._normalize_enum_example_value(example) not in enum_lookup
        ]

    @staticmethod
    def _normalize_enum_example_value(value: any) -> any:
        if not isinstance(value, str):
            return value
        return BuildSchema._clean_enum_ontology_annotation(value)

    @staticmethod
    def _clean_enum_ontology_annotation(value: any) -> any:
        """Remove ontology annotations displayed between brackets from enum labels."""
        if not isinstance(value, str):
            return value
        return re.sub(r"\s*\[[^\]]+\]", "", value).strip()

    def _clean_template_enum_values(self, values: any) -> any:
        """Return enum values as displayed in the Excel template dropdowns."""
        if not isinstance(values, list):
            return values
        return self._unique_enum_values(
            [self._clean_enum_ontology_annotation(value) for value in values]
        )

    @staticmethod
    def _is_empty_validation_value(value: any) -> bool:
        if value is None or isinstance(value, list):
            return value is None
        return pd.isna(value)

    @staticmethod
    def _parse_examples_for_validation(example_value: any) -> list[any]:
        """Parse the examples cell using the same separator used for schema examples."""
        if isinstance(example_value, str):
            return [
                value.strip() for value in example_value.split("; ") if value.strip()
            ]
        if isinstance(example_value, datetime):
            return [example_value.strftime("%Y-%m-%d")]
        if isinstance(example_value, float) and example_value.is_integer():
            return [int(example_value)]
        return [example_value]

    def get_available_projects(self, json):
        """Get list of available software in configuration

        Args:
            json (str): Path to bioinfo configuration json file.

        Returns:
            available_software: List containing available software defined in json.
        """
        config = relecov_tools.utils.read_json_file(json)
        # available_software = list(config.keys())
        available_projects = list(config.get("projects", {}).keys())
        return available_projects

    def read_database_definition(self, sheet_id="main"):
        """Reads the database definition from an Excel sheet and converts it into JSON format.

        Args:
            sheet_id (str): The sheet name or ID in the Excel file to read from. Defaults to "main".

        Returns:
            json_data (dict): The JSON data representing the database definition.
        """
        caller_method = inspect.stack()[1][3]
        # Read excel file
        df = pd.read_excel(
            self.excel_file_path,
            sheet_name=sheet_id,
            na_values=["nan", "N/A", "NA", ""],
        )
        property_column = df.columns[0]
        df = df[df[property_column].notna()]
        df = df[df[property_column].astype(str).str.strip() != ""]
        # Convert database to json format
        json_data = {}
        for row in df.itertuples(index=False):
            property_name = row[0]
            values = row[1:]
            json_data[property_name] = dict(zip(df.columns[1:], values))

        # Check json is not empty
        if len(json_data) == 0:
            self.log.error(f"{caller_method}{sheet_id}) No data found in xlsx database")
            stderr.print(
                f"{caller_method}{sheet_id}) [red]No data found in xlsx database"
            )
            sys.exit(1)

        # Perform validation of database content
        validation_out = self.validate_database_definition(json_data)
        if validation_out:
            sys.exit()
        else:
            return json_data

    def create_schema_draft_template(self):
        """
        Create a JSON Schema draft template based on the draft version.
        Available drafts: [2020-12]

        Returns:
            draft_template(dict): The JSON Schema draft template.
        """
        draft_template = (
            relecov_tools.assets.schema_utils.jsonschema_draft.create_draft(
                draft_version=self.draft_version, required_items=True
            )
        )
        return draft_template

    def _cast_example_to_type(
        self, property_id: str, expected_type: str | None, value: any
    ) -> any:
        """Cast a single example value to the declared JSON-schema type when possible."""
        if not isinstance(expected_type, str):
            return value
        expected = expected_type.strip().lower()
        if expected == "string":
            return str(value)
        if expected == "integer":
            try:
                parsed_number = float(value)
            except (TypeError, ValueError):
                self.log.warning(
                    "Example value %r for property '%s' does not match expected type 'integer'. Keeping original value.",
                    value,
                    property_id,
                )
                return value
            if not parsed_number.is_integer():
                self.log.warning(
                    "Example value %r for property '%s' does not match expected type 'integer'. Keeping original value.",
                    value,
                    property_id,
                )
                return value
            return int(parsed_number)
        if expected == "number":
            try:
                return float(value)
            except (TypeError, ValueError):
                self.log.warning(
                    "Example value %r for property '%s' does not match expected type 'number'. Keeping original value.",
                    value,
                    property_id,
                )
                return value
        if expected == "boolean":
            if isinstance(value, bool):
                return value
            if isinstance(value, str):
                normalized = value.strip().lower()
                if normalized in ("true", "1", "yes", "y"):
                    return True
                if normalized in ("false", "0", "no", "n"):
                    return False
            self.log.warning(
                "Example value %r for property '%s' does not match expected type 'boolean'. Keeping original value.",
                value,
                property_id,
            )
            return value
        return value

    def _cast_examples_to_declared_type(
        self, property_id: str, expected_type: str | None, values: list[any]
    ) -> list[any]:
        return [
            self._cast_example_to_type(property_id, expected_type, item)
            for item in values
        ]

    def jsonschema_object(
        self,
        property_id: str,
        property_feature_key: str,
        value: any,
        expected_type: str | None = None,
    ) -> dict[str, any]:
        """
        Process a property keyword with their value and return a dictionary with fields for a property.

        Args:
            property_id (str): Name of the property.
            property_feature_key (str): Property keyword.
            value (any): Property keyword value.

        Returns:
            jsonschema_value (dict): {keyword: value}, parsed for each of the options
        """

        jsonschema_value = {}
        # Match/Case statement to evaluate the key:value pairs in the database and transform them to schema-compliant dictionaries.
        match property_feature_key, value:
            case "options", str(value):
                options_list = [option.split(":") for option in value.split(",")]
                # Handling float/ints stored as str
                for key, value in options_list:
                    key = key.strip()
                    value = value.strip()
                    try:
                        value = float(value)
                        value = int(value) if value.is_integer() else value
                    except ValueError:
                        pass
                    jsonschema_value[key] = value
            # FIXME multiple examples will always be loaded as str, regardless of actual type
            case "examples", str(value):
                parsed_examples = value.split("; ")
                parsed_examples = self._cast_examples_to_declared_type(
                    property_id, expected_type, parsed_examples
                )
                jsonschema_value = {property_feature_key: parsed_examples}
            case "examples", datetime():
                value = value.strftime("%Y-%m-%dT%H:%M:%S")
                value = value.replace("T00:00:00", "")
                parsed_examples = self._cast_examples_to_declared_type(
                    property_id, expected_type, [value]
                )
                jsonschema_value = {property_feature_key: parsed_examples}
            case "examples", int(value) | float(value):
                value = float(value)
                parsed_examples = [int(value) if value.is_integer() else value]
                parsed_examples = self._cast_examples_to_declared_type(
                    property_id, expected_type, parsed_examples
                )
                jsonschema_value = {property_feature_key: parsed_examples}
            case "enum", str():
                jsonschema_value = {"$ref": f"#/$defs/enums/{property_id}"}
            case _, value if not pd.isna(value):
                # Non-serializable JSON value check and parsing (e.g. datetimes)
                try:
                    json.dumps(value)
                except (TypeError, OverflowError):
                    value = str(value)
                jsonschema_value = {property_feature_key: value}
            case _, _:
                pass

        return jsonschema_value

    def handle_properties(self, json_data: dict[str, dict]) -> tuple[dict, dict, dict]:
        """
        Handle the generation of simple and nested properties from the database definition.

        Args:
            json_data (dict): dictionary with structure {property_name: database_definition_dictionary}

        Returns:
            jsonschema_value (tuple): tuple containing the properties, required properties identified during the handling, and enums.
        """
        schema_property = {}
        required_properties = []
        definitions = {"$defs": {"enums": {}}}

        mapping_features = self.configurables.get("database_mapping_features", {})
        exclude_fields = self.configurables.get("database_exclude_features", [])
        # Flag property values that belong outside the property:
        # - is_required: if required, goes to root 'required' keyword
        # - has_enum: if there is an enum, store it for '$defs'
        for property_id, db_features_dic in json_data.items():
            is_required = db_features_dic.get("required (Y/N)", "") == "Y"
            has_enum = db_features_dic.get("enum", False)
            if property_id in [
                "collecting_institution",
                "submitting_institution",
                "sequencing_institution",
            ]:
                lab_values = self._lab_uniques.get(property_id, [])
                if lab_values:
                    has_enum = "; ".join(lab_values)

            # Create empty placeholder
            schema_property[property_id] = {}
            # If property is complex, call build schema again; else, continue function
            is_complex = db_features_dic.get("complex_field (Y/N)", "") == "Y"
            if is_complex:
                schema_draft = {"type": "object", "properties": {}, "required": []}
                subschema = self.read_database_definition(property_id)
                complex_json_feature = self.build_new_schema(
                    subschema,
                    schema_draft,
                    root_schema=False,
                )
                if complex_json_feature:
                    if complex_json_feature.get("$defs"):
                        # Prune the defs from the complex property
                        complex_defs = complex_json_feature.pop("$defs")
                        complex_defs["enums"] = {property_id: complex_defs["enums"]}
                        definitions["$defs"]["enums"].update(complex_defs["enums"])
                        # Fix the "$refs" adding the name of the parent property
                        for property_key, value in complex_json_feature[
                            "properties"
                        ].items():
                            if "$ref" in value:
                                value["$ref"] = value["$ref"].replace(
                                    f"/{property_key}", f"/{property_id}/{property_key}"
                                )
                    schema_property[property_id]["type"] = "array"
                    schema_property[property_id]["items"] = complex_json_feature
            else:
                for db_feature_key, db_feature_value in db_features_dic.items():
                    if db_feature_key in exclude_fields:
                        continue
                    # Extra check to avoid non-mapping properties.
                    if db_feature_key in mapping_features:
                        if db_feature_key == "enum" and isinstance(has_enum, str):
                            db_feature_value = has_enum
                        std_json_feature = self.jsonschema_object(
                            property_id,
                            mapping_features[db_feature_key],
                            db_feature_value,
                            expected_type=db_features_dic.get("type"),
                        )
                        if std_json_feature:
                            schema_property[property_id].update(std_json_feature)

            # If property is required, add it to list
            if is_required:
                required_properties.append(property_id)
            # If there is an enum in the property, parse it and add it to definitions
            if isinstance(has_enum, str):
                enum = self._parse_enum_values(has_enum)
                definitions["$defs"]["enums"][property_id] = {}
                definitions["$defs"]["enums"][property_id]["enum"] = enum

        # Just to be completely sure, but it should be unique
        required_properties = (
            {"required": list(set(required_properties))} if required_properties else {}
        )

        # Check that there are definitions
        definitions = definitions if definitions["$defs"]["enums"].values() else {}

        return schema_property, required_properties, definitions

    def schema_build_all_of(self, json_data: dict) -> dict:
        """
        Build the subschemas in 'allOf' keyword from the database definition.

        Args:
            json_data (dict): dictionary with structure {property_name: database_definition_dictionary}

        Returns:
            all_of_base (list): list containing all the subschemas to test in 'allOf'
        """
        all_of_base = []

        # Generate all the anyOf within
        all_any_of = []
        conditional_required = {
            key: value.get("conditional_required_group").strip()
            for key, value in json_data.items()
            if not pd.isna(value.get("conditional_required_group"))
        }
        groups = list(set(conditional_required.values()))
        conditional_required_by_group = {
            group: [
                key
                for key in conditional_required.keys()
                if conditional_required[key] == group
            ]
            for group in groups
        }
        for group, keys in conditional_required_by_group.items():
            any_of = [{"required": [key]} for key in keys]
            all_any_of.append({"anyOf": any_of})

        all_of_base.extend(all_any_of)

        # For future: generate if_then within (for required props when specific value)
        # FUTURE: all_of_base.extend(all_if_then)

        return {"allOf": all_of_base} if all_of_base else {}

    def build_new_schema(
        self,
        json_data: dict[str, dict],
        schema_draft: dict,
        root_schema: bool = True,
    ) -> dict[str, any]:
        """
        Build a new JSON Schema based on the provided JSON data and draft template, in three stages:
        - Pre-properties: all the operations needed prior to handling the properties (e.g. creation of root properties)
        - properties: handling both simple and complex properties on a separate function
        - Post-properties: All the operations needed after handling properties (e.g. defining which properties are required)

        Parameters:
        json_data (dict): Dictionary containing the properties and values of the database definition.
        schema_draft (dict): The JSON Schema draft template.
        root_schema (bool): True if is root of schema, False if not (e.g. complex property generation)

        Returns:
            schema_draft (dict): The newly created JSON Schema.
        """
        # Pre-properties
        new_schema = schema_draft
        json_data = {
            property_id: property_definition
            for property_id, property_definition in json_data.items()
            if not self._is_template_only_property(property_definition)
        }
        if root_schema:
            # Fill schema header
            package_name = relecov_tools.utils.get_package_name()
            branch_name = relecov_tools.utils.get_git_branch()
            package_path = package_name.replace("-", "_")
            new_schema["$id"] = (
                "https://github.com/BU-ISCIII/"
                f"{package_name}/blob/{branch_name}/{package_path}/schema/{self.schema_output_filename}"
            )
            new_schema["title"] = f"{self.project}-schema"
            new_schema["description"] = (
                "Json Schema that specifies the structure, content, and validation "
                f"rules for {self.project}"
            )
            new_schema["version"] = self.version

        # Fill schema properties
        # Properties
        try:
            properties, required, defs = self.handle_properties(json_data)
        except Exception as e:
            self.log.error(f"Error building properties: {str(e)}")
            stderr.print(f"[red]Error building properties: {str(e)}")
            raise e

        # Post-properties
        # Finally, send schema_property object to the new json schema draft.
        new_schema["properties"] = properties
        new_schema.update(required)
        new_schema.update(defs)

        # Build the allOf keyword
        all_of = self.schema_build_all_of(json_data)
        new_schema.update(all_of)
        # From here it can be extended to build other keywords at the end following the example above

        return new_schema

    def verify_schema(self, schema):
        """
        Verify that the given schema adheres to the JSON Schema specification for the specified draft version.

        Args:
            schema (dict): The JSON Schema to be verified.

        Raises:
            ValueError: If the schema does not conform to the JSON Schema specification.
        """
        self.validate_schema_enum_duplicates(schema)
        relecov_tools.assets.schema_utils.jsonschema_draft.check_schema_draft(
            schema, self.draft_version
        )

    @staticmethod
    def _find_duplicate_values(values: list) -> list:
        """Return duplicated values preserving first duplicate encounter order."""
        seen = set()
        duplicates = []
        duplicate_seen = set()
        for value in values:
            lookup_value = value.strip() if isinstance(value, str) else value
            try:
                is_seen = lookup_value in seen
            except TypeError:
                lookup_value = json.dumps(value, sort_keys=True, ensure_ascii=False)
                is_seen = lookup_value in seen

            if is_seen and lookup_value not in duplicate_seen:
                duplicates.append(value)
                duplicate_seen.add(lookup_value)
            else:
                seen.add(lookup_value)
        return duplicates

    def validate_schema_enum_duplicates(self, schema: dict):
        """Validate that every enum list in a generated schema has unique values."""
        duplicate_enums = {}

        def walk_schema(node, path="$"):
            if isinstance(node, dict):
                enum_values = node.get("enum")
                if isinstance(enum_values, list):
                    duplicates = self._find_duplicate_values(enum_values)
                    if duplicates:
                        duplicate_enums[path] = duplicates
                for key, value in node.items():
                    walk_schema(value, f"{path}.{key}")
            elif isinstance(node, list):
                for index, value in enumerate(node):
                    walk_schema(value, f"{path}[{index}]")

        walk_schema(schema)

        if duplicate_enums:
            df_errors = pd.DataFrame(
                [
                    {
                        "Enum Path": enum_path,
                        "Duplicate Values": "; ".join(map(str, duplicates)),
                    }
                    for enum_path, duplicates in duplicate_enums.items()
                ]
            )
            error_file_path = f"{self.output_dir}/schema_enum_duplicates.csv"
            df_errors.to_csv(error_file_path, index=False, encoding="utf-8")
            relecov_tools.utils.display_dataframe_to_user(
                name="Schema Enum Duplicates", dataframe=df_errors
            )
            stderr.print("[red]Duplicated enum values found. Log saved to:")
            stderr.print(f"\t{error_file_path}")
            raise ValueError("Duplicated enum values found in generated schema.")

    def get_schema_diff(self, base_schema, new_schema):
        """
        Print the differences between the base schema and the newly generated schema.

        Args:
            base_schema (dict): The base JSON Schema to compare against.
            new_schema (dict): The newly generated JSON Schema to compare.

        Returns:
            bool: True if differences are found, False otherwise.
        """
        # Set diff input
        base_schema_lines = json.dumps(base_schema, indent=4).splitlines()
        new_schema_lines = json.dumps(new_schema, indent=4).splitlines()

        # Get diff lines
        diff_lines = list(
            difflib.unified_diff(
                base_schema_lines,
                new_schema_lines,
                fromfile="base_schema.json",
                tofile="new_schema.json",
            )
        )

        if not diff_lines:
            self.log.info(
                "No differences were found between already installed and new generated schema. Exiting. No changes made"
            )
            stderr.print(
                "[yellow]No differences were found between already installed and new generated schema. Exiting. No changes made"
            )
            return None
        else:
            self.log.info(
                "Differences found between the existing schema and the newly generated schema."
            )
            stderr.print(
                "[yellow]Differences found between the existing schema and the newly generated schema."
            )
            if self.show_diff:
                return self.print_save_schema_diff(diff_lines)
            else:
                return None

    def print_save_schema_diff(self, diff_lines=None):
        # Set user's choices
        choices = ["Print to standard output (stdout)", "Save to file", "Both"]
        diff_output_choice = (
            "Save to file"
            if self.non_interactive
            else relecov_tools.utils.prompt_selection(
                "How would you like to print the diff between schemes?:", choices
            )
        )
        if diff_output_choice in ["Print to standard output (stdout)", "Both"]:
            for line in diff_lines:
                print(line)
            return True
        if diff_output_choice in ["Save to file", "Both"]:
            diff_filepath = os.path.join(
                os.path.realpath(self.output_dir) + "/build_schema_diff.txt"
            )
            with open(diff_filepath, "w") as diff_file:
                diff_file.write("\n".join(diff_lines))
            self.log.info(f"[green]Schema diff file saved to {diff_filepath}")
            stderr.print(f"[green]Schema diff file saved to {diff_filepath}")
            return True

    # FIXME: Add version tag to file name
    def save_new_schema(self, json_data):
        """
        Save the generated JSON Schema to the output folder.

        Args:
            json_data (dict): The JSON Schema to be saved.

        Returns:
            bool: True if the schema was successfully saved, False otherwise.
        """
        try:
            path_to_save = os.path.join(self.output_dir, self.schema_output_filename)
            with open(path_to_save, "w") as schema_file:
                json.dump(json_data, schema_file, ensure_ascii=False, indent=4)
            self.log.info(f"New JSON schema saved to: {path_to_save}")
            stderr.print(f"[green]New JSON schema saved to: {path_to_save} ")
            return True
        except PermissionError as perm_error:
            self.log.error(f"Permission error: {perm_error}")
            stderr.print(f"[red]Permission error: {perm_error}")
        except IOError as io_error:
            self.log.error(f"I/O error: {io_error}")
            stderr.print(f"[red]I/O error: {io_error}")
        except Exception as e:
            self.log.error(f"An unexpected error occurred: {str(e)}")
            stderr.print(f"[red]An unexpected error occurred: {str(e)}")
        return False

    def _sort_enum_values(self, enum_values: list[str]) -> list[str]:
        """
        Deduplicate and sort enum values alphabetically while keeping configured special terms last.

        Args:
            enum_values (list[str]): Enum values to sort.

        Returns:
            list[str]: Sorted enum values.
        """
        last_values = {
            "not sequenced": 0,
            "no enrichment": 1,
            "not applicable": 2,
            "not collected": 3,
            "not provided": 4,
            "missing": 5,
            "restricted access": 6,
            "other": 7,
            "none": 8,
        }

        unique_values = self._unique_enum_values(enum_values)

        def sort_key(value: str):
            normalized_value = value.strip().casefold()
            for last_value, last_order in last_values.items():
                if last_value in normalized_value:
                    return (1, last_order)
            return (0, normalized_value)

        return sorted(unique_values, key=sort_key)

    @staticmethod
    def _unique_enum_values(enum_values: list) -> list:
        """Return enum values without duplicates, preserving first occurrence order."""
        unique_values = []
        seen = set()
        for value in enum_values:
            lookup_value = value.strip() if isinstance(value, str) else value
            try:
                is_seen = lookup_value in seen
            except TypeError:
                lookup_value = json.dumps(value, sort_keys=True, ensure_ascii=False)
                is_seen = lookup_value in seen
            if not is_seen:
                unique_values.append(value)
                seen.add(lookup_value)
        return unique_values

    def _parse_enum_values(self, enum_value):
        """
        Resolve enum definitions from lists, Excel sheet references, txt files, or cells.

        Supported references:
        - @sheet:sheet_name:column_name
        - @sheet:sheet_name:column_name:filter_column=filter_value
        - @file:file_name.txt or @file_name.txt

        Args:
            enum_value: Raw enum value from the input Excel or an already parsed list.

        Returns:
            list[str] | pandas.NA: Parsed and sorted enum values, or pd.NA if empty.
        """
        if isinstance(enum_value, list):
            return self._sort_enum_values(enum_value)
        if isinstance(enum_value, str):
            loaded_values = self._load_enum_values_from_sheet(enum_value)
            if loaded_values or enum_value.strip().startswith("@sheet:"):
                return self._sort_enum_values(loaded_values)
            loaded_values = self._load_enum_values_from_file(enum_value)
            if loaded_values or enum_value.strip().startswith(("@file:", "@")):
                return self._sort_enum_values(loaded_values)
            return self._sort_enum_values(
                [value.strip() for value in enum_value.split("; ") if value.strip()]
            )
        return pd.NA

    def _load_enum_source_sheet(self, sheet_name: str) -> pd.DataFrame:
        """
        Read and cache an enum source sheet from the input Excel file.

        Args:
            sheet_name (str): Name of the Excel sheet to load.

        Returns:
            pd.DataFrame: Sheet contents, or an empty DataFrame if the sheet is absent.
        """
        if sheet_name not in self._enum_sheet_cache:
            try:
                df = pd.read_excel(
                    self.excel_file_path,
                    sheet_name=sheet_name,
                    na_values=["nan", "N/A", "NA", ""],
                )
            except ValueError:
                self._enum_sheet_cache[sheet_name] = pd.DataFrame()
            else:
                df.columns = [str(column).strip() for column in df.columns]
                self._enum_sheet_cache[sheet_name] = df

        return self._enum_sheet_cache[sheet_name]

    def _filter_enum_source_sheet(
        self, df: pd.DataFrame, filter_references: list[str]
    ) -> pd.DataFrame:
        """
        Apply case-insensitive column=value filters to an enum source sheet.

        For example, the reference
        @sheet:amr_gene_list:Name:Category=allele
        loads the Name column from the amr_gene_list sheet, keeping rows where Category is allele.

        Args:
            df (pd.DataFrame): Enum source sheet contents.
            filter_references (list[str]): Filters from @sheet references.

        Returns:
            pd.DataFrame: Filtered data, or an empty DataFrame if a filter is invalid.
        """
        for filter_reference in filter_references:
            if "=" not in filter_reference:
                return pd.DataFrame()
            column_name, expected_value = [
                item.strip() for item in filter_reference.split("=", maxsplit=1)
            ]
            if not column_name or column_name not in df.columns:
                return pd.DataFrame()
            df = df[
                df[column_name].astype(str).str.strip().str.casefold()
                == expected_value.casefold()
            ]
        return df

    def _load_enum_values_from_sheet(self, enum_value: str) -> list[str]:
        """
        Load enum values from an Excel sheet reference in the enum column.

        Args:
            enum_value (str): Reference using @sheet:sheet_name:column_name syntax.

        Returns:
            list[str]: Unique enum values from the referenced sheet column.
        """
        enum_reference = enum_value.strip()
        if not enum_reference.startswith("@sheet:"):
            return []

        reference_parts = [
            part.strip() for part in enum_reference.removeprefix("@sheet:").split(":")
        ]
        if len(reference_parts) < 2:
            return []

        sheet_name, column_name = reference_parts[:2]
        if not sheet_name or not column_name:
            return []

        df = self._load_enum_source_sheet(sheet_name)
        df = self._filter_enum_source_sheet(df, reference_parts[2:])
        if df.empty or column_name not in df.columns:
            return []

        values = [
            str(value).strip()
            for value in df[column_name].dropna().tolist()
            if str(value).strip()
        ]
        return values

    def _load_enum_values_from_file(self, enum_value: str) -> list[str]:
        """
        Load enum values from a txt file in relecov_tools/conf.

        Args:
            enum_value (str): File reference using @file:file_name.txt or @file_name.txt.

        Returns:
            list[str]: Enum values read from the file, ignoring blank/comment lines.
        """
        enum_file = enum_value.strip()
        if enum_file.startswith("@file:"):
            enum_file = enum_file.removeprefix("@file:").strip()
        elif enum_file.startswith("@"):
            enum_file = enum_file.removeprefix("@").strip()

        if not enum_file.endswith(".txt"):
            return []

        enum_path = os.path.join(os.path.dirname(__file__), "conf", enum_file)
        if not os.path.exists(enum_path):
            return []

        with open(enum_path, encoding="utf-8") as fh:
            return [
                line.strip()
                for line in fh
                if line.strip() and not line.lstrip().startswith("#")
            ]

    def _template_only_properties_to_df(self, database_definition: dict | None):
        """
        Build metadata template rows for properties marked as ONLY.

        Properties with submitting_lab_form set to ONLY are excluded from the JSON
        schema, so they need to be converted directly from the input Excel definition
        into template rows.

        Args:
            database_definition (dict | None): Full database definition from the input Excel.

        Returns:
            pd.DataFrame: Template rows for ONLY properties.
        """
        if not database_definition:
            return pd.DataFrame()

        mapping_features = self.configurables.get("database_mapping_features", {})
        template_rows = []
        for property_id in database_definition:
            db_features = database_definition[property_id]
            if not self._is_template_only_property(db_features):
                continue

            row = {
                "property_id": property_id,
                "field_id": property_id,
                "parent_property_id": None,
                "parent_label": "",
                "parent_classification": "",
                "is_required": db_features.get("required (Y/N)", "") == "Y",
            }
            for db_feature_key, db_feature_value in db_features.items():
                schema_key = mapping_features.get(db_feature_key)
                if not schema_key:
                    continue
                if schema_key == "enum":
                    row["enum"] = self._parse_enum_values(db_feature_value)
                    continue
                std_json_feature = self.jsonschema_object(
                    property_id,
                    schema_key,
                    db_feature_value,
                    expected_type=db_features.get("type"),
                )
                row.update(std_json_feature)

            template_rows.append(row)

        return pd.DataFrame(template_rows)

    @staticmethod
    def _format_template_required_value(value):
        """Return the visible required label used in the metadata template."""
        required_value = str(value or "").strip()
        if required_value.upper() == "Y":
            return "YES"
        if required_value.upper() in ["N", "NO"]:
            return "NO"
        if required_value.lower().startswith("y if "):
            condition = required_value[5:].strip()
            if condition.lower() == "sequenced":
                condition = "sequenced"
            return f"YES if {condition}"
        return required_value

    def create_metadatalab_excel(self, json_schema, database_definition=None):
        """
        Generates an Excel template file for Metadata LAB with four sheets:
        Overview, Metadata LAB, Data Validation, and Version History.

        Args:
            json_schema (dict): The JSON schema used to generate the template.
                                It should include properties and required fields.

        Returns:
            None: If an error occurs during the process.
        """
        try:
            default_note = (
                "Initial version" if self.initial_version else "Auto-generated update"
            )
            notes_control_input = (
                default_note
                if self.non_interactive
                else input(
                    "\033[93mEnter a note about changes made to the schema: \033[0m"
                )
                or default_note
            )

            # ------------------------------------------------------------------ #
            # 1.  Versioning & paths
            # ------------------------------------------------------------------ #
            next_version = self.version
            version_info = {
                "FILE_VERSION": f"{self.template_output_prefix}_v{next_version}",
                "CODE": next_version,
                "NOTES CONTROL": notes_control_input,
                "DATE": datetime.now().strftime("%Y-%m-%d"),
            }

            if self.initial_version:
                version_history = pd.DataFrame([version_info])
            else:
                try:
                    wb = openpyxl.load_workbook(self.excel_template)
                    ws_version = wb["VERSION"]
                    data = ws_version.values
                    columns = next(data)
                    version_history = pd.DataFrame(data, columns=columns)
                except Exception as e:
                    msg = (
                        f"Error reading previous VERSION sheet from "
                        f"{self.excel_template}: {e}"
                    )
                    self.log.error(msg)
                    stderr.print(f"[red]{msg}")
                    raise ValueError(msg)

                version_history = pd.concat(
                    [version_history, pd.DataFrame([version_info])],
                    ignore_index=True,
                )
            out_file = os.path.join(
                self.output_dir, f"{self.template_output_prefix}_v{next_version}.xlsx"
            )

            # ------------------------------------------------------------------ #
            # 2.  Schema filtering and dataframe preparation
            # ------------------------------------------------------------------ #
            default_classification_filter = [
                "Database Identifiers",
                "Sample collection and processing",
                "Host information",
                "Sequencing",
                "Pathogen diagnostic testing",
                "Contributor Acknowledgement",
                "Public databases",
                "Bioinformatics and QC metrics fields",
            ]
            required_properties = set(json_schema.get("required", []))
            schema_properties = json_schema.get("properties")
            enum_defs = json_schema.get("$defs", {}).get("enums", {})

            try:
                schema_properties_flatten = relecov_tools.assets.schema_utils.metadatalab_template.schema_to_flatten_json(
                    schema_properties, required_properties
                )
                df = relecov_tools.assets.schema_utils.metadatalab_template.schema_properties_to_df(
                    schema_properties_flatten
                )
                template_only_df = self._template_only_properties_to_df(
                    database_definition
                )
                if not template_only_df.empty:
                    df = pd.concat([df, template_only_df], ignore_index=True)

                classification_overrides = self.configurables.get(
                    "classification_filters", {}
                )
                classification_filter = classification_overrides.get(
                    self.project, default_classification_filter
                )
                if classification_filter and "classification" in df.columns:
                    df = df[df["classification"].isin(classification_filter)]
                if "is_required" in df.columns:
                    df["required"] = df["is_required"].apply(
                        lambda value: "Y" if bool(value) else "N"
                    )
                else:
                    df["required"] = df["property_id"].apply(
                        lambda x: "Y" if x in required_properties else "N"
                    )
                if database_definition:
                    required_values = {
                        property_id: self._format_template_required_value(
                            features.get("required (Y/N)")
                        )
                        for property_id, features in database_definition.items()
                    }
                    df["required"] = df.apply(
                        lambda row: required_values.get(
                            row["property_id"], row["required"]
                        )
                        or row["required"],
                        axis=1,
                    )
                df["required"] = df["required"].apply(
                    self._format_template_required_value
                )

                def resolve_enum_ref(ref: str, enum_defs: dict) -> list[str]:
                    property_key = ref.split("enums/")[-1]
                    property_id = property_key.split("/")
                    try:
                        values = enum_defs
                        for property_node in property_id:
                            values = values[property_node]
                        values = values["enum"]
                    except KeyError:
                        self.log.error(
                            f"Error finding enum for property '{'.'.join(property_id)}'; not found in $defs"
                        )
                        stderr.print(
                            f"[red]Error finding enum for property '{'.'.join(property_id)}'; not found in $defs"
                        )
                        return []
                    return self._clean_template_enum_values(values)

                resolved_enums = df["$ref"].apply(
                    lambda row: (
                        resolve_enum_ref(row, enum_defs=enum_defs)
                        if not pd.isna(row)
                        else row
                    )
                )
                if "enum" in df.columns:
                    df["enum"] = df["enum"].where(
                        pd.notnull(df["enum"]), resolved_enums
                    )
                else:
                    df["enum"] = resolved_enums

                df["enum"] = df["enum"].apply(self._clean_template_enum_values)
                common_dropdown = self._lab_dropdowns["collecting_institution"]

                lab_fields = [
                    "collecting_institution",
                    "submitting_institution",
                    "sequencing_institution",
                ]

                mask = df["property_id"].isin(lab_fields)

                df.loc[mask, "enum"] = pd.Series(
                    [common_dropdown] * mask.sum(), index=df.loc[mask].index
                )

            except Exception as e:
                self.log.error(f"Error processing schema properties: {e}")
                stderr.print(f"Error processing schema properties: {e}")
                return None

            # ------------------------------------------------------------------ #
            # 3.  Headers / filtering
            # ------------------------------------------------------------------ #
            if "header" in df.columns:
                df["header"] = df["header"].astype(str).str.strip().str.upper()
                df_filtered = df[df["header"].isin(["Y", "ONLY"])]
            else:
                self.log.warning(
                    "No se encontró la columna 'header', usando df sin filtrar."
                )
                df_filtered = df

            # ------------------------------------------------------------------ #
            # 4.  Construction of OVERVIEW, METADATA_LAB and DATA_VALIDATION
            # ------------------------------------------------------------------ #
            # -- OVERVIEW
            try:
                overview_header = [
                    "Label name",
                    "Description",
                    "Group",
                    "Mandatory (Y/N)",
                    "Example",
                ]
                df_overview = pd.DataFrame(columns=overview_header)
                df_overview["Label name"] = df_filtered["label"]
                df_overview["Description"] = df_filtered["description"]
                df_overview["Group"] = df_filtered["classification"]
                df_overview["Mandatory (Y/N)"] = df_filtered["required"]
                df_overview["Example"] = df_filtered["examples"].apply(
                    lambda x: x[0] if isinstance(x, list) else x
                )
            except Exception as e:
                self.log.error(f"Error creating overview sheet: {e}")
                stderr.print(f"Error creating overview sheet: {e}")
                return None

            # -- METADATA_LAB
            try:
                metadatalab_header = ["CAMPO", "DESCRIPCIÓN", "EJEMPLOS", "REQUERIDO"]
                df_metadata = pd.DataFrame(columns=metadatalab_header)
                df_metadata["REQUERIDO"] = df_filtered["required"].apply(
                    lambda x: (
                        "YES"
                        if str(x).upper() in ["Y", "YES"]
                        else "" if str(x).upper() in ["N", "NO"] else x
                    )
                )
                df_metadata["EJEMPLOS"] = df_filtered["examples"].apply(
                    lambda x: x[0] if isinstance(x, list) else x
                )
                df_metadata["DESCRIPCIÓN"] = df_filtered["description"]
                df_metadata["CAMPO"] = df_filtered["label"]
                df_metadata = df_metadata.transpose()
            except Exception as e:
                self.log.error(f"Error creating MetadataLab sheet: {e}")
                stderr.print(f"[red]Error creating MetadataLab sheet: {e}")
                return None

            # -- DATA_VALIDATION
            try:
                datavalidation_header = ["CAMPO", "DESCRIPCIÓN", "EJEMPLOS"]
                df_hasenum = df[pd.notnull(df.enum)]
                df_hasenum = df_hasenum[df_hasenum["label"].isin(df_filtered["label"])]
                df_validation = pd.DataFrame(columns=datavalidation_header)
                df_validation["tmp_property"] = df_hasenum["property_id"]
                df_validation["EJEMPLOS"] = df_hasenum["examples"].apply(
                    lambda x: x[0] if isinstance(x, list) else x
                )
                df_validation["DESCRIPCIÓN"] = df_hasenum["description"]
                df_validation["CAMPO"] = df_hasenum["label"]
            except Exception as e:
                self.log.error(f"Error creating DataValidation sheet: {e}")
                stderr.print(f"[red]Error creating DataValidation sheet: {e}")
                return None

            # ------------------------------------------------------------------ #
            # 5.  Prepare DATA_VALIDATION
            # ------------------------------------------------------------------ #
            try:
                enum_dict = {prop: [] for prop in df_hasenum["property_id"]}
                enum_maxitems = 0
                for key in enum_dict.keys():
                    enum_values = df_hasenum[df_hasenum["property_id"] == key][
                        "enum"
                    ].values
                    if enum_values.size > 0:
                        enum_list = enum_values[0]
                        enum_dict[key] = enum_list
                        enum_maxitems = max(enum_maxitems, len(enum_list))
                    else:
                        enum_dict[key] = []

                for key in enum_dict.keys():
                    enum_dict[key].extend([""] * (enum_maxitems - len(enum_dict[key])))

                new_df = pd.DataFrame(enum_dict)
                valid_index = df_validation["tmp_property"].values
                valid_transposed = df_validation.transpose()
                valid_transposed.columns = valid_index
                df_validation = pd.concat([valid_transposed, new_df]).drop(
                    index=["tmp_property"]
                )
            except Exception as e:
                self.log.error(f"Error processing enums and combining data: {e}")
                stderr.print(f"[red]Error processing enums and combining data: {e}")
                return None

            # ------------------------------------------------------------------ #
            # 6.  Cleaning of NaN's
            # ------------------------------------------------------------------ #
            df_overview = (
                df_overview.replace([float("inf"), float("-inf")], "")
                .fillna("")
                .infer_objects()
            )
            df_metadata = (
                df_metadata.replace([float("inf"), float("-inf")], "")
                .fillna("")
                .infer_objects()
            )
            df_validation = (
                df_validation.replace([float("inf"), float("-inf")], "")
                .fillna("")
                .infer_objects()
            )

            # ------------------------------------------------------------------ #
            # 7.  Initial writing of the XLSX
            # ------------------------------------------------------------------ #
            try:
                writer = pd.ExcelWriter(out_file, engine="xlsxwriter")
                mt = relecov_tools.assets.schema_utils.metadatalab_template

                mt.excel_formater(
                    df_overview,
                    writer,
                    "OVERVIEW",
                    out_file,
                    have_index=False,
                    have_header=False,
                )
                mt.excel_formater(
                    df_metadata,
                    writer,
                    "METADATA_LAB",
                    out_file,
                    have_index=True,
                    have_header=False,
                )
                mt.excel_formater(
                    df_validation,
                    writer,
                    "DATA_VALIDATION",
                    out_file,
                    have_index=True,
                    have_header=False,
                )
                version_history.to_excel(writer, sheet_name="VERSION", index=False)
                writer.close()

                self.log.info(
                    f"Metadata lab template successfuly created in: {out_file}"
                )
                stderr.print(
                    f"[green]Metadata lab template successfuly created in: {out_file}"
                )
            except Exception as e:
                self.log.error(f"Error writing to Excel: {e}")
                stderr.print(f"[red]Error writing to Excel: {e}")
                return None

            # ------------------------------------------------------------------ #
            # 8.  OpenPyXL post-processing (conditions, dropdowns, etc.)
            # ------------------------------------------------------------------ #
            try:
                wb = openpyxl.load_workbook(out_file)
                ws_metadata = wb["METADATA_LAB"]
                ws_metadata.freeze_panes = self.configurables.get("freeze_panel", "D1")
                ws_metadata.delete_rows(5)
                mt.format_metadata_data_entry_area(
                    ws_metadata, end_col=len(df_filtered) + 1
                )
                mt.create_condition(ws_metadata, self.project_config, df_filtered)
                mt.add_duplicate_value_formatting(ws_metadata)
                mt.add_conditional_format_age_check(ws_metadata, df_filtered)

                # Hidden sheet for dropdowns
                ws_dropdowns = (
                    wb.create_sheet("DROPDOWNS")
                    if "DROPDOWNS" not in wb.sheetnames
                    else wb["DROPDOWNS"]
                )
                for row in ws_dropdowns.iter_rows():
                    for cell in row:
                        cell.value = None

                # Dynamic lists for the three special fieldss
                common_dropdown = self._lab_dropdowns["collecting_institution"]
                special_dropdowns = {
                    "collecting_institution": common_dropdown,
                    "submitting_institution": common_dropdown,
                    "sequencing_institution": common_dropdown,
                }
                # ------------------------------------------------------------------------------

                # We scroll through the columns of METADATA_LAB (original order of df)
                column = 1
                for col_idx, property_id in enumerate(
                    df_filtered["property_id"], start=1
                ):
                    if property_id not in df_hasenum["property_id"].values:
                        continue
                    # Select list of values
                    if property_id in special_dropdowns:
                        enum_values = special_dropdowns[property_id]
                    else:
                        enum_values = df.loc[
                            df["property_id"] == property_id, "enum"
                        ].values[0]

                    if not isinstance(enum_values, list) or len(enum_values) == 0:
                        continue
                    column += 1
                    # Write on sheet DROPDOWNS
                    col_letter = openpyxl.utils.get_column_letter(column)
                    for i, val in enumerate(enum_values, start=1):
                        ws_dropdowns[f"{col_letter}{i}"].value = val

                    # Create validation in METADATA_LAB
                    dv_range = (
                        f"DROPDOWNS!${col_letter}$1:${col_letter}${len(enum_values)}"
                    )
                    meta_col = ws_metadata.cell(row=4, column=col_idx + 1).column_letter
                    meta_rng = f"{meta_col}5:{meta_col}1000"

                    dv = DataValidation(
                        type="list",
                        formula1=dv_range,
                        allow_blank=False,
                        showErrorMessage=True,
                    )
                    dv.error = "Valor no permitido. Elija un elemento de la lista."
                    dv.errorTitle = "Error de validación"
                    dv.prompt = f"Seleccione un valor para {property_id}"
                    dv.promptTitle = "Selección de valor"

                    ws_metadata.add_data_validation(dv)
                    dv.add(meta_rng)

                # ------------------------------------------------------------------ #
                # 9.  Visual adjustments (widths, text-wrap, hide sheet)
                # ------------------------------------------------------------------ #
                if "OVERVIEW" in wb.sheetnames:
                    ws_overview = wb["OVERVIEW"]
                    column_width = 35
                    for col in ws_overview.columns:
                        max_length = max(
                            len(str(cell.value)) for cell in col if cell.value
                        )
                        ws_overview.column_dimensions[col[0].column_letter].width = min(
                            max_length + 2, column_width
                        )
                    for row in ws_overview.iter_rows():
                        for cell in row:
                            cell.alignment = openpyxl.styles.Alignment(wrap_text=True)

                    ws_version = wb["VERSION"]
                    for i, col in enumerate(ws_version.columns, start=1):
                        max_len = max(len(str(c.value)) for c in col if c.value)
                        ws_version.column_dimensions[
                            openpyxl.utils.get_column_letter(i)
                        ].width = (max_len + 2)

                ws_dropdowns.sheet_state = "hidden"
                wb.save(out_file)
            except Exception as e:
                self.log.error(f"Error adding dropdowns: {e}")
                stderr.print(f"[red]Error adding dropdowns: {e}")
                return None

        except ValueError:
            raise
        except Exception as e:
            self.log.error(f"Error in create_metadatalab_excel: {e}")
            stderr.print(f"[red]Error in create_metadatalab_excel: {e}")
            return None

    def summarize_schema(self, json_schema):
        """
        Generate summary statistics for a JSON Schema and display it in tabular format.

        Args:
            json_schema (dict): The JSON Schema to analyze.

        Returns:
            None: Displays the table directly to stdout.
        """
        properties = json_schema.get("properties", {})
        # Initialize counters
        total_properties = len(properties)
        type_counts = {}
        enum_count = 0
        free_text_count = 0

        # Iterate over properties
        for _, prop_details in properties.items():
            prop_type = prop_details.get("type", "unknown")

            # Count types
            type_counts[prop_type] = type_counts.get(prop_type, 0) + 1

            # Count enum vs free text
            if "enum" in prop_details:
                enum_count += 1
            else:
                free_text_count += 1

        # Prepare summary data
        summary_data = {
            "Total Properties": [total_properties],
            "String Properties": [type_counts.get("string", 0)],
            "Integer Properties": [type_counts.get("integer", 0)],
            "Number Properties": [type_counts.get("number", 0)],
            "Boolean Properties": [type_counts.get("boolean", 0)],
            "Object Properties": [type_counts.get("object", 0)],
            "Array Properties": [type_counts.get("array", 0)],
            "Enum Properties": [enum_count],
            "Free Text Properties": [free_text_count],
        }
        summary_df = pd.DataFrame(summary_data)

        # Display summary using rich table (if available) or print raw
        try:
            relecov_tools.utils.display_dataframe_to_user(
                name="JSON Schema Summary", dataframe=summary_df
            )
        except AttributeError:
            print(summary_df.to_string(index=False))

        def shorten_path(path, max_length=50):
            """Shortens long paths by keeping first and last segments while adding ellipsis in the middle."""
            if len(path) <= max_length:
                return path  # No need to shorten

            parts = path.split(os.sep)  # Split into parts
            if len(parts) > 3:
                return os.sep.join(
                    [parts[0], "..."] + parts[-2:]
                )  # Keep first, last two, and replace middle
            return path  # If it's already short, return as is

        # Folder containing results
        outdir_data = {
            "Description": [
                "Output Folder",
                "New JSON Schema",
                "Old JSON Schema",
                "Schema Diff File",
                "Metadata Template File",
            ],
            "Path": [
                shorten_path(self.output_dir),
                shorten_path(
                    os.path.join(self.output_dir, self.schema_output_filename)
                ),
                shorten_path(self.base_schema_path),
                shorten_path(f"{self.output_dir}/build_schema_diff.txt"),
                shorten_path(
                    os.path.join(
                        self.output_dir, f"{self.template_output_prefix}_v*.xlsx"
                    )
                ),
            ],
        }

        # Convert to DataFrame
        outdir_df = pd.DataFrame(outdir_data)

        # Display summary using rich table or print raw if unavailable
        try:
            relecov_tools.utils.display_dataframe_to_user(
                name="JSON Results Overview", dataframe=outdir_df
            )
        except AttributeError:
            print(outdir_df.to_string(index=False))

    def handle_build_schema(self):
        # Load xlsx database and convert into json format
        self.log.info("Start reading xlsx database")
        stderr.print("[white]Start reading xlsx database")
        database_dic = self.read_database_definition()

        # Verify current schema used as base for this project:
        base_schema_json = relecov_tools.utils.read_json_file(self.base_schema_path)
        if not base_schema_json:
            self.log.error(f"Couldn't find {self.project} base schema.)")
            stderr.print(f"[red]Couldn't find {self.project} base schema. Exiting...)")
            sys.exit(1)

        # Create schema draft template (leave empty to be prompted to list of available schema versions)
        schema_draft_template = self.create_schema_draft_template()

        # build new schema draft based on database definition.
        new_schema_json = self.build_new_schema(database_dic, schema_draft_template)

        # Verify new schema follows json schema specification rules.
        self.verify_schema(new_schema_json)

        # Compare base vs new schema and print/saves differences (--diff = True)
        self.get_schema_diff(base_schema_json, new_schema_json)

        # Saves new JSON schema
        self.save_new_schema(new_schema_json)
        self._save_amr_genes_json()

        # Create metadata lab template
        if self.non_interactive or relecov_tools.utils.prompt_yn_question(
            "Do you want to create a metadata lab file?:"
        ):
            self.create_metadatalab_excel(new_schema_json, database_dic)

        # Return new schema
        return new_schema_json
