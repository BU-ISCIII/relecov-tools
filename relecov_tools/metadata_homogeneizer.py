#!/usr/bin/env python
import os
import sys
import logging
import rich.console
import openpyxl
from itertools import islice
import relecov_tools.utils
from relecov_tools.config_json import ConfigJson

log = logging.getLogger(__name__)
stderr = rich.console.Console(
    stderr=True,
    style="dim",
    highlight=False,
    force_terminal=relecov_tools.utils.rich_force_colors(),
)


class MetadataHomogeneizer:
    """MetadataHomogeneizer object"""

    def __init__(self, institution=None, directory=None, output_folder=None):
        self.config_json = ConfigJson()
        self.heading = self.config_json.get_configuration("metadata_lab_heading")

        if institution is None:
            self.institution = relecov_tools.utils.prompt_selection(
                msg="Select the available mapping institution",
                choices=["isciii", "hugtip", "hunsc-iter"],
            )
        else:
            self.institution = institution.upper()
        mapping_json_file = os.path.join(
            os.path.dirname(__file__),
            "schema",
            "institution_schemas",
            self.config_json.get_topic_data("mapping_file", self.institution),
        )
        if directory is None:
            directory = relecov_tools.utils.prompt_path(
                msg="Select the directory which contains additional files for metadata"
            )
        if not os.path.exists(directory):
            log.error("Folder for additional files %s does not exist ", directory)
            stderr.print(
                "[red] Folder for additional files " + directory + " does not exist"
            )
            sys.exit(1)
        self.mapping_json_data = relecov_tools.utils.read_json_file(mapping_json_file)

        try:
            lab_metadata = self.mapping_json_data["required_files"]["meta_file"][
                "file_name"
            ]
        except KeyError:
            log.error("Metadata File is not defined in schema")
            stderr.print("[red] Metadata File is not defined in schema")
            sys.exit(1)
        metadata_path = os.path.join(directory, lab_metadata)
        if not os.path.isfile(metadata_path):
            log.error("Metadata File %s does not exists", metadata_path)
            stderr.print("[red] Metadata File " + metadata_path + "does not exists")
            sys.exit(1)
        self.lab_metadata = metadata_path
        # Check if python file is defined
        function_file = self.mapping_json_data["pyhon_file"]
        if function_file == "":
            self.function_file = None
        else:
            self.function_file = os.path.join(
                os.path.dirname(__file__), "institution_scripts", function_file
            )
            if not os.path.isfile(self.function_file):
                log.error("File with functions %s does not exist ", self.function_file)
                stderr.print(
                    "[red] File with functions "
                    + self.function_file
                    + " does not exist"
                )
                sys.exit(1)
        self.additional_files = []
        if len(self.mapping_json_data["required_files"]) > 1:
            for key, values in self.mapping_json_data["required_files"].items():
                if key == "meta_file":
                    continue
                if values["file_name"] == "":
                    self.additional_files.append(values)
                    continue
                f_path = os.path.join(directory, values["file_name"])
                if not os.path.isfile(f_path):
                    log.error("Additional file %s does not exist ", f_path)
                    stderr.print("[red] Additional file " + f_path + " does not exist")
                    sys.exit(1)
                values["file_name"] = f_path
                self.additional_files.append(values)
        if output_folder is None:
            self.output_folder = relecov_tools.utils.prompt_path(
                msg="Select the output folder"
            )
        else:
            self.output_folder = output_folder

    def read_metadata_file(self):
        """Read the input metadata file"""
        wb_file = openpyxl.load_workbook(self.lab_metadata, data_only=True)
        ws_metadata_lab = wb_file["Sheet"]
        heading = [i.value.strip() for i in ws_metadata_lab[1] if i.value]
        ws_data = []
        for row in islice(ws_metadata_lab.values, 1, ws_metadata_lab.max_row):
            l_row = list(row)
            data_row = {}
            # Ignore the empty rows
            # guessing that row 1 and 2 with no data are empty rows
            if l_row[0] is None and l_row[1] is None:
                continue
            for idx in range(0, len(heading)):
                data_row[heading[idx]] = l_row[idx]
            ws_data.append(data_row)
        # import pdb; pdb.set_trace()
        return ws_data

    def mapping_metadata(self, ws_data):
        map_fields = self.mapping_json_data["required_files"]["meta_file"][
            "mapped_fields"
        ]
        map_data = []
        for row in ws_data:
            row_data = {}
            for dest_map, orig_map in map_fields.items():

                row_data[dest_map] = row[orig_map]
            map_data.append(row_data)

        return map_data

    def additional_fields(self, mapped_data):
        add_data = [self.heading]
        fixed_fields = self.mapping_json_data["required_files"]["meta_file"][
            "fixed_fields"
        ]
        for row in mapped_data:
            new_row_data = []
            for field in self.heading:
                if field in row:
                    data = row[field]
                elif field in fixed_fields:
                    data = fixed_fields[field]
                else:
                    data = ""
                new_row_data.append(data)
            add_data.append(new_row_data)
        return add_data

    def handling_additional_files(self, additional_data):

        for additional_file in self.additional_files:
            if additional_file["file_name"] != "":
                f_name = additional_file["file_name"]
                stderr.print("[blue] Start processing additional file " + f_name)
                if f_name.endswith(".json"):
                    data = relecov_tools.utils.read_json_file(f_name)
                elif f_name.endswith(".tsv"):
                    data = relecov_tools.utils.read_csv_file_return_dict(f_name, "\t")
                elif f_name.endswith(".csv"):
                    data = relecov_tools.utils.read_csv_file_return_dict(f_name, ",")
                else:
                    log.error("Additional file extension %s is not supported ", f_name)
                    stderr.print(
                        "[red] Additional file extension "
                        + f_name
                        + " is not supported"
                    )
                    sys.exit(1)
                sample_idx = self.heading.index("Sample ID given for sequencing")
            else:
                data = ""

            if additional_file["function"] == "None":
                for row in additional_data[1:]:
                    # new_row_data = []
                    s_value = str(row[sample_idx])
                    try:
                        item_data = data[s_value]
                    except KeyError:
                        pass
                        """
                        log.error(
                            "Additional file %s does not have the information for %s ",
                            f_name,
                            s_value,
                        )
                        stderr.print(
                            "[red] Additional file "
                            + f_name
                            + " does not have information for "
                            + str(s_value)
                        )
                        continue
                        """
                        # sys.exit(1)
                    for m_field, f_field in additional_file["mapped_fields"].items():
                        try:
                            meta_idx = self.heading.index(m_field)
                        except ValueError as e:
                            log.error("Field %s does not exist in Metadata ", e)
                            stderr.print(f"[red] Field {e} does not exist")
                            sys.exit(1)
                        row[meta_idx] = item_data[f_field]

                # import pdb; pdb.set_trace()
            else:
                func_name = additional_file["function"]
                stderr.print("[yellow] Start processing function " + func_name)
                exec(
                    "from relecov_tools.institution_scripts."
                    + self.institution
                    + " import "
                    + func_name
                )
                eval(
                    func_name
                    + "(additional_data, data, additional_file['mapped_fields'], self.heading)"
                )

        stderr.print("[green] Succesful processing of additional file ")
        return additional_data

    def write_to_excel_file(self, data, f_name):
        book = openpyxl.Workbook()
        sheet = book.active
        for row in data:
            sheet.append(row)
        sheet.title = "METADATA_LAB"
        book.save(f_name)
        return

    def converting_metadata(self):
        stderr.print("[blue] Reading the metadata file to convert")
        ws_data = self.read_metadata_file()
        mapped_data = self.mapping_metadata(ws_data)
        stderr.print("[green] Successful conversion mapping to ISCIII metadata")
        stderr.print("[blue] Adding fixed information")
        additional_data = self.additional_fields(mapped_data)
        # Fetch the additional files and include the information in metadata
        stderr.print("[blue] reading and mapping de information that cames in files")
        converted_data = self.handling_additional_files(additional_data)
        f_name = os.path.join(self.output_folder, "converted_metadata_lab.xlsx")
        self.write_to_excel_file(converted_data, f_name)
        return
