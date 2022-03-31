#!/usr/bin/env python
from itertools import islice

# from geopy.geocoders import Nominatim
import json
import logging
import rich.console

# from openpyxl import Workbook
import openpyxl
import os
import sys
import relecov_tools.utils
from relecov_tools.config_json import ConfigJson
import relecov_tools.json_schema

log = logging.getLogger(__name__)
stderr = rich.console.Console(
    stderr=True,
    style="dim",
    highlight=False,
    force_terminal=relecov_tools.utils.rich_force_colors(),
)


class RelecovMetadata:
    def __init__(self, metadata_file=None, sample_list_file=None, output_folder=None):
        if metadata_file is None:
            self.metadata_file = relecov_tools.utils.prompt_path(
                msg="Select the excel file which contains metadata"
            )
        else:
            self.metadata_file = metadata_file
        if not os.path.exists(self.metadata_file):
            log.error("Metadata file %s does not exist ", self.metadata_file)
            stderr.print("[red] Metadata file " + self.meta_file + " does not exist")
            sys.exit(1)
        if sample_list_file is None:
            self.sample_list_file = relecov_tools.utils.prompt_path(
                msg="Select the file which contains the sample information"
            )
        else:
            self.sample_list_file = sample_list_file
        if not os.path.exists(self.sample_list_file):
            log.error("Sample information file %s does not exist ", self.metadata_file)
            stderr.print(
                "[red] Sample information " + self.sample_list_file + " does not exist"
            )
            sys.exit(1)
        if output_folder is None:
            self.output_folder = relecov_tools.utils.prompt_path(
                msg="Select the output folder"
            )
        else:
            self.output_folder = output_folder
        config_json = ConfigJson()
        relecov_schema = config_json.get_topic_data("json_schemas", "relecov_schema")
        relecov_sch_path = os.path.join(
            os.path.dirname(os.path.realpath(__file__)), "schema", relecov_schema
        )
        with open(relecov_sch_path, "r") as fh:
            self.relecov_sch_json = json.load(fh)
        self.label_prop_dict = {}
        for prop, values in self.relecov_sch_json["properties"].items():
            try:
                self.label_prop_dict[values["label"]] = prop
            except KeyError:
                continue

    def check_new_metadata(folder):
        """Check if there is a new metadata to be processed
        folder  Directory to be checked
        """
        pass

    def fetch_metadata_file(folder, file_name):
        """Fetch the metadata file folder  Directory to fetch metadata file
        file_name   metadata file name
        """
        wb_file = openpyxl.load_workbook(file_name, data_only=True)
        ws_metadata_lab = wb_file["METADATA_LAB"]
        heading = []
        for cell in ws_metadata_lab[1]:
            heading.append(cell.value)

    def read_json_file(self, j_file):
        """Read json file."""
        with open(j_file, "r") as fh:
            data = json.load(fh)
        return data

    def get_laboratory_data(self, lab_json, geo_loc_json, lab_name):
        """Fetch the laboratory location  and return a dictionary"""
        data = {}
        for lab in lab_json:
            if lab_name == lab["collecting_institution"]:
                for key, value in lab_name.items():
                    data[key] = value
                break
        for city in geo_loc_json:
            if city["geo_loc_city"] == data["geo_loc_city"]:
                data[city["geo_loc_city"]] = {}
                data[city["geo_loc_city"]]["geo_loc_latitude"] = city[
                    "geo_loc_latitude"
                ]
                data[city["geo_loc_city"]]["geo_loc_longitude"] = city[
                    "geo_loc_longitude"
                ]
                break
        return data

    def add_extra_data(self, metadata, lab_json_file, geo_loc_file):
        """Add the additional information that must be included in final metadata
        metadata Origin metadata
        extra_data  additional data to be included
        result_metadata    final metadata after adding the additional data
        """
        lab_data = {}
        extra_metadata = []
        extra_data = ""
        lab_json = self.read_json_file(lab_json_file)
        geo_loc_json = self.read_json_file(geo_loc_file)
        samples_json = self.read_json_file(self.sample_list_file)

        for row in metadata:
            for new_field, value in extra_data.items():
                row[new_field] = value
            if row["collecting_institution"] not in lab_data:
                # from collecting_institution find city, and geo location latitude and longitude
                l_data = self.get_laboratory_data(
                    lab_json, geo_loc_json, row["collecting_institution"]
                )
                row.update(l_data)
                lab_data[row["collecting_institution"]] = l_data
            else:
                row.update(lab_data[row["collecting_institution"]])

            try:
                s_data = samples_json[row["collecting_lab_sample_id"]]
                for key, values in s_data.items():
                    if key.endswith("_R1_fastq.gz"):
                        row["sequence_file_R1_fastq"] = key
                        row["r1_fastq_filepath"] = values["local_folder"]
                        row["fastq_md5"] = values["md5"]
                    elif key.endswith("_R2_fastq.gz"):
                        row["sequence_file_R2_fastq"] = key
                        row["r2_fastq_filepath"] = values["local_folder"]
                        # # WARNING:  no md5 value for R2 is deficned on schena
                        # row["fastq_md5"] = values["md5"]
                    elif key.endswith(".fasta"):
                        file_path = os.path.join(values["local_folder"], key)
                        row["consensus_sequence_filepath"] = file_path
                        # # WARNING:  no md5 value for fasta is deficned on schena
                        # row["fastq_md5"] = values["md5"]
            except KeyError:
                log.error(
                    "There is no files for sample %s", row["collecting_lab_sample_id"]
                )

            # update isolate qith the name of the sample
            row["isolate"] = row["collecting_lab_sample_id"]
            row["host_scientific_name"] = extra_data["host_scientific_name"][
                row["host_common_name"]
            ]
            row["sequencing_instrument_platform"] = "To change"
            extra_metadata.append(row)
        return extra_metadata

        # def compare_sample_in_metadata(self, completed_metadata):
        """Compare the samples defined in metadata file and the ones in the
        sample file
        """
        """
        not_found_samples = []
        if not os.path.exists(self.sample_list_file):
            return False
        # get the smaples defined in json
        with open(self.sample_list_file, "r") as fh:
            samples = fh.read().split("\n")
        for line_metadata in completed_metadata:
            if line_metadata["collecting_lab_sample_id"] not in samples:
                not_found_samples.append(line_metadata["collecting_lab_sample_id"])
        if len(not_found_samples) > 0:
            return not_found_samples
        return True
        """

    def request_information(external_url, request):
        """Get information from external database server using Rest API

        external_url
        request
        """
        pass

    def store_information(external_url, request, data):
        """Update information"""
        pass

        # def get_geo_location_data(self, state, country):
        """Get the geo_loc_latitude and geo_loc_longitude from state"""
        """
        geolocator = Nominatim(user_agent="geoapiRelecov")
        loc = geolocator.geocode(state + "," + country)
        return [str(loc.latitude), str(loc.longitude)]
        """

        # def update_heading_to_json(self, heading, meta_map_json):
        """Change the heading values from the metadata file for the ones defined
        in the json schema
        """
        """
        mapped_heading = []
        for cell in heading:
            if cell in meta_map_json:
                mapped_heading.append(meta_map_json[cell])
            else:
                mapped_heading.append(cell)
        return mapped_heading
        """

    def read_metadata_file(self):
        """Read the input metadata file, changing the metadata heading with
        their property name values defined in schema.
        Convert the date colunms value to the dd/mm/yyyy format.
        Return list of dict with data, and errors
        """
        wb_file = openpyxl.load_workbook(self.metadata_file, data_only=True)
        ws_metadata_lab = wb_file["METADATA_LAB"]
        # removing the None columns in excel heading row
        heading = [i.value.strip() for i in ws_metadata_lab[4] if i.value]
        # heading = self.update_heading_to_json(heading_without_none, meta_map_json)
        metadata_values = []
        errors = {}
        for row in islice(ws_metadata_lab.values, 4, ws_metadata_lab.max_row):
            sample_data_row = {}
            # Ignore the empty rows
            if row[2] is None:
                continue
            for idx in range(1, len(heading)):
                if "date" in heading[idx]:
                    try:
                        sample_data_row[self.label_prop_dict[heading[idx]]] = row[
                            idx
                        ].strftime("%Y/%m/%d")
                    except AttributeError:
                        if row[2] not in errors:
                            errors[row[2]] = {}
                        errors[row[2]][heading[idx]] = "Invalid date format"
                        log.error("Invalid date format in sample %s", row[2])
                        stderr.print(
                            "[red] Invalid date format in sample",
                            row[2] + " column " + heading[idx],
                        )
                else:
                    try:
                        sample_data_row[self.label_prop_dict[heading[idx]]] = (
                            row[idx] if row[idx] else ""
                        )
                    except KeyError as e:
                        print(e)
            metadata_values.append(sample_data_row)
        # import pdb; pdb.set_trace()
        return metadata_values, errors

    def write_json_fo_file(self, completed_metadata, file_name):
        """Write metadata to json file"""
        os.makedirs(self.output_folder, exist_ok=True)
        json_file = os.path.join(self.output_folder, file_name)
        with open(json_file, "w", encoding="utf-8") as fh:
            fh.write(
                json.dumps(
                    completed_metadata, indent=4, sort_keys=True, ensure_ascii=False
                )
            )
        return True

    def create_metadata_json(self):
        config_json = ConfigJson()
        """
        schema_location = config_json.get_topic_data(
            "json_schemas", "phage_plus_schema"
        )
        schema_location_file = os.path.join(
            os.path.dirname(os.path.realpath(__file__)), "schema", schema_location
        )
        schema_json = self.read_json_file(schema_location_file)
        phage_plus_schema = relecov_tools.schema_json.PhagePlusSchema(schema_json)
        properties_in_schema = phage_plus_schema.get_schema_properties()
        """
        geo_loc_json = config_json.get_configuration("geo_location_data")
        geo_loc_file = os.path.join(
            os.path.dirname(os.path.realpath(__file__)), "conf", geo_loc_json
        )
        lab_json = config_json.get_configuration("laboratory_data")
        lab_json_file = os.path.join(
            os.path.dirname(os.path.realpath(__file__)), "conf", lab_json
        )
        """
        metadata_mapping_json = config_json.get_configuration("mapping_metadata_json")
        meta_map_json_file = os.path.join(
            os.path.dirname(os.path.realpath(__file__)), "schema", metadata_mapping_json
        )
        meta_map_json = self.read_json_file(meta_map_json_file)
        """
        valid_metadata_rows, errors = self.read_metadata_file()

        completed_metadata = self.add_extra_data(
            valid_metadata_rows,
            lab_json_file,
            geo_loc_file,
        )
        comp_result = self.compare_sample_in_metadata(completed_metadata)
        if isinstance(comp_result, list):
            missing_samples = ",".join(comp_result)
            log.error("Missing samples %s", missing_samples)
        elif comp_result:
            log.info("Samples in metadata matches with the ones uploaded")
        else:
            log.error("There is missing samples in metadata and/or uploaded")
        self.write_json_fo_file(completed_metadata, "completed_metadata.json")
        return True
