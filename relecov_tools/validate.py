#!/usr/bin/env python
import rich.console
from jsonschema import Draft202012Validator, FormatChecker
import os
import re
import openpyxl
from datetime import datetime
from collections import defaultdict

import relecov_tools.utils
import relecov_tools.assets.schema_utils.jsonschema_draft
import relecov_tools.assets.schema_utils.custom_validators
import relecov_tools.sftp_client
from relecov_tools.config_json import ConfigJson
from relecov_tools.base_module import BaseModule
from relecov_tools.rest_api import RestApi


stderr = rich.console.Console(
    stderr=True,
    style="dim",
    highlight=False,
    force_terminal=relecov_tools.utils.rich_force_colors(),
)


class Validate(BaseModule):
    def __init__(
        self,
        json_file=None,
        json_schema_file=None,
        metadata=None,
        output_dir=None,
        excel_sheet=None,
        registry=None,
        upload_files=False,
        logsum_file=None,
        check_db=False,
    ):
        """Validate json file against the schema"""
        super().__init__(output_dir=output_dir, called_module=__name__)
        self.config = ConfigJson(extra_config=True)
        self.log.info("Initiating validation process")
        req_conf = (
            ["validate"]
            + ["download"] * bool(upload_files)
            + ["update_db"] * bool(check_db)
        )
        missing = [
            conf for conf in req_conf if self.config.get_configuration(conf) is None
        ]
        if missing:
            self.log.error(
                "Extra config file () is missing required sections: %s"
                % ", ".join(missing)
            )
            self.log.error(
                "Please use add-extra-config to add them to the config file."
            )
            stderr.print(
                f"[red]Config file is missing required sections: {', '.join(missing)}"
            )
            stderr.print(
                "[red]Please use add-extra-config to add them to the config file."
            )
            raise ValueError(
                f"Config file is missing required sections: {', '.join(missing)}"
            )
        if json_schema_file is None:
            schema_name = self.config.get_topic_data("generic", "relecov_schema")
            json_schema_file = os.path.join(
                os.path.dirname(os.path.realpath(__file__)), "schema", schema_name
            )

        self.json_schema = relecov_tools.utils.read_json_file(json_schema_file)

        if json_file is None:
            json_file = relecov_tools.utils.prompt_path(
                msg="Select the json file to be validated"
            )

        if output_dir is None:
            self.out_folder = relecov_tools.utils.prompt_path(
                msg="Select the folder where excel file with invalid data will be saved"
            )
        else:
            self.out_folder = output_dir

        # Read and check json to validate file
        if not os.path.isfile(json_file):
            stderr.print("[red] Json file does not exist")
            self.log.error("Json file does not exist")
            raise ValueError(f"Json file '{json_file}' does not exist")
        self.json_data_file = json_file
        out_path = os.path.dirname(os.path.realpath(self.json_data_file))

        stderr.print("[blue] Reading the json file")
        self.log.info("Reading the json file")
        self.json_data = relecov_tools.utils.read_json_file(json_file)
        if not isinstance(self.json_data, list):
            stderr.print(f"[red]Invalid json file content in {json_file}.")
            stderr.print("Should be a list of dicts. Create it with read-lab-metadata")
            self.log.error(f"[red]Invalid json file content in {json_file}.")
            self.log.error(
                "Should be a list of dicts. Create it with read-lab-metadata"
            )
            raise TypeError(f"Invalid json file content in {json_file}")
        try:
            batch_id = self.get_batch_id_from_data(self.json_data)
        except ValueError:
            raise ValueError(f"Provided json file {json_file} is empty")
        except AttributeError as e:
            raise ValueError(f"Invalid json file content in {json_file}: {e}")
        try:
            unique_institutions = set(
                [x.get("submitting_institution_id") for x in self.json_data]
            )
            if len(unique_institutions) > 1:
                self.log.warning(
                    f"All samples in {json_file} should be from the same submitting institution. Found {unique_institutions}"
                )
            self.lab_code = self.json_data[0]["submitting_institution_id"]
            self.log.info(f"Laboratory code set to {self.lab_code}")
        except Exception as e:
            self.log.warning(f"Could not extract lab_code from json_data: {e}")
            self.lab_code = out_path.split("/")[-2]
        self.logsum = self.parent_log_summary(
            output_dir=self.out_folder, lab_code=self.lab_code, path=out_path
        )
        self.set_batch_id(batch_id)
        self.metadata = metadata

        # TODO: Include this field in configuration.json
        sample_id_ontology = self.config.get_topic_data("generic", "sample_id_ontology")
        try:
            self.sample_id_field = Validate.get_field_from_schema(
                sample_id_ontology, self.json_schema
            )
        except ValueError as e:
            self.sample_id_field = None
            self.log.error(f"Could not extract sample_id_field: {e}. Set to None")
        self.excel_sheet = excel_sheet
        self.registry_path = registry
        self.logsum_file = logsum_file
        self.upload_files = upload_files
        self.check_db = check_db
        if upload_files:
            upload_config = self.config.get_configuration("download")
            if upload_config:
                self.user = upload_config.get("user")
                self.password = upload_config.get("password")
                self.subfolder = upload_config.get("subfolder")
                self.sftp_port = self.config.get_topic_data("sftp_handle", "sftp_port")
            else:
                raise ValueError("Could not find")
        if check_db:
            platform_config = self.config.get_configuration("update_db")
            if platform_config:
                self.db_user = platform_config.get("user")
                self.db_pass = platform_config.get("password")
                self.db_platform = platform_config.get("platform")

    def validate_schema(self):
        """Validate json schema against draft and check if all properties have label"""
        relecov_tools.assets.schema_utils.jsonschema_draft.check_schema_draft(
            self.json_schema, "2020-12"
        )
        for prop_name, prop_def in self.json_schema["properties"].items():
            if "label" not in prop_def:
                self.log.debug(f"Property {prop_name} is missing 'label'")
        return

    @staticmethod
    def get_field_from_schema(ontology, schema_json):
        """Find the name of the field used to track the samples in the given schema
        using its ontology value

        Args:
            field (str): Name of the field
            schema_json (dict): Loaded json schema as a dictionary

        Returns:
            sample_id_field (str): Name of the FIRST field that matches the given ontology
        """
        ontology_match = [
            x
            for x, y in schema_json["properties"].items()
            if y.get("ontology") == ontology
        ]
        if ontology_match:
            sample_id_field = ontology_match[0]
        else:
            error_text = f"No valid sample ID field ({ontology}) in schema"
            raise ValueError(error_text)
        return sample_id_field

    @staticmethod
    def validate_instances(
        json_data, json_schema, sample_id_field=None, validator=None
    ):
        """Validate data instances against a validated JSON schema

        Args:
            json_data (list(dict)): List of samples with processed metadata
            json_schema (dict): Loaded JSON schema as a dictionary
            sample_id_field (str, optional): Metadata field used as ID to
            identify the samples associated with each error.
            validator (jsonschema.Validator(), optional): Validator with any custom
            characteristics included. Default is Draft202012Validator()

        Returns:
            validated_json_data (list(dict)): List of successfully validated samples
            errors (dict): Custom dict used to summarize validation errors.
            '''
                errors = {
                    "fields": {
                        'host_age is a required property': 'host_age'
                        'Collection_date is not a valid date': 'Collection_date'
                    },
                    "samples": {
                        'host_age is a required property': ["sample_1", "sample_2"]
                        'Collection_date is not a valid date': ["sample_1"]
                    }
                }
            '''
        """

        def get_property_label(schema_props, error_field):
            """Extract the label for the given property given a list of schema properties"""
            try:
                err_field_label = schema_props[error_field]["label"]
            except KeyError:
                return error_field
            return err_field_label

        # Create default validator if not given.
        if not validator:
            validator = Draft202012Validator(
                json_schema, format_checker=FormatChecker()
            )
        schema_props = json_schema["properties"]

        validated_json_data = []
        errors = defaultdict(dict)

        stderr.print("[blue] Start processing the JSON file")

        # Start validation
        for idx, item_row in enumerate(json_data):
            sample_id_value = item_row.get(sample_id_field, f"UnknownSample#{idx}")

            # Collect all errors (don't raise immediately)
            validation_errors = list(validator.iter_errors(item_row))

            # Run the custom validator to check if errors should be ignored
            validation_errors = relecov_tools.assets.schema_utils.custom_validators.validate_with_exceptions(
                json_schema, item_row, validation_errors
            )

            if not validation_errors:
                validated_json_data.append(item_row)

            else:
                # Process remaining errors
                for error in validation_errors:
                    if error.cause:
                        # Probably generated by a custom validator or format_checker
                        error.message = str(error.cause)
                    try:
                        if error.validator == "required":
                            error_field = list(error.message.split("'"))[1]
                        elif error.validator == "anyOf":
                            # AnyOf errors include multiple clauses so they need some processing
                            multi_errdict = {}
                            for suberror in error.context:
                                error_type = suberror.validator
                                suberr_label = get_property_label(
                                    schema_props, suberror.validator_value[0]
                                )
                                label_message = suberror.message.replace(
                                    suberror.validator_value[0], suberr_label
                                )
                                if suberror.validator in multi_errdict:
                                    multi_errdict[error_type].append(
                                        (suberr_label, label_message)
                                    )
                                else:
                                    multi_errdict[error_type] = [
                                        (suberr_label, label_message)
                                    ]
                            error_field = ""
                            multi_message = {}
                            # Combine the different error messages from this AnyOf in a single message
                            for errtype, fieldtups in multi_errdict.items():
                                failed_fields = " or ".join([t[0] for t in fieldtups])
                                clean_message = (
                                    fieldtups[0][1]
                                    .replace(fieldtups[0][0], "")
                                    .strip("'")
                                )
                                if error_field:
                                    error_field = error_field + " and"
                                error_field = error_field + failed_fields
                                multi_message[errtype] = (
                                    f"{failed_fields}: {clean_message}"
                                )
                            # Override error.message with the combination of the sub-messages
                            error.message = "Any of the following: " + " --- ".join(
                                multi_message.values()
                            )
                        elif error.absolute_path:
                            error_field = str(error.absolute_path[0])
                        else:
                            error_field = error.validator + " error: " + error.message
                    except Exception as ex:
                        errtxt = f"Error extracting error_field from: {error.validator_value}, {ex}"
                        error_field = str(error)
                        errors["fields"][errtxt] = error.validator_value
                        errors["samples"].setdefault(errtxt, []).append(sample_id_value)
                        continue

                    # Try to get the human-readable label from the schema
                    err_field_label = get_property_label(schema_props, error_field)
                    # Format the error message
                    error.message = error.message.replace(error_field, err_field_label)
                    error_text = f"Error in column {err_field_label}: {error.message}"

                    # Log errors for summary
                    errors["fields"][error_text] = error_field
                    errors["samples"].setdefault(error_text, []).append(sample_id_value)

                # Add the invalid row to the list
        return validated_json_data, errors

    def summarize_errors(self, errors):
        """Summarize errors from validation process and add them to log_summary

        Args:
            errors (dict): Error dict from validate_instances()
        """

        def truncate_error_message(error_text, max_length):
            truncated_msg = (
                error_text[:max_length] + "..."
                if len(error_text) > max_length
                else error_text
            )
            return truncated_msg

        stderr.print("[blue] --------------------")
        stderr.print("[blue] VALIDATION SUMMARY")
        stderr.print("[blue] --------------------")
        self.log.info("Validation summary:")
        max_length = 250
        for error_type, failed_samples in errors["samples"].items():
            count = len(failed_samples)
            field_with_error = errors["fields"][error_type]
            error_text = f"{count} samples failed validation for {field_with_error}: {error_type}"
            truncated_msg = truncate_error_message(error_text, max_length)
            self.logsum.add_warning(entry=truncated_msg)
            for failsamp in failed_samples:
                err_msg = truncate_error_message(error_type, max_length)
                self.logsum.add_error(sample=failsamp, entry=err_msg)
            stderr.print(f"[red]{truncated_msg}")
            stderr.print("[red] --------------------")
        return

    def create_invalid_metadata(self, invalid_json, metadata, out_folder):
        """Create a new sub excel file having only the samples that were invalid.
        Samples name are checking the Sequencing sample id which are in
        column B (index 1).
        The rows that match the value collected from json file on tag
        collecting_lab_sample_id are removed from excel
        """
        if self.sample_id_field is None:
            log_text = f"Invalid excel file won't be created: {self.SAMPLE_FIELD_ERROR}"
            self.logsum.add_error(entry=log_text)
            return
        self.log.info("Trying to create invalid metadata excel file...")
        if metadata is None:
            metadata = relecov_tools.utils.prompt_path(
                msg="Select the metadata file to select those not-validated samples."
            )
        if not os.path.isfile(metadata):
            self.log.error("Metadata file %s does not exist", metadata)
            stderr.print(
                "[red] Unable to create excel file for invalid samples. Metadata file ",
                metadata,
                " does not exist",
            )
            raise FileNotFoundError(
                f"Unable to create excel file for invalid samples. Metadata file {metadata} does not exist"
            )
        sample_list = []
        stderr.print("Start preparation of invalid samples...")
        self.log.info("Start preparation of invalid samples")
        for row in invalid_json:
            sample_list.append(str(row[self.sample_id_field]))
        wb = openpyxl.load_workbook(metadata)
        try:
            ws_sheet = wb[self.excel_sheet]
        except KeyError:
            logtxt = f"No sheet named {self.excel_sheet} could be found in {metadata}"
            self.log.error(logtxt)
            raise
        tag = "Sample ID given for sequencing"
        # Check if mandatory colum ($tag) is defined in metadata.
        try:
            header_row = [idx + 1 for idx, x in enumerate(ws_sheet.values) if tag in x][
                0
            ]
        except IndexError:
            self.log.error(
                f"Column with tag '{tag}' not found in any row of the Excel sheet."
            )
            stderr.print(f"[red]Column with tag '{tag}' not found. Cannot continue.")
            raise
        row_to_del = []
        row_iterator = ws_sheet.iter_rows(
            min_row=header_row + 1, max_row=ws_sheet.max_row
        )
        consec_empty_rows = 0
        id_col = [
            idx for idx, val in enumerate(ws_sheet[header_row]) if val.value == tag
        ][0]
        for row in row_iterator:
            # if no data in 10 consecutive rows, break loop
            if not any(row[x].value for x in range(10)):
                row_to_del.append(row[0].row)
                consec_empty_rows += 1
            if consec_empty_rows > 10:
                break
            consec_empty_rows = 0
            if str(row[id_col].value) not in sample_list:
                row_to_del.append(row[0].row)
        stderr.print("Collected rows to create the excel file")
        if len(row_to_del) > 0:
            row_to_del.sort(reverse=True)
            for idx in row_to_del:
                try:
                    ws_sheet.delete_rows(idx)
                except TypeError as e:
                    self.log.error(
                        "Unable to delete row %s from metadata file because of",
                        idx,
                        e,
                    )
                    stderr.print(f"[red] Unable to delete row {idx} becuase of {e}")
                    raise
        os.makedirs(out_folder, exist_ok=True)
        new_name = "invalid_" + os.path.basename(metadata)
        m_file = os.path.join(out_folder, new_name)
        self.log.info(f"Saving excel file with the invalid samples: {m_file}")
        wb.save(m_file)
        stderr.print(f"Saved excel file with the invalid samples: {m_file}")
        return m_file

    def create_validated_json(self, valid_json_data, out_folder):
        """Create a copy of the input json file, keeping only the validated samples

        Args:
            valid_json_data (list(dict)): List of valid samples metadata as dictionaries
            out_folder (str): path to folder where file will be created
        """
        file_name = "_".join(["validated", os.path.basename(self.json_data_file)])
        file_path = os.path.join(out_folder, file_name)
        self.log.info("Saving Json file with the validated samples in %s", file_path)
        stderr.print(f"Saving Json file with the validated samples in {file_path}")
        relecov_tools.utils.write_json_to_file(valid_json_data, file_path)
        return file_path

    def validate_registry_file(self):
        """Validate specified registry file path. Try to get it from config if invalid."""
        # Parse file containing sample IDs registry
        if self.registry_path is not None and relecov_tools.utils.file_exists(
            os.path.abspath(os.path.expanduser(self.registry_path))
        ):
            self.registry_path = os.path.abspath(os.path.expanduser(self.registry_path))
        else:
            config_json = ConfigJson(extra_config=True)
            try:
                default_path = config_json.get_topic_data(
                    "generic", "default_sample_id_registry"
                )
            except KeyError:
                default_path = None

            if default_path and relecov_tools.utils.file_exists(default_path):
                self.registry_path = default_path
            else:
                stderr.print(
                    "[yellow]No valid ID registry found. Please select the file manually."
                )
                prompted_path = relecov_tools.utils.prompt_path(
                    "Select the JSON file with registered unique sample IDs"
                )
                if relecov_tools.utils.file_exists(prompted_path):
                    self.registry_path = prompted_path
                else:
                    stderr.print(
                        "[red]No valid ID registry file could be found or selected."
                    )
                    raise FileNotFoundError("No valid ID registry file could be found")

        # Read id registry file
        try:
            self.id_registry = relecov_tools.utils.read_json_file(self.registry_path)
        except Exception as e:
            stderr.print(f"[red]Failed to read ID registry JSON: {e}")
            self.log.error(f"Failed to read ID registry JSON: {e}")
            raise
        return

    def validate_invexcel_args(self):
        """Validate arguments needed to create invalid_samples.xlsx file"""
        if not self.excel_sheet:
            conf_subdata = self.config.get_topic_data(
                "sftp_handle", "metadata_processing"
            )
            try:
                self.excel_sheet = conf_subdata["excel_sheet"]
            except KeyError:
                self.log.error("Default metadata sheet name should be in config file")
                raise
        return

    def update_unique_id_registry(self, valid_json_data):
        """Prepare ID registry. Validated samples will be asigned with an unique ID"""
        new_ids_to_save = {}
        for sample in valid_json_data:
            # If sample has validated, then assign an unique id
            if "unique_sample_id" not in sample:
                new_id = self.generate_incremental_unique_id(
                    {**self.id_registry, **new_ids_to_save}
                )
                sample["unique_sample_id"] = new_id
            # If sample has validated, then assign an unique id
            if "unique_sample_id" not in sample:
                new_id = self.generate_incremental_unique_id(
                    {**self.id_registry, **new_ids_to_save}
                )
                sample["unique_sample_id"] = new_id
                new_ids_to_save[new_id] = {
                    "sequencing_sample_id": sample["sequencing_sample_id"],
                    "lab_code": self.lab_code,
                    "generated_at": datetime.now().isoformat(timespec="seconds"),
                }
        self.save_new_ids(new_ids_to_save)
        return valid_json_data

    def generate_incremental_unique_id(self, current_registry, prefix="RLCV"):
        """Generates an incremental unique ID using as baseline the latest record
        in a registry.

        Args:
            current_registry (dict): dict containing sample's unique IDs
            prefix (str, optional): String that preceeds the unique ID. Defaults to "RLCV".

        Returns:
            string: An unique ID
        """
        existing_numbers = []
        for uid in current_registry:
            if uid.startswith(prefix):
                try:
                    number = int(uid.replace(f"{prefix}-", ""))
                    existing_numbers.append(number)
                except ValueError:
                    continue
        next_number = max(existing_numbers, default=0) + 1
        return f"{prefix}-{next_number:09d}"

    def save_new_ids(self, new_ids_dict):
        """Updates sample id registry by adding sample unique ids of new validated samples.

        Args:
            new_ids_dict (dict): Dict of new unique sample IDs.
        """
        updated_registry = {**self.id_registry, **new_ids_dict}
        relecov_tools.utils.write_json_to_file(updated_registry, self.registry_path)
        self.log.info(f"Added {len(new_ids_dict)} new sample IDs to registry")

    def create_validation_files(self, valid_json_data, invalid_json):
        """Creates validated and invalid metadata files based on the provided JSON data and
        adds valid samples to the unique ID registry.

        Args:
            valid_json_data (list[dict]): List of valid sample metadata dictionaries.
            invalid_json (list[dict]): List of invalid sample metadata dictionaries.

        Returns:
            valid_file (str): Path to the created file with valid samples, or None if not created
            invalid_file (str): Path to the created file with invalid samples, or None if not created
        """
        # Add all valid samples to the unique_id registry file
        self.validate_registry_file()
        valid_json_data = self.update_unique_id_registry(valid_json_data)
        valid_file = invalid_file = None
        if valid_json_data:
            self.log.info("Creating json_file with validated samples...")
            valid_file = self.create_validated_json(valid_json_data, self.out_folder)
        else:
            log_text = "All the samples were invalid. No valid file created"
            self.logsum.add_error(entry=log_text)
            stderr.print(f"[red]{log_text}")

        if invalid_json:
            log_text = "Summary: %s valid and %s invalid samples"
            self.logsum.add_warning(
                entry=log_text % (len(valid_json_data), len(invalid_json))
            )
            self.validate_invexcel_args()
            invalid_file = self.create_invalid_metadata(
                invalid_json, self.metadata, self.out_folder
            )
        else:
            stderr.print("[green]Sucessful validation, no invalid file created!!")
            self.log.info("Sucessful validation, no invalid file created.")
        return valid_file, invalid_file

    def validate_upload_args(self):
        """Validates that all required arguments for the upload process are present
        and that the log summary file exists.

        Raises:
            ValueError: If any required argument is missing or the log summary file does not exist.
        """
        required_args = ["user", "password"]
        self.log.debug(f"Trying to validate args for upload: {required_args}")
        missing_args = []
        for arg in required_args:
            if not self.__dict__.get(arg):
                missing_args.append(arg)
        if missing_args:
            raise ValueError(
                f"Missing mandatory args to upload validated files: {missing_args}"
            )
        if not self.subfolder:
            self.log.warning(
                "No subfolder provided. Uploading files to main lab folder"
            )
        if not self.logsum_file or not os.path.isfile(self.logsum_file):
            raise ValueError(
                f"Provided log_summary.json from previous processes does not exist: {self.logsum_file}"
            )
        return

    def update_invalid_with_logsum(self, invalid_json, previous_logsum):
        """Adds previously invalid samples from a prior log summary to the current invalid list.

        Args:
            invalid_json (list[dict]): Current list of invalid sample metadata.
            previous_logsum (dict): Log summary containing past validation results.

        Returns:
            list[dict]: Updated list of invalid samples, including any re-detected ones from the log.
        """
        log_invalid_samples = []
        self.log.info("Updating invalid samples with previous log_summary")
        if self.lab_code in previous_logsum.keys():
            for samp, logs in previous_logsum[self.lab_code].get("samples", {}).items():
                if not logs["valid"]:
                    self.log.debug(f"Found sample {samp} invalid in log_summary")
                    log_invalid_samples.append(samp)
        else:
            errtxt = f"Lab code {self.lab_code} not found in {self.logsum_file} to update invalid samples"
            self.log.warning(errtxt)
            stderr.print(f"[orange]{errtxt}")
            return invalid_json

        samp_id = "sequencing_sample_id"
        updated_invalid = [
            x
            for x in self.json_data
            if x.get(samp_id) in log_invalid_samples and x not in invalid_json
        ]
        invalid_json.extend([x for x in updated_invalid if x not in invalid_json])
        return invalid_json

    def upload_validation_results(self, invalid_json, invalid_excel):
        """Uploads invalid sample files and related reports to a remote SFTP server.
        Checks that required remote directories exist or creates them.

        Args:
            invalid_json (list[dict]): List of invalid samples containing file paths.
            invalid_excel (str): Path to the Excel metadata report file.

        Raises:
            FileNotFoundError: If expected files or folders are not found locally or remotely.
        """

        def upload_and_clean(local_file, remote_dest, clean=True):
            """Upload file to remote sftp and log the process."""
            self.log.debug(f"Uploading {local_file} → {remote_dest}")
            if not sftp_client.upload_file(local_file, remote_dest):
                self.log.error(f"Could not upload {local_file} to sftp.")
                return False
            self.log.debug(f"{remote_dest} uploaded successfully")
            if clean:
                try:
                    os.remove(local_file)
                except OSError as e:
                    self.log.error(f"Could not remove {local_file}: {e}")
                    return False
            return True

        # ── 1· Connection and remote folders ──────────────────────────
        self.log.info("Initiating sftp client to upload invalid files")
        sftp_client = relecov_tools.sftp_client.SftpClient(
            username=self.user, password=self.password
        )
        sftp_client.sftp_port = self.sftp_port

        remote_labfold = (
            os.path.join(self.lab_code, self.subfolder)
            if self.subfolder
            else self.lab_code
        )

        self.log.info(f"Output folder set to {remote_labfold}")
        if f"./{remote_labfold}" not in sftp_client.list_remote_folders(
            ".", recursive=bool(self.subfolder)
        ):
            raise FileNotFoundError(f"Couldn't find remote lab folder {remote_labfold}")

        invalid_remote_folder = f"{self.batch_id}_invalid_samples"
        self.remote_outfold = os.path.join(f"./{remote_labfold}", invalid_remote_folder)
        if invalid_remote_folder not in sftp_client.list_remote_folders(
            f"./{remote_labfold}"
        ):
            self.log.info(f"{invalid_remote_folder} not found, creating it…")
            sftp_client.make_dir(self.remote_outfold)

        stderr.print(f"[blue]Uploading invalid files to remote {remote_labfold}...")

        # ── 2· Build map path → sample_id  ─────────────────────
        path_fields = [
            ("sequence_file_path_R1", "sequence_file_R1"),
            ("sequence_file_path_R2", "sequence_file_R2"),
        ]
        path_to_sample = {}
        invalid_files = []
        for row in invalid_json:
            samp_id = row.get(self.sample_id_field, "UnknownSample")
            for p, f in path_fields:
                if p in row and f in row:
                    file_path = os.path.join(row[p], row[f])
                    invalid_files.append(file_path)
                    path_to_sample[file_path] = samp_id

        # ── 3· Try to upload each FASTQ ────────────────────────────
        failed_uploads = []
        for file in invalid_files:
            if not os.path.isfile(file):
                failed_uploads.append(file)
                self.log.warning(f"Local file not found, skipping upload: {file}")
                self.logsum.add_error(
                    sample=path_to_sample[file],
                    entry="File missing when uploading invalid FastQ",
                )
                continue
            remote_dest = os.path.join(self.remote_outfold, os.path.basename(file))
            if not upload_and_clean(file, remote_dest):
                failed_uploads.append(file)
                self.logsum.add_error(
                    sample=path_to_sample[file], entry="Failed to upload invalid FastQ"
                )

        # ── 4· Upload invalid.xlsx ────────────────────────
        remote_dest = os.path.join(self.remote_outfold, os.path.basename(invalid_excel))
        if not upload_and_clean(invalid_excel, remote_dest, clean=False):
            failed_uploads.append(invalid_excel)

        # ── 5· Locate & upload the *_metadata_report.xlsx ───────────
        pattern = re.compile(
            rf"^{re.escape(self.lab_code)}_{re.escape(self.batch_id)}.*_metadata_report\.xlsx$"
        )
        for rpt in (f for f in os.listdir(self.out_folder) if pattern.match(f)):
            local_path = os.path.join(self.out_folder, rpt)
            remote_dest = os.path.join(self.remote_outfold, rpt)
            if not upload_and_clean(local_path, remote_dest, clean=False):
                failed_uploads.append(local_path)

        # ── 6· Screen summary & log_summary ───────────────────
        if failed_uploads:
            preview = ", ".join(os.path.basename(x) for x in failed_uploads[:3])
            stderr.print(
                f"[yellow]{len(failed_uploads)} files could not be uploaded "
                f"(first ones: {preview})"
            )
            self.log.warning(f"Files failed to upload: {failed_uploads}")
        else:
            stderr.print("[green]Finished uploading files to remote sftp successfully")

    def process_validation_upload(self, valid_json_data, invalid_json, previous_logsum):
        """
        Coordinates the process of validating and uploading metadata files.

        - Merges current and previous log summaries.
        - Generates an updated Excel metadata report.
        - Uploads invalid samples and reports to SFTP if any are found.

        Args:
            valid_json_data (list[dict]): List of valid sample metadata dictionaries.
            invalid_json (list[dict]): Current list of invalid sample metadata.
            previous_logsum (dict): Dictionary containing results from a prior validation run.
        """
        merged_logsum = self.logsum.merge_logs(
            [previous_logsum, self.logsum.logs], key_name=self.lab_code
        )
        logsum_basename = self.tag_filename(self.lab_code) + "_metadata_report.json"
        sumfile = os.path.join(self.out_folder, logsum_basename)
        self.parent_create_error_summary(
            called_module="metadata",
            to_excel=True,
            logs=merged_logsum,
            filepath=sumfile,
        )
        _, invalid_file = self.create_validation_files(valid_json_data, invalid_json)
        if invalid_file:
            self.log.info("Starting upload process of invalid_samples...")
            self.upload_validation_results(invalid_json, invalid_file)
        else:
            stderr.print("[green]No invalid samples were found, no upload needed")
            self.log.info("No invalid samples were found, no upload needed")
        return

    def validate_db_args(self):
        """Validates that all required arguments to check if samples are
        already in platform database are present.

        Raises:
            ValueError: If any required argument or configuration is missing.
        """
        required_args = ["db_user", "db_pass", "db_platform"]
        self.log.debug(f"Trying to validate args to check database: {required_args}")
        missing_args = []
        for arg in required_args:
            if not self.__dict__.get(arg):
                missing_args.append(arg)
        if missing_args:
            raise ValueError(
                f"Missing mandatory args to check samples in db: {missing_args}"
            )
        p_settings = self.config.get_topic_data("upload_database", "platform")
        if self.db_platform not in p_settings:
            raise ValueError(f"No configuration found for platform {self.db_platform}")
        if "server_url" not in p_settings[self.db_platform]:
            raise ValueError(f"Missing 'server_url' in config for {self.db_platform}")
        if "api_url" not in p_settings[self.db_platform]:
            raise ValueError(f"Missing 'api_url' in config for {self.db_platform}")
        return

    def search_sample_dups_in_db(self, valid_json_data, invalid_json):
        """Connect to configured platform and turn invalid those samples that are already
        uploaded to the database from the workflow. Update jsons based on this clause"""
        p_settings = self.config.get_topic_data("upload_database", "platform")
        server_url = p_settings[self.db_platform]["server_url"]
        api_url = p_settings[self.db_platform]["api_url"]
        credentials = {
            "user": self.db_user,
            "pass": self.db_pass,
        }
        api_rest = RestApi(server_url, api_url)
        apifunc = p_settings[self.db_platform]["check_sample"]
        for sample in self.json_data:
            sample_seqid = sample.get("sequencing_sample_id")
            self.log.debug(f"Checking sample {sample_seqid} in {self.db_platform} db")
            try:
                samp_in_db = api_rest.sample_already_in_db(apifunc, credentials, sample)
            except ValueError as e:
                errtxt = f"Could not check for sample {sample_seqid} in db: {e}"
                stderr.print(f"[red]{errtxt}")
                self.log.error(errtxt)
                self.logsum.add_error(errtxt, sample=sample_seqid)
                continue
            if samp_in_db:
                errtxt = f"Sample {sample_seqid} already defined in db. Skipped"
                stderr.print(f"[yellow]{errtxt}")
                self.log.error(errtxt)
                self.logsum.add_error(errtxt, sample=sample_seqid)
                if sample in valid_json_data:
                    valid_json_data.remove(sample)
                    invalid_json.append(sample)
        return valid_json_data, invalid_json

    def validate(self):
        """Validate samples from metadata, create an excel with invalid samples,
        and a json file with the validated ones.
        """
        self.log.info("Validate the given schema")
        self.validate_schema()
        self.log.info("Preparing validator based on config")
        starting_date = self.config.get_topic_data("generic", "starting_date")
        date_checker = (
            relecov_tools.assets.schema_utils.custom_validators.make_date_checker(
                datetime.strptime(starting_date, "%Y-%m-%d").date(),
                datetime.now().date(),
            )
        )
        validator = Draft202012Validator(self.json_schema, format_checker=date_checker)
        self.log.info("Starting validation process of JSON file against schema")
        valid_json_data, errors = Validate.validate_instances(
            self.json_data,
            self.json_schema,
            sample_id_field=self.sample_id_field,
            validator=validator,
        )
        for sample in valid_json_data:
            sample_id_value = sample.get(self.sample_id_field)
            self.logsum.feed_key(sample=sample_id_value)
        if errors:
            self.summarize_errors(errors)
            stderr.print(
                f"[red]{len(valid_json_data)}/{len(self.json_data)} samples were valid"
            )
        else:
            stderr.print("[green]No errors found during metadata validation!")
        invalid_json = [x for x in self.json_data if x not in valid_json_data]
        return valid_json_data, invalid_json

    def execute_validation_process(self):
        """Execute all the validation process start to end"""
        valid_json_data, invalid_json = self.validate()
        if self.check_db:
            stderr.print(
                "[blue]Checking if samples are already uploaded to platform..."
            )
            try:
                self.validate_db_args()
                valid_json_data, invalid_json = self.search_sample_dups_in_db(
                    valid_json_data, invalid_json
                )
            except ValueError as e:
                stderr.print(f"[red]Could not check if samples are already in db: {e}")
                self.log.error(f"Could not check if samples are already in db: {e}")
                self.create_validation_files(valid_json_data, invalid_json)
                raise
        if self.logsum_file:
            previous_logsum = relecov_tools.utils.read_json_file(self.logsum_file)
            merged_logs = self.logsum.merge_logs([previous_logsum, self.logsum.logs])
            invalid_json = self.update_invalid_with_logsum(invalid_json, merged_logs)
        else:
            invalid_json = self.update_invalid_with_logsum(
                invalid_json, self.logsum.logs
            )
        valid_json_data = [x for x in valid_json_data if x not in invalid_json]
        if self.upload_files:
            stderr.print(f"Starting uploading process for {self.lab_code}...")
            self.log.info(f"Starting uploading process for {self.lab_code}...")
            try:
                self.validate_upload_args()
            except ValueError as e:
                stderr.print(f"[red]Could not upload validation files: {e}")
                self.log.error(f"Could not upload validation files: {e}")
                self.create_validation_files(valid_json_data, invalid_json)
                raise
            self.process_validation_upload(
                valid_json_data, invalid_json, previous_logsum
            )
        else:
            self.create_validation_files(valid_json_data, invalid_json)
            self.parent_create_error_summary(called_module="validate")
        return
