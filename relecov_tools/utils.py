#!/usr/bin/env python
"""
Common utility function used for relecov_tools package.
"""
import os
import sys
import glob
import hashlib
import logging
import questionary
import json
import openpyxl
import yaml
import gzip
import re
import shutil
from itertools import islice, product
from Bio import SeqIO
from rich.console import Console
from rich.table import Table
from datetime import datetime
from tabulate import tabulate
from secrets import token_hex
import openpyxl.utils
import openpyxl.styles
import pandas as pd
import semantic_version
import subprocess
import importlib.metadata

import relecov_tools.config_json

log = logging.getLogger(__name__)


def file_exists(file_to_check):
    """
    Input:
        file_to_check   # file name to check if exists
    Return:
        True if exists
    """
    if os.path.isfile(file_to_check):
        return True
    return False


def safe_remove(file_path):
    try:
        os.remove(file_path)
    except OSError:
        return False
    return True


def get_files_match_condition(condition):
    """find all path names that matches with the condition"""
    return glob.glob(condition)


def read_json_file(j_file):
    """Read json file."""
    with open(j_file, "r") as fh:
        try:
            data = json.load(fh)
        except (UnicodeDecodeError, ValueError):
            raise

    return data


def read_excel_file(f_name, sheet_name, header_flag, leave_empty=True):
    """Read the input excel file and return the data as a list of dictionaries.
    If openpyxl fails, fall back to pandas but return in the same format.
    """
    try:
        wb_file = openpyxl.load_workbook(f_name, data_only=True)
        ws_metadata_lab = wb_file[sheet_name]
        try:
            heading_row = [
                idx + 1
                for idx, x in enumerate(ws_metadata_lab.values)
                if header_flag in x
            ][0]
        except IndexError:
            raise KeyError(
                f"Header flag '{header_flag}' could not be found in {f_name}"
            )
        heading = [
            str(i.value).strip() for i in ws_metadata_lab[heading_row] if i.value
        ]
        ws_data = []
        for row in islice(ws_metadata_lab.values, heading_row, ws_metadata_lab.max_row):
            l_row = list(row)
            if all(cell is None for cell in l_row):
                continue
            data_row = {}
            for idx in range(0, len(heading)):
                if l_row[idx] is None:
                    data_row[heading[idx]] = (
                        None
                        if leave_empty
                        else relecov_tools.config_json.ConfigJson().get_topic_data(
                            "generic", "not_provided_field"
                        )
                    )
                else:
                    data_row[heading[idx]] = l_row[idx]
            ws_data.append(data_row)
        return ws_data, heading_row

    except Exception as e:
        try:
            df = pd.read_excel(f_name, sheet_name=sheet_name, header=None)

            heading_row_idx = df.apply(lambda row: header_flag in row.values, axis=1)
            if not heading_row_idx.any():
                raise KeyError(
                    f"Header flag '{header_flag}' could not be found in {f_name}"
                )
            heading_row = heading_row_idx.idxmax()
            heading = [str(h).strip() for h in df.iloc[heading_row] if pd.notna(h)]

            ws_data = []
            for _, row in df.iloc[heading_row + 1 :].iterrows():
                if row.isna().all():
                    continue
                data_row = {}
                for idx in range(len(heading)):
                    val = row.iloc[idx] if idx < len(row) else None
                    if pd.isna(val):
                        data_row[heading[idx]] = (
                            None
                            if leave_empty
                            else relecov_tools.config_json.ConfigJson().get_topic_data(
                                "generic", "not_provided_field"
                            )
                        )
                    else:
                        data_row[heading[idx]] = val
                ws_data.append(data_row)
            return (
                ws_data,
                heading_row + 1,
            )  # +1 to maintain consistency with openpyxl (1-based)

        except Exception as fallback_e:
            raise RuntimeError(
                f"Failed to read file with both openpyxl and pandas:\n- openpyxl error: {e}\n- pandas error: {fallback_e}"
            )


def string_to_date(string):
    """Convert date (Y-M-D...) from string to date. Tries iteratively with variable
    number of digits and multiple separators, starting from seconds up to year.
    args:
        string (str): Date in string format to be parsed: e.g. 2020-08-07-12-00-00
    returns:
        res_date (datetime.datetime): String converted to date format
    """

    def rec_date_extraction(string, digits, sep):
        regex = r"^\d{4}"  # The string should start from year
        for _ in range(0, digits - 4, 2):
            new_reg = sep + "\d{2}"  # Each date param occupies 2 digits (4 for year)
            regex = regex + new_reg
        match = re.match(regex, string)
        if not match:
            match = re.match(regex, string.replace(":", sep))
            if not match:
                raise ValueError(f"Could not match date to given string: {string}")
        matchdate = match.group(0)
        # Getting the equivalent format from given digits
        full_date = "%Y%m%d%H%M%S"[0 : digits - 2]
        datepattern = f"{sep}%".join(full_date.split("%")).strip(sep)
        res_date = datetime.strptime(matchdate, datepattern)
        return res_date

    seps = ["", " ", "/", "-", "_"]
    digits_list = [x for x in range(4, 16, 2)]
    combinations = sorted(product(digits_list, seps), reverse=True)
    for digits, sep in combinations:
        try:
            res_date = rec_date_extraction(string, digits, sep)
        except ValueError:
            continue
        return res_date
    else:
        return None


def excel_date_to_num(date):
    """Transform a date object formatted by excel to a numeric value"""
    try:
        return date.toordinal() - datetime(1899, 12, 30).toordinal()
    except AttributeError:
        return None


def read_csv_file_return_dict(file_name, sep=None, key_position=0):
    """Read csv or tsv file, according to separator (sep), and return a dictionary
    where the main key is the first column, if key position is None otherwise
    the index value of the key position is used as key. If sep is None then
    try to assert a separator automatically depending on file extension.
    """
    if sep is None:
        file_extension = os.path.splitext(file_name)[1]
        extdict = {".csv": ",", ".tsv": "\t", ".tab": "\t"}
        # Use space as a default separator, None would also be valid
        sep = extdict.get(file_extension, " ")
    try:
        # Read all columns as strings to avoid parsing IDs as float buy try to infer datatypes afterwards
        file_df = pd.read_csv(file_name, sep=sep, dtype="string").convert_dtypes()
    except FileNotFoundError:
        raise
    key_column = file_df.columns[key_position]
    file_data = file_df.set_index(key_column).to_dict(orient="index")
    return file_data


def read_fasta_return_SeqIO_instance(file_name):
    """Read fasta and return SeqIO instance"""
    try:
        return SeqIO.read(file_name, "fasta")
    except FileNotFoundError:
        raise


def read_yml_file(file_name):
    """Read yml file"""
    file_name = os.path.expanduser(file_name)
    with open(file_name, "r") as fh:
        try:
            return yaml.safe_load(fh)
        except yaml.YAMLError:
            raise


def get_md5_from_local_folder(local_folder):
    """Fetch the md5 values for each file in the file list"""
    md5_results = {}
    reg_for_md5 = os.path.join(local_folder, "*.md5")
    # reg_for_non_md5 = os.path.join(local_folder, "*[!.md5]")
    md5_files = glob.glob(reg_for_md5)
    if not md5_files:
        return False
    else:
        for md5_file in md5_files:
            file_path_name, f_ext = os.path.splitext(md5_file)
            if not file_exists(file_path_name):
                log.error("Found md5 file but not %s", file_path_name)
                continue
            file_name = os.path.basename(file_path_name)
            fh = open(md5_file, "r")
            md5_results[file_name] = fh.read()
            fh.close()
    return md5_results


def read_md5_checksum(file_name, avoid_chars=list()):
    """Read MD5_checksum file and return a dict of {file: md5_hash}

    Args:
        file_name (str): file containing "md5hash  file" in tab separated format
        avoid_chars (list(str), optional): Lines with any of these elements
        will be skipped. Defaults to list().

    Returns:
        hash_dict(dict): dictionary of {file: md5_hash}
    """
    try:
        with open(file_name, "r") as file:
            content = file.read()
    except FileNotFoundError:
        raise
    clean_content = content.replace("*", "")
    lines = clean_content.splitlines()
    translation = str.maketrans("", "", "'\"")
    if any("\t" in line for line in lines):
        lines = [line.strip().translate(translation).split("\t") for line in lines]
    elif any("," in line for line in lines):
        lines = [line.strip().translate(translation).split(",") for line in lines]
    else:
        lines = [line.strip().translate(translation).split() for line in lines]
    clean_lines = [
        x for x in lines if not any(ch in string for ch in avoid_chars for string in x)
    ]
    # md5sum should always have 2 columns: hash - path
    md5_lines = [line for line in clean_lines if len(line) == 2]
    if not md5_lines:
        return False
    # split paths for both windows "\" and linux "/" using regex [\\/]
    hash_dict = {re.split(r"[\\/]", line[1])[-1]: line[0].lower() for line in md5_lines}
    return hash_dict


def delete_local_folder(folder):
    """Delete download folder because files does not complain requisites"""
    log.info("Deleting local folder %s", folder)
    shutil.rmtree(folder, ignore_errors=True)
    return True


def calculate_md5(file_name):
    """Calculate the md5 value for the file name"""
    return hashlib.md5(open(file_name, "rb").read()).hexdigest()


def write_md5_file(file_name, md5_value):
    """Write md5 to file"""
    with open(file_name, "w") as fh:
        fh.write(md5_value)
    return


def create_md5_files(local_folder, file_list):
    """Create the md5 files and return their value"""
    md5_results = {}
    for file_name in file_list:
        md5_results[file_name] = [
            local_folder,
            calculate_md5(os.path.join(local_folder, file_name)),
        ]
        md5_file_name = file_name + ".md5"
        write_md5_file(
            os.path.join(local_folder, md5_file_name), md5_results[file_name][1]
        )
    return md5_results


def save_local_md5(file_name, md5_value):
    """Save the MD5 value"""
    with open(file_name, "w") as fh:
        fh.write(md5_value)
    return True


def write_json_to_file(data, file_name):
    """Write metadata to json file"""
    with open(file_name, "w", encoding="utf-8") as fh:
        fh.write(json.dumps(data, indent=4, sort_keys=True, ensure_ascii=False))
    return True


def compress_file(file):
    """compress a given file with gzip, adding .gz extension afterwards

    Args:
        file (str): path to the given file
    """
    try:
        with open(file, "rb") as raw, gzip.open(f"{file}.gz", "wb") as comp:
            comp.writelines(raw)
        return True
    except FileNotFoundError:
        return False


def check_gzip_integrity(file_path):
    """Check if a compressed file can be decompressed"""
    chunksize = 100000000  # 10 Mbytes
    with gzip.open(file_path, "rb") as f:
        try:
            while f.read(chunksize) != b"":
                pass
        except Exception:
            # Not a gzip file
            return False
        # EOFError: Compressed file is truncated
        except EOFError:
            return False
    return True


def lower_keys(data):
    """Transform all keys to lowercase strings in a dictionary"""
    return {str(key).lower(): v for key, v in data.items()}


def rich_force_colors():
    """
    Check if any environment variables are set to force Rich to use coloured output
    """
    if (
        os.getenv("GITHUB_ACTIONS")
        or os.getenv("FORCE_COLOR")
        or os.getenv("PY_COLORS")
    ):
        return True
    return None


stderr = Console(
    stderr=True, style="dim", highlight=False, force_terminal=rich_force_colors()
)


def prompt_text(msg):
    source = questionary.text(msg).unsafe_ask()
    return source


def prompt_password(msg):
    source = questionary.password(msg).unsafe_ask()
    return source


def prompt_tmp_dir_path():
    stderr.print("Temporal directory destination to execute service")
    source = questionary.path("Source path").unsafe_ask()
    return source


def prompt_selection(msg, choices):
    selection = questionary.select(msg, choices=choices).unsafe_ask()
    return selection


def prompt_path(msg):
    source = questionary.path(msg).unsafe_ask()
    return source


def prompt_yn_question(msg):
    confirmation = questionary.confirm(msg).unsafe_ask()
    return confirmation


def prompt_skip_folder_creation():
    stderr.print("Do you want to skip folder creation? (Y/N)")
    confirmation = questionary.confirm("Skip?", default=False).unsafe_ask()
    return confirmation


def prompt_checkbox(msg, choices):
    selected_options = questionary.checkbox(msg, choices=choices).unsafe_ask()
    return selected_options


def get_file_date(file_path):
    """Get the modification date of a file."""
    try:
        # Get the modification time of the file
        mtime = os.path.getmtime(file_path)
        # Convert the modification time to a datetime object
        file_date = datetime.fromtimestamp(mtime)
        # Format date
        formatted_date = file_date.strftime("%Y%m%d")
        return formatted_date
    except FileNotFoundError:
        # Handle file not found error
        print(f"File not found: {file_path}")
        return None


def select_most_recent_files_per_sample(paths_list):
    """Selects the most recent file for each sample among potentially duplicated files.
    Input:
        - paths_list: a list of sample's file paths.
    Output:
        - List of file paths containig the most recent/up-to-date file for each sample.
    """
    filename_groups = {}
    # Count occurrences of each filename and group files by sample names
    for file in paths_list:
        # TODO: So far, it uses split method to identify this pattern:
        # [sample1.pangolin.csv, sample1.pangolin_20240310.csv]. It should be
        # improve to parse files based on a different character matching field.
        file_name = os.path.basename(file).split(".")[0]
        if file_name in filename_groups:
            filename_groups[file_name].append(file)
        else:
            filename_groups[file_name] = [file]
    # Filter out sample names with only one file
    duplicated_files = [
        (sample_name, file_paths)
        for sample_name, file_paths in filename_groups.items()
        if len(file_paths) > 1
    ]
    # Iterate over duplicated files to select the most recent one for each sample
    for sample_name, file_paths in duplicated_files:
        stderr.print(
            f"\tMore than one file found for sample {sample_name}. Selecting the most recent one."
        )
        # Sort files by modification time (most recent first)
        sorted_files = sorted(
            file_paths, key=lambda file_path: os.path.getmtime(file_path), reverse=True
        )
        # Select the most recent file
        selected_file = sorted_files[0]
        stderr.print(f"\tSelected file for sample {sample_name}: {selected_file}")
        # Remove other files for the same sample from the filtered_files dictionary
        filename_groups[sample_name] = [selected_file]
    # Update filename_groups with filtered files
    filename_groups = [
        (sample_name, file_path)
        for sample_name, file_paths in filename_groups.items()
        for file_path in file_paths
    ]
    return [sample_file_path for _, sample_file_path in filename_groups]


def print_log_report(
    log_report, categories=None, sections=["warning", "valid", "error"]
):
    color_codes = {
        "error": "\033[91m",  # Red
        "warning": "\033[93m",  # Orange
        "valid": "\033[92m",  # Green
        "reset": "\033[0m",  # Reset color
    }
    table_data = []
    for section_name, section_data in log_report.items():
        if section_name in sections:
            for category, items in section_data.items():
                if categories is None or category in categories:
                    colored_category = (
                        f"{color_codes[section_name]}{category}{color_codes['reset']}"
                    )
                    for item in items:
                        colored_message = (
                            f"{color_codes[section_name]}{item}{color_codes['reset']}"
                        )
                        table_data.append(
                            [section_name, colored_category, colored_message]
                        )
    print(
        tabulate(
            table_data,
            headers=["Log type", "Category", "Message"],
            tablefmt="presto",
        )
    )


def prompt_create_outdir(
    path=None, folder_name=None, prompt_message="Define path to store the output:"
):
    """Ensure the directory exists or prompt the user to define and create it."""
    # Check path
    if not path:
        path = prompt_path(prompt_message)
        stderr.print(f"Chosen directory: {path}")

    # Check folder_name
    if not folder_name:
        default_folder = prompt_yn_question(
            "Do you want to use the default directory ('results') to store the results? (yes/no):"
        )
        if not default_folder:
            folder_name = prompt_text("Write your output directory: ")
        else:
            folder_name = "results"

    # Prevent duplicate folder names in the path
    if os.path.basename(path) == folder_name:
        global_path = path
    else:
        global_path = os.path.join(path, folder_name)

    if not os.path.exists(global_path):
        create_folder = prompt_yn_question(
            f"The directory does not exist. Do you want to create '{folder_name}' folder in this path? (yes/no):"
        )
        if create_folder:
            os.makedirs(global_path)
            stderr.print(f"[green]Folder '{folder_name}' created at {path}")
        else:
            stderr.print("[red]Directory creation aborted.")
            sys.exit(1)
    elif os.path.isdir(global_path):
        os.makedirs(global_path, exist_ok=True)
        stderr.print(f"[green]Defining '{folder_name}' as output folder")
    else:
        stderr.print("[red]The provided path is not a directory.")
        sys.exit(1)

    return global_path


def adjust_sheet_size(sheet, wrap_text=True, col_width=30):
    """Adjust column width and row heights depending on the max number of
    characters in each one.

    Args:
        sheet (openpyxl.worksheet): active openpyxl worksheet object
        wrap_text (bool): Wether to use excel wrap_text function for each cell. Defaults to True
        col_width (int): Minimum columns width value. Also used to define maximum
        number of characters in each cell when wrap_text is True. Defaults to 30.
    """
    dims = {}
    for _, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row), start=2):
        for cell in row:
            if wrap_text:
                cell.alignment = openpyxl.styles.Alignment(wrapText=True)
            if cell.value is not None:
                max_length = max((dims.get(cell.column, 0), len(str(cell.value))))
                dims[cell.column] = max_length / col_width
    for col_num, value in dims.items():
        if value < col_width:
            value = col_width
        sheet.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = value


def validate_semantic_version(version):
    try:
        ver = semantic_version.Version(version)
        return ver
    except ValueError:
        return None


def get_package_name():
    """Get project name"""
    try:
        package_name = importlib.metadata.metadata(__name__.split(".")[0])["Name"]
        return package_name
    except importlib.metadata.PackageNotFoundError:
        return "unknown_package"


def get_git_branch():
    """Get current git branch"""
    try:
        branch = (
            subprocess.check_output(["git", "rev-parse", "--abbrev-ref", "HEAD"])
            .strip()
            .decode("utf-8")
        )
        return branch
    except Exception:
        return "main"  # if not able to retrieve git branch, add fixed value


def get_schema_url():
    """Generates the schema url dinamically"""
    package_name = get_package_name()
    branch_name = get_git_branch()
    base_url = "https://github.com/BU-ISCIII"
    schema_path = f"{package_name}/blob/{branch_name}/{package_name.replace('-', '_')}/schema/relecov_schema.json"

    return f"{base_url}/{schema_path}"


def display_dataframe_to_user(name: str, dataframe: pd.DataFrame):
    """
    Display a Pandas DataFrame in a formatted table using Rich.

    Args:
        name (str): Title of the table.
        dataframe (pd.DataFrame): The DataFrame to display.
    """
    console = Console()

    # If DataFrame is empty, show a message
    if dataframe.empty:
        console.print(f"[bold red]{name} - No Data Available[/bold red]")
        return

    # Create a Rich Table
    table = Table(title=name, show_lines=True)

    # Add columns
    for col in dataframe.columns:
        table.add_column(col, justify="left", style="cyan", no_wrap=True)

    # Add rows
    for _, row in dataframe.iterrows():
        table.add_row(*[str(value) for value in row])

    # Display the table
    console.print(table)


def load_schema(schema_path: str) -> dict:
    """
    Load a JSON schema from the specified file path.

    Args:
        schema_path (str): The file path to the JSON schema.

    Returns:
        dict: Parsed schema as a Python dictionary.
    """
    with open(schema_path) as f:
        return json.load(f)


def get_available_software(json_path: str) -> list:
    """
    Retrieve available software names from a bioinfo configuration JSON file.

    Args:
        json_path (str): Path to the bioinfo configuration file.

    Returns:
        list: A list of available software/tools defined in the configuration.
    """
    config = read_json_file(json_path)
    return list(config.keys())


def cast_value_to_schema_type(value, expected_type: str):
    """
    Cast a value to the expected JSON schema type.

    Args:
        value (any): The input value to be cast.
        expected_type (str): Target data type from the schema. Options: "integer", "number", "boolean", "string".

    Returns:
        any: The value cast to the appropriate type, or a string fallback if casting fails.
    """
    if expected_type == "integer":
        try:
            return int(float(value))
        except (ValueError, TypeError):
            return str(value).strip()
    elif expected_type == "number":
        try:
            return float(value)
        except (ValueError, TypeError):
            return str(value).strip()
    elif expected_type == "boolean":
        return str(value).strip().lower() in ["true", "yes", "1"]
    elif expected_type == "string":
        return str(value).strip()
    else:
        try:
            return str(value).strip()
        except Exception:
            return value


def get_safe_hex(output_dir, length=3):
    """Return an unique hexadecimal code that does not repeat in any file
    of the given output directory

    Args:
        output_dir (str): Folder where files will be checked
        length (int): Number of bytes for the code. Character length will be doubled

    Returns:
        hex_id (str): Non-repeating hexadecimal code
    """

    def get_new_hex(hex_id, output_dir):
        """Recursive search for new hexadecimal codes if exist"""
        if any(hex_id in x for x in os.listdir(output_dir)):
            return get_new_hex(token_hex(length).upper(), output_dir)
        else:
            return token_hex(length).upper()

    hex_id = get_new_hex(token_hex(length).upper(), output_dir)
    return hex_id


def generate_fingerprint(
    sequencing_sample_id,
    collecting_lab_sample_id,
    submitting_institution,
    collecting_institution,
):
    combined = f"{sequencing_sample_id}|{collecting_lab_sample_id}|{submitting_institution}|{collecting_institution}".lower()
    return hashlib.sha256(combined.encode()).hexdigest()[:24]
