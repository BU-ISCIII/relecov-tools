#!/usr/bin/env python
import logging
import rich.console
import pandas as pd


import relecov_tools.utils
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import column_index_from_string
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill


log = logging.getLogger(__name__)
stderr = rich.console.Console(
    stderr=True,
    style="dim",
    highlight=False,
    force_terminal=relecov_tools.utils.rich_force_colors(),
)


def schema_to_flatten_json(json_data, required_properties=None):
    """Return the schema flattened to a list while keeping parent metadata."""
    try:
        required_set = set(required_properties or [])
        flatten_rows = []
        for property_id, features in json_data.items():
            try:
                is_complex_array = (
                    isinstance(features, dict)
                    and features.get("type") == "array"
                    and features.get("items", {}).get("type") == "object"
                )
                if is_complex_array:
                    items_schema = features.get("items", {})
                    complex_properties = items_schema.get("properties", {})
                    required_list = items_schema.get(
                        "required", features.get("required", [])
                    )
                    complex_required = set(required_list)
                    for (
                        complex_property_id,
                        complex_feature,
                    ) in complex_properties.items():
                        row = dict(complex_feature)
                        row["property_id"] = f"{property_id}.{complex_property_id}"
                        row["field_id"] = complex_property_id
                        row["parent_property_id"] = property_id
                        row["parent_label"] = features.get("label", "")
                        row["parent_classification"] = features.get(
                            "classification", ""
                        )
                        row["is_required"] = complex_property_id in complex_required
                        flatten_rows.append(row)
                else:
                    row = dict(features)
                    row["property_id"] = property_id
                    row["field_id"] = property_id
                    row["parent_property_id"] = None
                    row["parent_label"] = ""
                    row["parent_classification"] = ""
                    row["is_required"] = property_id in required_set
                    flatten_rows.append(row)
            except Exception as e:
                stderr.print(f"[red]Error processing property {property_id}: {e}")
        return flatten_rows
    except Exception as e:
        stderr.print(f"[red]Error in schema_to_flatten_json: {e}")
        return []


def schema_properties_to_df(json_data):
    try:
        if json_data is None:
            return pd.DataFrame()
        if isinstance(json_data, dict):
            rows = []
            for property_id, property_features in json_data.items():
                try:
                    row = {"property_id": property_id}
                    row.update(property_features)
                    rows.append(row)
                except Exception as e:
                    stderr.print(f"[red]Error processing property {property_id}: {e}")
            return pd.DataFrame(rows)

        if isinstance(json_data, list):
            return pd.DataFrame(json_data)

        stderr.print("[yellow]schema_properties_to_df received unsupported data.")
        return pd.DataFrame()
    except Exception as e:
        stderr.print(f"[red]Error in schema_properties_to_df: {e}")
        return None


def excel_formater(df, writer, sheet, out_file, have_index=True, have_header=True):
    try:

        # Write the DataFrame to the specified sheet
        df.to_excel(
            writer, sheet_name=sheet, startrow=1, index=have_index, header=have_header
        )

        # Get the xlsxwriter workbook and worksheet objects.
        workbook = writer.book
        worksheet = writer.sheets[sheet]

        # Set up general column width
        worksheet.set_column(0, len(df.columns), 30)

        # General header format
        header_formater = workbook.add_format(
            {
                "bold": True,
                "text_wrap": False,
                "valign": "top",
                "fg_color": "#B9DADE",  # Light blue
                "border": 1,
                "locked": True,
            }
        )

        # Custom header format for METADATA_LAB (red text starting from column 2)
        red_header_formater = workbook.add_format(
            {
                "bold": True,
                "text_wrap": False,
                "valign": "top",
                "fg_color": "#B9DADE",  # Light blue background
                "color": "#f60606",  # Red text color
                "border": 1,
                "locked": True,
            }
        )

        # First column format
        first_col_formater = workbook.add_format(
            {
                "bold": True,
                "text_wrap": False,
                "valign": "center",
                "fg_color": "#B9DADE",  # Light blue
                "border": 1,
                "locked": True,
            }
        )

        cell_formater = workbook.add_format(
            {
                "border": 1,  # Apply border to every cell
                "locked": True,
            }
        )

        if sheet == "OVERVIEW":
            # Write the column headers with the defined format.
            for col_num, value in enumerate(df.columns.values):
                try:
                    worksheet.write(0, col_num, value, header_formater)
                except Exception as e:
                    stderr.print(f"Error writing header at column {col_num + 1}: {e}")

            # Write the first column with the defined format.
            for row_num in range(1, len(df) + 1):
                try:
                    worksheet.write(
                        row_num, 0, df.iloc[row_num - 1, 0], first_col_formater
                    )
                except Exception as e:
                    stderr.print(f"Error writing first column at row {row_num}: {e}")

            for row_num in range(1, len(df) + 1):
                for col_num in range(1, len(df.columns)):
                    try:
                        worksheet.write(
                            row_num,
                            col_num,
                            df.iloc[row_num - 1, col_num],
                            cell_formater,
                        )
                    except Exception as e:
                        stderr.print(
                            f"Error writing cell at row {row_num}, column {col_num}: {e}"
                        )

        if sheet == "METADATA_LAB" or sheet == "DATA_VALIDATION":
            # Write the column headers with the defined format.
            for col_num in range(0, len(df.columns)):
                for row_num in range(0, len(df)):
                    if row_num < 4:
                        try:
                            worksheet.write(
                                row_num,
                                col_num + 1,
                                df.iloc[row_num, col_num],
                                header_formater,
                            )
                        except Exception as e:
                            stderr.print(
                                f"Error writing first column at row {row_num}: {e}"
                            )
                        if row_num == 0 and col_num >= 0 and sheet == "METADATA_LAB":
                            try:
                                worksheet.write(
                                    row_num,
                                    col_num + 1,
                                    df.iloc[row_num, col_num],
                                    red_header_formater,
                                )
                            except Exception as e:
                                stderr.print(
                                    f"Error writing first row at column {col_num}: {e}"
                                )
            # Write the first column with the defined format.
            for index_num, index_val in enumerate(df.index):
                try:
                    worksheet.write(index_num, 0, index_val, first_col_formater)
                except Exception as e:
                    stderr.print(f"Error writing first column at row {row_num}: {e}")
    except Exception as e:
        stderr.print(f"Error in excel_formater: {e}")


def create_condition(ws_metadata, conditions, df_filtered):
    """This function creates conditions on METADATA_LAB template sheet"""
    label_to_property = dict(zip(df_filtered["label"], df_filtered["property_id"]))
    column_map = {}

    for cell in ws_metadata[4]:
        property_id = label_to_property.get(cell.value)
        if property_id in conditions:
            column_map[property_id] = cell.column_letter

    for property_id, rules in conditions.items():
        col_letter = column_map.get(property_id)
        if col_letter:
            start_row = rules.get("header_row_idx", 5)
            end_row = rules.get("max_rows", 1000)
            validation_type = rules.get("validation_type")
            formula1 = rules.get("formula1", "").replace("{col_letter}", col_letter)
            formula2 = rules.get("formula2")
            error_message = rules.get("error_message", "Valor no permitido")
            error_title = rules.get("error_title", "Error")

            cell_range = f"{col_letter}{start_row}:{col_letter}{end_row}"

            if validation_type:
                validation = DataValidation(
                    type=validation_type,
                    operator=rules.get("operator", "between"),
                    formula1=formula1,
                    formula2=formula2,
                    showErrorMessage=True,
                )
                validation.error = error_message
                validation.errorTitle = error_title
                ws_metadata.add_data_validation(validation)
                validation.add(cell_range)

            if rules.get("format_cells_as_date", False):
                col_idx = column_index_from_string(col_letter)
                for row in ws_metadata.iter_rows(
                    min_row=start_row, max_row=end_row, min_col=col_idx, max_col=col_idx
                ):
                    for cell in row:
                        cell.number_format = "yyyy-mm-dd"
    return ws_metadata


def add_conditional_format_age_check(
    ws_metadata,
    df_filtered,
    prop1="host_age_years",
    prop2="host_age_months",
    start_row=5,
    end_row=1000,
):
    """
    Applies a conditional formatting that marks in red if both cells are filled in
    (host_age_years and host_age_months) in the same row.
    """

    label_to_property = dict(zip(df_filtered["label"], df_filtered["property_id"]))

    column_map = {}
    for cell in ws_metadata[4]:
        property_id = label_to_property.get(cell.value)
        if property_id:
            column_map[property_id] = cell.column_letter

    col1_letter = column_map.get(prop1)
    col2_letter = column_map.get(prop2)
    if not col1_letter or not col2_letter:
        return

    red_fill = PatternFill(start_color="F54627", end_color="F54627", fill_type="solid")
    for row in range(start_row, end_row + 1):
        formula = f"=AND(NOT(ISBLANK(${col1_letter}{row})), NOT(ISBLANK(${col2_letter}{row})))"
        rule = FormulaRule(formula=[formula], fill=red_fill)

        for col in [col1_letter, col2_letter]:
            ws_metadata.conditional_formatting.add(f"{col}{row}", rule)
