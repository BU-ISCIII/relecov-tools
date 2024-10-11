#!/usr/bin/env python
import copy
import logging
import json
import sys
import os
import yaml
import warnings
import rich.console
import paramiko
import relecov_tools.utils
import relecov_tools.sftp_client
from datetime import datetime
from itertools import islice
from secrets import token_hex
from csv import writer as csv_writer, Error as CsvError
from openpyxl import load_workbook as openpyxl_load_workbook
from pandas import read_excel, ExcelWriter, concat
from pandas.errors import ParserError, EmptyDataError
from relecov_tools.config_json import ConfigJson
from relecov_tools.log_summary import LogSum


# from relecov_tools.rest_api import RestApi

log = logging.getLogger(__name__)
stderr = rich.console.Console(
    stderr=True,
    style="dim",
    highlight=False,
    force_terminal=relecov_tools.utils.rich_force_colors(),
)


class MetadataError(Exception):
    def __init__(self, message):
        super().__init__(message)


class DownloadManager:
    def __init__(
        self,
        user=None,
        passwd=None,
        conf_file=None,
        download_option=None,
        output_location=None,
        target_folders=None,
    ):
        """Initializes the sftp object"""
        config_json = ConfigJson()
        self.allowed_file_ext = config_json.get_topic_data(
            "sftp_handle", "allowed_file_extensions"
        )
        sftp_user = user
        sftp_passwd = passwd
        self.target_folders = target_folders
        self.allowed_download_options = config_json.get_topic_data(
            "sftp_handle", "allowed_download_options"
        )
        if download_option not in self.allowed_download_options:
            self.download_option = relecov_tools.utils.prompt_selection(
                "Options", self.allowed_download_options
            )
        else:
            self.download_option = download_option
        if conf_file is None:
            # self.sftp_server = config_json.get_topic_data("sftp_handle", "sftp_server")
            # self.sftp_port = config_json.get_topic_data("sftp_handle", "sftp_port")
            self.platform_storage_folder = config_json.get_topic_data(
                "sftp_handle", "platform_storage_folder"
            )
            self.abort_if_md5_mismatch = (
                True
                if config_json.get_topic_data("sftp_handle", "abort_if_md5_mismatch")
                == "True"
                else False
            )
        else:
            if not os.path.isfile(conf_file):
                log.error("Configuration file %s does not exists", conf_file)
                stderr.print(
                    "[red] Configuration file does not exist. " + conf_file + "!"
                )
                sys.exit(1)
            with open(conf_file, "r") as fh:
                config = yaml.load(fh, Loader=yaml.FullLoader)
            try:
                # self.sftp_server = config["sftp_server"]
                # self.sftp_port = config["sftp_port"]
                self.target_folders = config["target_folders"]
                try:
                    self.platform_storage_folder = config["platform_storage_folder"]
                except KeyError:
                    self.platform_storage_folder = config_json.get_topic_data(
                        "sftp_handle", "platform_storage_folder"
                    )
                sftp_user = config["sftp_user"]
                sftp_passwd = config["sftp_passwd"]
            except KeyError as e:
                log.error("Invalid configuration file %s", e)
                stderr.print(f"[red] Invalid configuration file {e} !")
                sys.exit(1)
        if output_location is not None:
            if os.path.isdir(output_location):
                self.platform_storage_folder = os.path.realpath(output_location)
            else:
                log.error("Output location does not exist, aborting")
                stderr.print("[red] Output location does not exist, aborting")
                sys.exit(1)
        if sftp_user is None:
            sftp_user = relecov_tools.utils.prompt_text(msg="Enter the user id")
        if isinstance(self.target_folders, str):
            self.target_folders = self.target_folders.split(",")
        self.logsum = LogSum(output_location=self.platform_storage_folder)
        if sftp_passwd is None:
            sftp_passwd = relecov_tools.utils.prompt_password(msg="Enter your password")
        self.metadata_lab_heading = config_json.get_topic_data(
            "lab_metadata", "metadata_lab_heading"
        )
        self.metadata_processing = config_json.get_topic_data(
            "sftp_handle", "metadata_processing"
        )
        self.avoidable_characters = config_json.get_topic_data(
            "sftp_handle", "skip_when_found"
        )
        self.samples_json_fields = config_json.get_topic_data(
            "lab_metadata", "samples_json_fields"
        )
        # initialize the sftp client
        self.relecov_sftp = relecov_tools.sftp_client.SftpRelecov(
            conf_file, sftp_user, sftp_passwd
        )
        self.finished_folders = {}

    def create_local_folder(self, folder):
        """Create folder to download files in local path using date

        Args:
            folder (str): name of remote folder to be downloaded

        Returns:
            local_folder_path(str): path to the new created folder
        """
        log.info("Creating folder %s to download files", folder)
        platform_storage_folder = self.platform_storage_folder
        if platform_storage_folder == folder:
            local_folder_path = platform_storage_folder
        else:
            folder = folder.strip("_tmp_processing")
            local_folder_path = os.path.join(platform_storage_folder, folder)
        os.makedirs(local_folder_path, exist_ok=True)
        log.info("created the folder to download files %s", local_folder_path)
        return local_folder_path

    def get_remote_folder_files(self, folder, local_folder, file_list):
        """Create the subfolder with the present date and fetch all files from
        the remote sftp server

        Args:
            folder (str): name of remote folder to be downloaded
            local_folder (str): name of local folder to store downloaded files
            file_list (list(str)): list of files in remote folder to be downloaded

        Returns:
            fetched_files(list(str)): list of successfully downloaded files
        """

        fetched_files = list()
        log.info("Trying to fetch files in remote server")
        stderr.print(f"Fetching {len(file_list)} files from {folder}")
        for file in file_list:
            file_to_fetch = os.path.join(folder, os.path.basename(file))
            output_file = os.path.join(local_folder, os.path.basename(file))
            if self.relecov_sftp.get_from_sftp(
                file_to_fetch, output_file, exist_ok=True
            ):
                fetched_files.append(os.path.basename(file))
            else:
                # Try to download again n times
                for _ in range(3):
                    if self.relecov_sftp.get_from_sftp(file_to_fetch, output_file):
                        fetched_files.append(os.path.basename(file))
                        break
                else:
                    log.warning("Couldn't fetch %s from %s after 3 tries", file, folder)
        return fetched_files

    def find_remote_md5sum(self, folder, pattern="md5sum"):
        """Search for a pattern in remote folder, by default is md5sum

        Args:
            folder (str): folder path in remote repository
            pattern (str, optional): Regex used to find file. Defaults to "md5sum".

        Returns:
            md5_file(str): file basename if found. If not found returns False
        """
        all_files = self.relecov_sftp.get_file_list(folder)
        md5_file = [file for file in all_files if pattern in file]
        if len(md5_file) == 1:
            return md5_file[0]
        else:
            return False

    def verify_md5_checksum(self, local_folder, fetched_files, fetched_md5):
        """Check if the md5 value from sftp matches with the one generated locally"""
        required_retransmition = []
        successful_files = []
        # fetch the md5 file if exists
        log.info("Searching for local md5 file")
        stderr.print("[blue]Verifying file integrity in md5 hashes")
        avoid_chars = self.avoidable_characters
        hash_dict = relecov_tools.utils.read_md5_checksum(fetched_md5, avoid_chars)
        if not hash_dict:
            error_text = "md5sum file could not be read, md5 hashes won't be validated"
            self.include_warning(error_text)
            return fetched_files, False
        # check md5 checksum for each file
        for f_name in hash_dict.keys():
            if f_name not in fetched_files:
                # Skip those files in md5sum that were not downloaded by any reason
                continue
            f_path = os.path.join(local_folder, f_name)
            if hash_dict[f_name] == relecov_tools.utils.calculate_md5(f_path):
                successful_files.append(f_name)
                log.info("Successful file download for %s", f_name)
            else:
                required_retransmition.append(f_name)
                log.warning("%s requested file re-sending", f_name)
        return successful_files, required_retransmition

    def create_files_with_metadata_info(
        self, local_folder, samples_dict, md5_dict, metadata_file
    ):
        """Copy metadata file from folder, extend samples_dict with md5hash for
        each file. Then create a Json file with this dict

        Args:
            local_folder (str): Path to folder with downloaded files and output
            samples_dict (dict{str:str}): same structure as validate_remote_files()
            md5_dict (dict(str:str)): Zipped dict of files_list and md5hash_list
            metadata_file (str): Name of the downloaded metadata file to rename it
        """
        samples_to_delete = []
        prefix_file_name = "_".join(local_folder.split("/")[-2:])
        prefix_file_name = prefix_file_name.replace("_tmp_processing", "")
        # TODO: Move these prefixes to configuration.json
        new_metadata_file = "lab_metadata_" + prefix_file_name + ".xlsx"
        sample_data_file = "samples_data_" + prefix_file_name + ".json"
        sample_data_path = os.path.join(local_folder, sample_data_file)
        os.rename(metadata_file, os.path.join(local_folder, new_metadata_file))
        error_text = "Sample %s incomplete. Not added to final Json"

        data = copy.deepcopy(samples_dict)
        for sample, values in data.items():
            if not all(val for val in values):
                self.include_error(str(error_text % sample), sample)
                samples_to_delete.append(sample)
                continue
            # TODO: Move these keys to configuration.json
            values["r1_fastq_filepath"] = local_folder
            values["fastq_r1_md5"] = md5_dict.get(values["sequence_file_R1_fastq"])
            if values.get("sequence_file_R2_fastq"):
                values["r2_fastq_filepath"] = local_folder
                values["fastq_r2_md5"] = md5_dict.get(values["sequence_file_R2_fastq"])
        if samples_to_delete:
            data = {k: v for k, v in data.items() if k not in samples_to_delete}
        with open(sample_data_path, "w", encoding="utf-8") as fh:
            fh.write(json.dumps(data, indent=4, sort_keys=True, ensure_ascii=False))
        # Feed accessible dict with necessary information for wrapper to work
        log.info("Successfully created samples json file %s", sample_data_path)
        return

    def remove_duplicated_values(self, sample_file_dict):
        """remove keys that share the same value due to duplication in sample_dict

        Args:
            sample_file_dict (dict(str:dict(str:str))): dictionary with sample_name
            as keys and a dict for both R1 filename and/or R2 if paired-end reads
            and fastq-file paths. e.g. {sample1:{r1_fastq_filepath:sample1.fastq.gz}}

        Returns:
            clean_sample_dict: sample_dictionary without duplications in values
        """
        inverted_dict = {}
        for sample, fastq_dict in sample_file_dict.items():
            # Dictionary values are not hashable, so you need to create a tuple of them
            samp_fastqs = tuple(fastq_dict.values())
            # Setting values as keys to find those samples refering to the same file
            for fastq in samp_fastqs:
                inverted_dict[fastq] = inverted_dict.get(fastq, []) + [sample]
        duplicated_dict = {k: v for k, v in inverted_dict.items() if len(v) > 1}
        dup_samples_list = [samp for dups in duplicated_dict.values() for samp in dups]
        non_duplicated_keys = {
            k: v for k, v in sample_file_dict.items() if k not in dup_samples_list
        }
        clean_sample_dict = {key: sample_file_dict[key] for key in non_duplicated_keys}
        if dup_samples_list:
            error_text = "Multiple samples in metadata pointing to the same file: %s"
            self.include_warning(error_text % duplicated_dict)
            stderr.print(f"[Orange]{error_text}")
            stderr.print("[Orange]These samples won't be processed: ", dup_samples_list)
            for fastq, samples in duplicated_dict.items():
                [self.include_error(str(error_text % fastq), samp) for samp in samples]

        return clean_sample_dict

    def read_metadata_file(self, meta_f_path, return_data=True):
        """Read excel file, check if the header matches with the one defined in config

        Args:
            meta_f_path (str): Path to the excel_file

        Raises:
            MetadataError: If the header in the excel is different from config

        Returns:
            ws_metadata_lab: openpyxl's workbook metadata sheet of the excel file
            metadata_header: column names of the header
            header_row: row where the header is located in the sheet
        """
        warnings.simplefilter(action="ignore", category=UserWarning)
        wb_file = openpyxl_load_workbook(meta_f_path, data_only=True)
        ws_metadata_lab = wb_file[self.metadata_processing.get("excel_sheet")]
        # find out the index for file names
        header_flag = self.metadata_processing.get("header_flag")
        try:
            header_row = [
                i + 1 for i, x in enumerate(ws_metadata_lab.values) if header_flag in x
            ][0]
        except IndexError:
            error_text = "Header could not be found for excel file %s"
            raise MetadataError(str(error_text % os.path.basename(meta_f_path)))
        for cell in ws_metadata_lab[header_row]:
            cell.value = cell.value.strip()
        metadata_header = [x.value for x in ws_metadata_lab[header_row]]
        meta_column_list = self.metadata_lab_heading
        if meta_column_list != metadata_header[1:]:
            diffs = [
                x
                for x in set(metadata_header[1:] + meta_column_list)
                if x not in meta_column_list or x not in metadata_header
            ]
            log.error(
                "Config field metadata_lab_heading is different from .xlsx header"
            )
            stderr.print(
                "[red]Header in metadata file is different from config file, aborting"
            )
            stderr.print("[red]Differences: ", diffs)
            raise MetadataError(f"Metadata header different from config: {diffs}")
        if return_data:
            return ws_metadata_lab, metadata_header, header_row
        else:
            return True

    def get_sample_fastq_file_names(self, local_folder, meta_f_path):
        """Read excel metadata template and create dictionary with files for each sample

        Args:
            local_folder (str): folder where the excel file has been downloaded
            meta_f_path (str): path to the downloaded excel file with metadata

        Returns:
            clean_sample_dict(dict(str:{str:str})): Nested dictionary for each sample
            {sample1: {"sequence_file_R1_fastq": "sample1_R1.fastq.gz",
                        "sequence_file_R2_fastq": "sample1_R2.fastq.gz"},
             sample2:{...} }
        """
        if not os.path.isfile(meta_f_path):
            log.error("Metadata file does not exist on %s", local_folder)
            stderr.print("[red] METADATA_LAB.xlsx do not exist in" + local_folder)
            return False
        sample_file_dict = {}
        metadata_ws, meta_header, header_row = self.read_metadata_file(meta_f_path)
        # TODO Include these columns in config
        index_sampleID = meta_header.index("Sample ID given for sequencing")
        index_layout = meta_header.index("Library Layout")
        index_fastq_r1 = meta_header.index("Sequence file R1 fastq")
        index_fastq_r2 = meta_header.index("Sequence file R2 fastq")
        counter = header_row
        for row in islice(metadata_ws.values, header_row, metadata_ws.max_row):
            counter += 1
            if row[index_sampleID] is not None:
                row_complete = True
                try:
                    s_name = str(row[index_sampleID]).strip()
                except ValueError as e:
                    stderr.print("[red]Unable to convert to string. ", e)
                    continue
                if s_name in sample_file_dict:
                    log_text = f"Found duplicated sample name: {s_name}. Skipped."
                    stderr.print(log_text)
                    self.include_warning(log_text, sample=s_name)
                    continue
                if row[index_layout] == "Paired" and row[index_fastq_r2] is None:
                    error_text = "Sample %s is paired-end, but no R2 given"
                    self.include_error(error_text % str(row[index_sampleID]), s_name)
                    row_complete = False
                if row[index_layout] == "Single" and row[index_fastq_r2] is not None:
                    error_text = "Sample %s is single-end, but R1&R2 given"
                    self.include_error(error_text % str(row[index_sampleID]), s_name)
                    row_complete = False
                if row_complete:
                    if row[index_fastq_r1] is not None:
                        sample_file_dict[s_name] = {}
                        # TODO: move these keys to configuration.json
                        sample_file_dict[s_name]["sequence_file_R1_fastq"] = row[
                            index_fastq_r1
                        ].strip()
                        if row[index_fastq_r2] is not None:
                            sample_file_dict[s_name]["sequence_file_R2_fastq"] = row[
                                index_fastq_r2
                            ].strip()
                    else:
                        log_text = "Fastq_R1 not defined in Metadata for sample %s"
                        stderr.print(f"[red]{str(log_text % s_name)}")
                        self.include_error(entry=str(log_text % s_name), sample=s_name)
            else:
                txt = f"Row {counter} in metadata skipped.No sequencing sample ID given"
                self.include_warning(entry=txt)
        # Remove duplicated files
        clean_sample_dict = self.remove_duplicated_values(sample_file_dict)
        return clean_sample_dict

    def get_metadata_file(self, remote_folder, local_folder):
        """Check if the metadata file exists

        Args:
            remote_folder (str): path to the folder in remote repository
            local_folder (str): path to the local folder

        Raises:
            FileNotFoundError: If missing metadata excel file or merging error.

        Returns:
            local_meta_file: Path to downloaded metadata file / merged metadata file.
        """
        remote_files_list = self.relecov_sftp.get_file_list(remote_folder)
        meta_files = [fi for fi in remote_files_list if fi.endswith(".xlsx")]

        def download_remote_metafile(target_meta_file):
            local_meta_file = os.path.join(
                local_folder, os.path.basename(target_meta_file)
            )
            try:
                self.relecov_sftp.get_from_sftp(target_meta_file, local_meta_file)
            except (IOError, PermissionError) as e:
                raise type(e)(f"[red]Unable to fetch metadata file {e}")
            log.info(
                "Obtained metadata file %s from %s",
                local_meta_file,
                remote_folder,
            )
            return local_meta_file

        if not meta_files:
            raise FileNotFoundError(f"Missing metadata file for {remote_folder}")
        os.makedirs(self.platform_storage_folder, exist_ok=True)
        if len(meta_files) > 1:
            # Merging multiple excel files into a single one
            log.warning(f"[yellow]Merging multiple metadata files in {remote_folder}")
            metadata_ws = self.metadata_processing.get("excel_sheet")
            header_flag = self.metadata_processing.get("header_flag")
            local_meta_list = []
            for remote_metafile in meta_files:
                local_meta_file = download_remote_metafile(remote_metafile)
                local_meta_list.append(local_meta_file)
            meta_df_list = []
            for loc_meta in local_meta_list:
                try:
                    loc_meta_df = self.excel_to_df(loc_meta, metadata_ws, header_flag)
                    meta_df_list.append(loc_meta_df)
                except (ParserError, EmptyDataError, MetadataError, KeyError) as e:
                    error_text = f"Could not process {os.path.basename(loc_meta)}: {e}"
                    self.include_error(error_text)
                os.remove(loc_meta)
            if meta_df_list:
                merged_df = meta_df_list[0]
            else:
                raise MetadataError("No single metadata file could be merged")
            for meta_df in meta_df_list[1:]:
                merged_df = self.merge_metadata(metadata_ws, merged_df, meta_df)
            folder_name = os.path.dirname(local_meta_file)
            excel_name = str(folder_name.split("/")[-1]) + "merged_metadata.xlsx"
            merged_excel_path = os.path.join(folder_name, excel_name)
            pd_writer = ExcelWriter(merged_excel_path, engine="xlsxwriter")
            for sheet in merged_df.keys():
                format_sheet = merged_df[sheet].astype(str)
                format_sheet.replace("nan", None, inplace=True)
                format_sheet.to_excel(pd_writer, sheet_name=sheet, index=False)
            pd_writer.close()
            local_meta_file = merged_excel_path
            return merged_excel_path
        else:
            target_meta_file = meta_files[0]
            local_meta_file = download_remote_metafile(target_meta_file)
        return local_meta_file

    def validate_remote_files(self, remote_folder, local_folder):
        """Check if the files in the remote folder are the ones defined in metadata file

        Args:
            remote_folder (str): Name of remote folder being validated
            local_folder (str): Name of folder where files are being downloaded

        Raises:
            FileNotFoundError: If none of the files in remote folder are valid

        Returns:
            sample_files_dict (dict): same structure as self.get_sample_fastq_file_names
            local_meta_file (str): location of downloaded metadata excel file
        """
        local_meta_file = self.get_metadata_file(remote_folder, local_folder)
        out_folder = os.path.dirname(local_meta_file)
        allowed_extensions = self.allowed_file_ext
        remote_files_list = [
            os.path.basename(file)
            for file in self.relecov_sftp.get_file_list(remote_folder)
        ]
        filtered_files_list = sorted(
            [fi for fi in remote_files_list if fi.endswith(tuple(allowed_extensions))]
        )
        sample_files_dict = self.get_sample_fastq_file_names(
            out_folder, local_meta_file
        )
        # Include the samples in the process log summary
        for sample in sample_files_dict.keys():
            self.include_new_key(sample=sample)
        metafiles_list = sorted(
            sum([list(fi.values()) for _, fi in sample_files_dict.items()], [])
        )
        if sorted(filtered_files_list) == sorted(metafiles_list):
            log.info("Files in %s match with metadata file", remote_folder)
        else:
            log_text = "Some files in %s do not match the ones described in metadata"
            log.warning(log_text % remote_folder)
            stderr.print(f"[gold1]{log_text % remote_folder}")
            set_list = set(metafiles_list)
            mismatch_files = [fi for fi in filtered_files_list if fi not in set_list]
            mismatch_rev = [fi for fi in set_list if fi not in filtered_files_list]

            if mismatch_files:
                error_text1 = "Files in folder missing in metadata %s"
                self.include_warning(error_text1 % str(mismatch_files))
            if mismatch_rev:
                error_text2 = "Files in metadata missing in folder %s"
                self.include_warning(error_text2 % str(mismatch_rev))
            # Try to check if the metadata filename lacks the proper extension
            log.info("Trying to match files without proper file extension")
            sample_files_dict = self.process_filedict(
                sample_files_dict, filtered_files_list
            )
        if not any(value for value in sample_files_dict.values()):
            raise FileNotFoundError(
                "No files from metadata found in %s" % remote_folder
            )
        stderr.print("[blue]Finished validating files based on metadata")
        return sample_files_dict, local_meta_file

    def delete_remote_files(self, remote_folder, files=None, skip_seqs=False):
        """Delete files from remote folder

        Args:
            remote_folder (str): path to folder in remote repository
            files (list(str), optional): list of target filenames in remote repository
            skip_seqs (bool, optional): Skip sequencing files based on extension
        """
        stderr.print(f"[blue]Deleting files in remote {remote_folder}...")
        if files is None:
            files_to_remove = self.relecov_sftp.get_file_list(remote_folder)
        else:
            files_to_remove = files
        for file in files_to_remove:
            if skip_seqs is True:
                if file.endswith(tuple(self.allowed_file_ext)):
                    continue
            try:
                self.relecov_sftp.remove_file(
                    os.path.join(remote_folder, os.path.basename(file))
                )
                log.info("%s Deleted from remote server", file)
            except (IOError, PermissionError) as e:
                log.error(f"Could not delete remote file {file}: {e}")
                stderr.print(f"Could not delete remote file {file}. Error: {e}")
        return

    def rename_remote_folder(self, remote_folder):
        if "tmp_processing" in remote_folder:
            new_name = remote_folder.replace("tmp_processing", "invalid_samples")
            if new_name == remote_folder:
                log.warning("Remote folder %s was already renamed", remote_folder)
                return
            try:
                self.relecov_sftp.rename_file(remote_folder, new_name)
                if self.finished_folders.get(remote_folder):
                    self.finished_folders[new_name] = self.finished_folders.pop(
                        remote_folder
                    )
                log.info("Successfully renamed %s to %s" % (remote_folder, new_name))
            except (OSError, PermissionError) as e:
                log_text = f"Could not rename remote {remote_folder}. Error: {e}"
                log.error(log_text)
        else:
            log.warning(
                "No `tmp_processing` pattern in %s, not renamed" % remote_folder
            )
        return

    def clean_remote_folder(self, remote_folder):
        """Delete a folder from remote sftp, check if it is empty or not first.

        Args:
            remote_folder (str): path to folder in remote repository
        """

        def remove_client_dir(remote_folder):
            # Never remove a folder in the top level
            if len(remote_folder.replace("./", "").split("/")) >= 2:
                log.info("Trying to remove %s", remote_folder)
                try:
                    self.relecov_sftp.remove_dir(remote_folder)
                    log.info("Successfully removed %s", remote_folder)
                except (OSError, PermissionError) as e:
                    log_text = f"Could not delete remote {remote_folder}. Error: {e}"
                    log.error(log_text)
                    stderr.print(log_text)
            else:
                log.info("%s is a top-level folder. Not removed", remote_folder)

        remote_folder_files = self.relecov_sftp.get_file_list(remote_folder)
        if remote_folder_files:
            self.rename_remote_folder(remote_folder)
            log_text = f"Remote folder {remote_folder} not empty. Not removed"
            log.warning(log_text)
        else:
            remove_client_dir(remote_folder)
        return

    def move_processing_fastqs(self, folders_with_metadata):
        """Gather all the files from any subfolder into a processing folder

        Args:
            folders_with_metadata (dict(str:list)): Dictionary updated from merge_md5sums()

        Returns:
            folders_with_metadata (dict(str:list)): Same dict updated with files successfully moved
        """
        log.info("Moving remote files to each temporal processing folder")
        stderr.print("[blue]Moving remote files to each temporal processing folder")
        for folder, files in folders_with_metadata.items():
            self.current_folder = folder.split("/")[0]
            successful_files = []
            for file in files:
                if not file.endswith(tuple(self.allowed_file_ext)):
                    continue
                file_dest = os.path.join(folder, os.path.basename(file))
                try:
                    # Paramiko.SSHClient.sftp_open does not have a method to copy files
                    self.relecov_sftp.rename_file(file, file_dest)
                    successful_files.append(file_dest)
                except OSError:
                    if file in folders_with_metadata[folder]:
                        error_text = "File named %s already in %s. Skipped"
                        self.include_warning(error_text % (file, self.current_folder))
                    else:
                        error_text = "Error while moving file %s"
                        self.include_error(error_text % file)
            folders_with_metadata[folder] = successful_files
        return folders_with_metadata

    def merge_md5sums(self, folders_with_metadata):
        """Download the md5sums for each folder, merge them into a single one,
        upload them to the remote processing folder.

        Args:
            folders_with_metadata (dict(str:list)): Dictionary with remote folders
            and their files. All subfolder filenames are merged into a single key.

        Raises:
            FileNotFoundError: If no md5sum file is found in the folder

        Returns:
            folders_with_metadata: Same dict updated with the merged md5sum file
        """
        output_location = self.platform_storage_folder

        # TODO: Include this function in relecov_tools.utils
        def md5_merger(md5_filelist, avoid_chars=None):
            """Merge all md5 files from a given list into a single multi-line md5sum"""
            md5dict_list = []
            for md5sum in md5_filelist:
                hash_dict = relecov_tools.utils.read_md5_checksum(md5sum, avoid_chars)
                if hash_dict:
                    md5dict_list.append(hash_dict)
            # Sort hashes and files back to the original order.
            merged_md5 = {
                md5: file for mdict in md5dict_list for file, md5 in mdict.items()
            }
            return merged_md5

        def md5_handler(md5sumlist, output_location):
            """Download all the remote md5sum files in a list, merge them
            into a single md5checksum and upload it back to sftp"""
            downloaded_md5files = []
            for md5sum in md5sumlist:
                md5_name = "_".join([token_hex(nbytes=12), "md5_temp.md5"])
                fetched_md5 = os.path.join(output_location, md5_name)
                if self.relecov_sftp.get_from_sftp(
                    file=md5sum, destination=fetched_md5
                ):
                    downloaded_md5files.append(fetched_md5)
            merged_md5 = md5_merger(downloaded_md5files, self.avoidable_characters)
            if merged_md5:
                merged_name = "_".join([folder.split("/")[0], "md5sum.md5"])
                merged_md5_path = os.path.join(output_location, merged_name)
                with open(merged_md5_path, "w") as md5out:
                    write_md5 = csv_writer(md5out, delimiter="\t")
                    write_md5.writerows(merged_md5.items())
                md5_dest = os.path.join(folder, os.path.basename(merged_md5_path))
                self.relecov_sftp.upload_file(merged_md5_path, md5_dest)
                # Remove local files once merged and uploaded
                os.remove(merged_md5_path)
                [os.remove(md5_file) for md5_file in downloaded_md5files]
                return md5_dest
            else:
                error_text = "No md5sum could be processed in remote folder"
                raise FileNotFoundError(error_text)

        for folder, files in folders_with_metadata.items():
            self.current_folder = folder.split("/")[0]
            log.info("Merging md5sum files from %s...", self.current_folder)
            stderr.print(f"[blue]Merging md5sum files from {self.current_folder}...")
            md5flags = [".md5", "md5sum", "md5checksum"]
            md5sumlist = [fi for fi in files if any(flag in fi for flag in md5flags)]
            if not md5sumlist:
                error_text = "No md5sum could be found in remote folder %s"
                stderr.print(f"[yellow]{error_text % folder}")
                self.include_warning(error_text % folder)
                continue
            folders_with_metadata[folder] = [fi for fi in files if fi not in md5sumlist]
            try:
                uploaded_md5 = md5_handler(md5sumlist, output_location)
            except (FileNotFoundError, OSError, PermissionError, CsvError) as e:
                error_text = "Could not merge md5files for %s. Reason: %s"
                stderr.print(f"[yellow]{error_text % (self.current_folder, str(e))}")
                self.include_warning(error_text % (self.current_folder, str(e)))
                continue
            if uploaded_md5:
                folders_with_metadata[folder].append(uploaded_md5)

        return folders_with_metadata

    def merge_metadata(self, meta_sheet=None, *metadata_tables):
        """Merge a variable number of metadata dataframes to the first one. Merge them
        only into a certain sheet from a multi-sheet excel file if sheetname is given.

        Args:
            meta_sheet (str): Name of the sheet containing metadata in excel file
            *metadata_tables (list(pandas.DataFrame)): Dataframes to be merged

        Returns:
            merged_df (pandas.DataFrame): A merged dataframe from the given tables
        """
        for idx, table in enumerate(metadata_tables):
            if idx == 0:
                merged_df = table
                continue
            if meta_sheet:
                merged_df[meta_sheet] = concat(
                    [merged_df[meta_sheet], table[meta_sheet]], ignore_index=True
                )
            else:
                merged_df = concat([merged_df, table], ignore_index=True)
        return merged_df

    def excel_to_df(self, excel_file, metadata_sheet, header_flag):
        """Read an excel file, return a dict with a dataframe for each sheet in it.
        Process the given sheet with metadata, removing all rows until header is found

        Args:
            excel_file (str): Path to the local excel file with metadata
            metadata_sheet (str): Name of the sheet containing metadata in excel file
            header_flag (str): Name of one of the columns from the metadata header

        Raises:
            MetadataError: If no header could be found matching header flag

        Returns:
            excel_df (dict(str:pandas.DataFrame)): Dict {name_of_excel_sheet:DataFrame}
            containing all sheets in the excel file as pandas dataframes.
        """
        # Get every sheet from the first excel file
        excel_df = read_excel(excel_file, dtype=str, sheet_name=None)
        meta_df = excel_df[metadata_sheet]
        if header_flag in meta_df.columns:
            return excel_df
        header_row = None
        for idx in range(len(meta_df)):
            if any(meta_df.loc[idx, x] == header_flag for x in meta_df.columns):
                header_row = idx
        meta_df.columns = meta_df.iloc[header_row]
        excel_df[metadata_sheet] = meta_df.drop(meta_df.index[: (header_row + 1)])
        excel_df[metadata_sheet] = excel_df[metadata_sheet].reset_index(drop=True)
        return excel_df

    def merge_subfolders(self, target_folders):
        """For each first-level folder in the sftp, merge all the subfolders within
        it in a single one called '*_tmp_processing' by moving all the fastq files from
        them. Merge the metadata excel and md5 files from each subfolder too.

        Args:
            target_folders (dict(str:list)): Dictionary with folders and their files

        Returns:
            clean_target_folders (dict(str:list)): Dict with '*_tmp_processing' folders
            and their content. All subfolder filenames are merged into a single key.
        """
        metadata_ws = self.metadata_processing.get("excel_sheet")
        header_flag = self.metadata_processing.get("header_flag")
        output_location = self.platform_storage_folder
        date_and_time = datetime.today().strftime("%Y%m%d%-H%M%S")
        exts = self.allowed_file_ext

        def upload_merged_df(merged_excel_path, last_main_folder, merged_df):
            """Upload metadata dataframe merged from all subfolders back to sftp"""
            self.relecov_sftp.make_dir(last_main_folder)
            pd_writer = ExcelWriter(merged_excel_path, engine="xlsxwriter")
            for sheet in merged_df.keys():
                format_sheet = merged_df[sheet].astype(str)
                format_sheet.replace("nan", None, inplace=True)
                format_sheet.to_excel(pd_writer, sheet_name=sheet, index=False)
            pd_writer.close()
            dest = os.path.join(last_main_folder, os.path.basename(merged_excel_path))
            self.relecov_sftp.upload_file(merged_excel_path, dest)
            os.remove(merged_excel_path)
            return

        def pre_validate_folder(folder, folder_files):
            """Check if remote folder has sequencing files and a valid metadata file"""
            if not any(file.endswith(tuple(exts)) for file in folder_files):
                error_text = "Remote folder %s skipped. No sequencing files found."
                self.include_error(error_text % folder)
                return
            try:
                downloaded_metadata = self.get_metadata_file(folder, output_location)
            except (FileNotFoundError, OSError, PermissionError, MetadataError) as err:
                error_text = "Remote folder %s skipped. Reason: %s"
                self.include_error(error_text % (folder, err))
                return
            try:
                self.read_metadata_file(downloaded_metadata, return_data=False)
            except (MetadataError, KeyError) as excel_error:
                error_text = f"Folder {self.current_folder} skipped: %s"
                os.remove(downloaded_metadata)
                self.include_error(error_text % excel_error)
                return
            return downloaded_metadata

        folders_with_metadata = {}
        processed_folders = []
        merged_df = merged_excel_path = last_main_folder = excel_name = None
        log.info("Setting %s remote folders...", str(len(target_folders.keys())))
        stderr.print(f"[blue]Setting {len(target_folders.keys())} remote folders...")
        for folder in sorted(target_folders.keys()):
            if "invalid_samples" in folder:
                log.warning("Skipped invalid_samples folder %s", folder)
                continue
            self.current_folder = folder
            # Include the folder in the final process log summary
            self.include_new_key()
            downloaded_metadata = pre_validate_folder(folder, target_folders[folder])
            if not downloaded_metadata:
                continue
            # Create a temporal name to avoid duplicated filenames
            meta_filename = "_".join([folder.split("/")[-1], "metadata_temp.xlsx"])
            local_meta = os.path.join(output_location, meta_filename)
            os.rename(downloaded_metadata, local_meta)

            # Taking the main folder for each lab as reference for merge and logs
            main_folder = folder.split("/")[0]
            self.current_folder = main_folder
            temporal_foldername = "_".join([date_and_time, "tmp_processing"])
            temp_folder = os.path.join(main_folder, temporal_foldername)
            # Get every file except the excel ones as they are going to be merged
            filelist = [fi for fi in target_folders[folder] if not fi.endswith(".xlsx")]
            if not folders_with_metadata.get(temp_folder):
                log_text = "Trying to merge metadata from %s in %s"
                log.info(log_text % (main_folder, temp_folder))
                stderr.print(f"[blue]{log_text % (main_folder, temp_folder)}")
                if merged_df:
                    # Write the previous merged metadata df before overriding it
                    try:
                        upload_merged_df(merged_excel_path, last_main_folder, merged_df)
                        folders_with_metadata[last_main_folder].append(excel_name)
                    except OSError:
                        error_text = "Error uploading merged metadata back to sftp: %s"
                        self.include_error(error_text % last_main_folder)
                        del folders_with_metadata[last_main_folder]
                try:
                    merged_df = self.excel_to_df(local_meta, metadata_ws, header_flag)
                except (ParserError, EmptyDataError, MetadataError, KeyError) as e:
                    meta_name = os.path.basename(downloaded_metadata)
                    error_text = "%s skipped. Error while processing excel %s: %s"
                    self.include_error(error_text % (main_folder, meta_name, str(e)))
                    os.remove(local_meta)
                    continue
                folders_with_metadata[temp_folder] = []
                folders_with_metadata[temp_folder].extend(filelist)
                # rename metadata file to avoid filename duplications
                excel_name = "_".join([folder.split("/")[0], "merged_metadata.xlsx"])
                merged_excel_path = os.path.join(output_location, excel_name)
                os.rename(local_meta, merged_excel_path)
                # Keep a track of the main_folder for next iteration
                last_main_folder = temp_folder
            else:
                # If temp_folder has subfolders in it, merge everything
                folders_with_metadata[temp_folder].extend(filelist)
                new_df = self.excel_to_df(local_meta, metadata_ws, header_flag)
                merged_df = self.merge_metadata(metadata_ws, merged_df, new_df)
                os.remove(local_meta)
            processed_folders.append(folder)
        # End of loop

        # Write last dataframe to file once loop is finished
        if folders_with_metadata.get(last_main_folder):
            if excel_name not in folders_with_metadata[last_main_folder]:
                upload_merged_df(merged_excel_path, last_main_folder, merged_df)
                folders_with_metadata[last_main_folder].append(excel_name)

        # Merge md5files and upload them to tmp_processing folder
        merged_md5_folders = self.merge_md5sums(folders_with_metadata)
        # Move all the files from each subfolder into its tmp_processing folder
        clean_target_folders = self.move_processing_fastqs(merged_md5_folders)
        log_text = "Remote folders merged into %s folders. Proceed with processing"
        log.info(log_text % len(clean_target_folders.keys()))
        stderr.print(f"[green]{log_text % len(clean_target_folders.keys())}")
        return clean_target_folders, processed_folders

    def select_target_folders(self):
        """Find the selected folders in remote if given, else select every folder

        Returns:
            folders_to_process (dict(str:list)): Dictionary with folders and their files
        """
        root_directory_list = self.relecov_sftp.list_remote_folders(".", recursive=True)
        clean_root_list = [folder.replace("./", "") for folder in root_directory_list]
        if not root_directory_list:
            log.error("Error while listing folders in remote. Aborting")
            sys.exit(1)
        if self.target_folders is None:
            target_folders = clean_root_list
        elif self.target_folders[0] == "ALL":
            log.info("Showing folders from remote SFTP for user selection")
            target_folders = relecov_tools.utils.prompt_checkbox(
                msg="Select the folders that will be targeted",
                choices=sorted(clean_root_list),
            )
        else:
            target_folders = [tf for tf in self.target_folders if tf in clean_root_list]
        if not target_folders:
            log.error("No remote folders matching selection %s", self.target_folders)
            stderr.print("Found no remote folders matching selection")
            stderr.print(f"List of remote folders: {str(clean_root_list)}")
            sys.exit(1)
        folders_to_process = {}
        for targeted_folder in target_folders:
            try:
                full_folders = self.relecov_sftp.list_remote_folders(
                    targeted_folder, recursive=True
                )
            except (FileNotFoundError, OSError) as e:
                log.error(f"Error during sftp listing. {targeted_folder} skipped:", e)
                continue
            for folder in full_folders:
                list_files = self.relecov_sftp.get_file_list(folder)
                if list_files:
                    folders_to_process[folder] = list_files
                else:
                    log.info("%s is empty", folder)
                    continue
        if len(folders_to_process) == 0:
            log.info("Exiting process, folders were empty.")
            log.error("There are no files in the selected folders.")
            self.relecov_sftp.close_connection()
            sys.exit(0)
        return folders_to_process

    def compress_and_update(self, fetched_files, files_to_compress, local_folder):
        """compress the given list of files_to_compress and update files_list

        Args:
            fetched_files (list(str)): list of all downloaded files
            files_to_compress (list(str)): list of files that are uncompressed

        Returns:
            fetched_files(list(str)): files list including the new compressed files
        """
        compressed_files = list()
        for file in files_to_compress:
            f_path = os.path.join(local_folder, file)
            compressed = relecov_tools.utils.compress_file(f_path)
            if not compressed:
                error_text = "Could not compress file %s, file not found" % str(file)
                self.include_error(error_text, f_path)
                continue
            # Remove file after compression is completed
            compressed_files.append(file)
            try:
                os.remove(f_path)
            except (FileNotFoundError, PermissionError) as e:
                log.warning(f"Could not delete file: {e}")
        fetched_files = [
            (fi + ".gz" if fi in compressed_files else fi) for fi in fetched_files
        ]
        return fetched_files

    def process_filedict(
        self, valid_filedict, clean_fetchlist, corrupted=[], md5miss=[]
    ):
        """Process the dictionary from validate_remote_files() to update filenames
        and remove samples that failed any validation process.

        Args:
            valid_filedict (dict{str:str}): same structure as validate_remote_files()
            clean_fetchlist (list(str)): List of files that passed validation process
            processed (bool): Indicates if filedict has been processed previously

        Returns:
            processed(dict{str:str}): Updated valid_filedict
        """
        processed_dict = {}
        error_text = "corrupted or md5 mismatch for %s"
        warning_text = "File %s not found in md5sum. Creating hash"
        for sample, vals in valid_filedict.items():
            processed_dict[sample] = {}
            for key, val in vals.items():
                processed_dict[sample][key] = None
                if val in corrupted:
                    self.include_error(error_text % val, sample=sample)
                if val in md5miss:
                    self.include_warning(warning_text % val, sample=sample)
                for file in clean_fetchlist:
                    if val in file:
                        processed_dict[sample][key] = file
            # remove sample if it has missing files
            if not all(x in clean_fetchlist for x in processed_dict[sample].values()):
                if not corrupted:
                    error_text = "Sample %s skipped: missing files in sftp"
                    self.include_error(str(error_text % sample), sample=sample)
                log.error(str(error_text % sample))
                del processed_dict[sample]
        return processed_dict

    def download(self, target_folders):
        """Manages all the different functions to download files, verify their
        integrity and create initial json with filepaths and md5 hashes

        Args:
            target_folders (dict): dictionary
        """
        log.info("Initiating download process")
        main_folder = self.platform_storage_folder
        try:
            os.makedirs(main_folder, exist_ok=True)
        except OSError as e:
            log.error("You do not have permissions to create folder %s", e)
            sys.exit(1)
        folders_to_download = target_folders
        for folder in folders_to_download.keys():
            self.current_folder = folder.split("/")[0]
            # Close previously open connection to avoid timeouts
            try:
                self.relecov_sftp.close_connection()
            except paramiko.SSHException:
                pass
            # Check if the connection has been closed due to time limit
            self.relecov_sftp.open_connection()
            log.info("Processing folder %s", folder)
            stderr.print("[blue]Processing folder " + folder)
            # Validate that the files are the ones described in metadata.

            local_folder = self.create_local_folder(folder)
            try:
                valid_filedict, meta_file = self.validate_remote_files(
                    folder, local_folder
                )
            except (FileNotFoundError, IOError, PermissionError, MetadataError) as fail:
                log.error(fail)
                stderr.print(f"[red]{fail}, skipped")
                self.include_error(fail)
                continue
            # Get the files in each folder
            files_to_download = [
                fi for vals in valid_filedict.values() for fi in vals.values()
            ]
            fetched_files = self.get_remote_folder_files(
                folder, local_folder, files_to_download
            )
            if not fetched_files:
                error_text = "No files could be downloaded in folder %s" % str(folder)
                stderr.print(f"{error_text}")
                self.include_error(error_text)
                continue
            log.info("Finished download for folder: %s", folder)
            stderr.print(f"Finished download for folder {folder}")
            remote_md5sum = self.find_remote_md5sum(folder)
            if remote_md5sum:
                # Get the md5checksum to validate integrity of files after download
                fetched_md5 = os.path.join(
                    local_folder, os.path.basename(remote_md5sum)
                )
                self.relecov_sftp.get_from_sftp(
                    file=remote_md5sum, destination=fetched_md5
                )
                successful_files, corrupted = self.verify_md5_checksum(
                    local_folder, fetched_files, fetched_md5
                )
                # try to download the files again to discard errors during download
                if corrupted:
                    stderr.print("[gold1]Found md5 mismatches, downloading again...")
                    self.get_remote_folder_files(folder, local_folder, corrupted)
                    saved_files, corrupted = self.verify_md5_checksum(
                        local_folder, corrupted, fetched_md5
                    )
                    if saved_files:
                        successful_files.extend(saved_files)
                    if corrupted:
                        corr_fold = os.path.join(local_folder, "corrupted")
                        os.mkdir(corr_fold)
                        error_text = "Found corrupted files: %s. Moved to: %s"
                        stderr.print(f"[red]{error_text % (str(corrupted), corr_fold)}")
                        self.include_warning(error_text % (str(corrupted), corr_fold))
                        for corr_file in corrupted:
                            path = os.path.join(local_folder, corr_file)
                            try:
                                os.rename(path, os.path.join(corr_fold, corr_file))
                            except (FileNotFoundError, PermissionError, OSError) as e:
                                error_text = (
                                    "Could not move corrupted file %s to %s: %s"
                                )
                                log.error(error_text % (path, corr_fold, e))
                                stderr.print(
                                    f"[red]{error_text % (path, corr_fold, e)}"
                                )
                        if self.abort_if_md5_mismatch:
                            error_text = "Stop processing %s due to corrupted files."
                            stderr.print(f"[red]{error_text % folder}")
                            self.include_error(error_text % "folder")
                            relecov_tools.utils.delete_local_folder(local_folder)
                            continue
                hash_dict = relecov_tools.utils.read_md5_checksum(
                    fetched_md5, self.avoidable_characters
                )
                log.info("Finished md5 check for folder: %s", folder)
                stderr.print(f"[blue]Finished md5 verification for folder {folder}")
            else:
                corrupted = []
                error_text = "No single md5sum file could be found in %s" % folder
                stderr.print(f"[red]{error_text}")
                self.include_warning(error_text)

            seqs_fetchlist = [
                fi for fi in fetched_files if fi.endswith(tuple(self.allowed_file_ext))
            ]
            seqs_fetchlist = [fi for fi in seqs_fetchlist if fi not in corrupted]
            # Checking for uncompressed files
            files_to_compress = [
                fi
                for fi in seqs_fetchlist
                if not fi.endswith(".gz") and not fi.endswith(".bam")
            ]
            if files_to_compress:
                comp_files = str(len(files_to_compress))
                log.info("Found %s uncompressed files, compressing...", comp_files)
                stderr.print(f"Found {comp_files} uncompressed files, compressing...")
                clean_fetchlist = self.compress_and_update(
                    seqs_fetchlist, files_to_compress, local_folder
                )
            else:
                clean_fetchlist = seqs_fetchlist
            clean_pathlist = [os.path.join(local_folder, fi) for fi in clean_fetchlist]
            not_md5sum = []
            if remote_md5sum:
                # Get hashes from provided md5sum, create them for those not provided
                files_md5_dict = {}
                for path in clean_pathlist:
                    f_name = os.path.basename(path)
                    if f_name in successful_files:
                        files_md5_dict[f_name] = hash_dict[f_name]
                    elif f_name in corrupted:
                        clean_fetchlist.remove(f_name)
                    else:
                        if not str(f_name).rstrip(".gz") in files_to_compress:
                            error_text = "File %s not found in md5sum. Creating hash"
                            log.warning(error_text % f_name)
                            not_md5sum.append(f_name)
                        else:
                            log.info("File %s was compressed, creating md5hash", f_name)
                        files_md5_dict[f_name] = relecov_tools.utils.calculate_md5(path)
            else:
                md5_hashes = [
                    relecov_tools.utils.calculate_md5(path) for path in clean_pathlist
                ]
                files_md5_dict = dict(zip(clean_fetchlist, md5_hashes))
            for file in files_md5_dict.keys():
                full_f_path = os.path.join(local_folder, file)
                if not relecov_tools.utils.check_gzip_integrity(full_f_path):
                    corrupted.append(file)
            files_md5_dict = {
                x: y for x, y in files_md5_dict.items() if x not in corrupted
            }
            processed_filedict = self.process_filedict(
                valid_filedict, clean_fetchlist, corrupted=corrupted, md5miss=not_md5sum
            )
            self.create_files_with_metadata_info(
                local_folder, processed_filedict, files_md5_dict, meta_file
            )
            if self.logsum.logs.get(self.current_folder):
                self.logsum.logs[self.current_folder].update({"path": local_folder})
                try:
                    folder_basename = os.path.basename(local_folder.rstrip("/"))
                    log_name = folder_basename + "_download_log_summary.json"
                    self.logsum.create_error_summary(
                        filepath=os.path.join(local_folder, log_name),
                        logs={
                            self.current_folder: self.logsum.logs[self.current_folder]
                        },
                    )
                except Exception as e:
                    log.error("Could not create logsum for %s: %s" % (folder, str(e)))
            stderr.print(f"[green]Finished processing {folder}")
            self.finished_folders[folder] = list(files_md5_dict.keys())
        return

    def include_new_key(self, sample=None):
        self.logsum.feed_key(key=self.current_folder, sample=sample)
        return

    def include_error(self, entry, sample=None):
        self.logsum.add_error(key=self.current_folder, entry=entry, sample=sample)
        return

    def include_warning(self, entry, sample=None):
        self.logsum.add_warning(key=self.current_folder, entry=entry, sample=sample)
        return

    def execute_process(self):
        """Executes different processes depending on the download_option"""
        if not self.relecov_sftp.open_connection():
            log.error("Unable to establish connection towards sftp server")
            stderr.print("[red]Unable to establish sftp connection")
            sys.exit(1)
        target_folders = self.select_target_folders()
        if self.download_option == "delete_only":
            log.info("Initiating delete_only process")
            processed_folders = target_folders.keys()
            for folder in processed_folders:
                self.current_folder = folder
                self.delete_remote_files(folder)
                self.clean_remote_folder(folder)
                stderr.print(f"Delete process finished in {folder}")
        else:
            target_folders, processed_folders = self.merge_subfolders(target_folders)
            self.download(target_folders)

        self.relecov_sftp.close_connection()
        stderr.print(f"Processed {len(processed_folders)} folders: {processed_folders}")
        if self.logsum.logs:
            log.info("Printing process summary to %s", self.platform_storage_folder)
            self.logsum.create_error_summary(called_module="download")
        else:
            log.info("Process log summary was empty. Not generated.")
        # If download_option is "download_clean", remove
        # sftp folder content after download is finished
        if self.download_option == "download_clean":
            for folder in processed_folders:
                self.delete_remote_files(folder, skip_seqs=True)
                self.clean_remote_folder(folder)
            folders_to_clean = copy.deepcopy(self.finished_folders)
            for folder, downloaded_files in folders_to_clean.items():
                self.delete_remote_files(folder, files=downloaded_files)
                self.delete_remote_files(folder, skip_seqs=True)
                self.clean_remote_folder(folder)
                stderr.print(f"Delete process finished in remote {folder}")
        stderr.print("Finished execution")
        return
